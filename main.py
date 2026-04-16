import streamlit as st
import pandas as pd
import requests
import re
import unicodedata
import os
import ast
import io
import json
import zipfile
import time
import threading
from datetime import datetime as dt_datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook

try:
    import pdfplumber
    _HAS_PDFPLUMBER = True
except ImportError:
    _HAS_PDFPLUMBER = False

st.set_page_config(page_title="Banana Import Club", layout="wide", page_icon="🍌")

# --- FONCTIONS DE NETTOYAGE ---
def to_clean_str(val):
    if pd.isna(val) or val == "":
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        return s[:-2]
    return s

def norm_piv(s): 
    if not s: return ""
    s = str(s).upper().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9]', '', s)
    return s[:20]

def clean_label_tva(label, code, do_fusion=True):
    if do_fusion and str(code).startswith(('2', '6', '7')):
        s = str(label)
        # 1. Supprimer les taux de TVA (20%, 5,5 %, 10.0%, etc.) avec les signes autour
        s = re.sub(r'[\s\-/\(]*\d+[.,\s]?\d*\s?%[\s\-/\)]*', ' ', s, flags=re.IGNORECASE)
        # 2. Supprimer les mentions TVA et toutes leurs variantes, y compris
        #    tronquées (EXO, INTR, AUTOLIQ...) grâce à des préfixes minimaux.
        #    Le \w* après le préfixe attrape les suffixes quelconques.
        s = re.sub(
            r'[\s\-/\(]*\b('
            r'T[\.\s]?V[\.\s]?A\.?'                   # TVA, T.V.A., T V A
            r'|EXON\w*'                              # EXONERE, EXONEREE, EXONERATION...
            r'|EXO\w*'                               # EXO, EXOTVA, EXO TVA...
            r'|INTRA\w*'                             # INTRA, INTRACOM, INTRACOMMUNAUTAIRE...
            r'|I[\.\s]?C\.?\w*'                        # IC, I.C., I C, I.C
            r'|AUTO[\-\s.]?LIQ\w*'                   # AUTOLIQ, AUTOLIQUIDATION, AUTO-LIQ...
            r'|REVERSE[\s\-]?CHARGE\w*'              # REVERSE CHARGE
            r'|NON[\s\-]?SOUMIS\w*'                  # NON SOUMIS, NON-SOUMISE
            r'|HORS[\s\-]?TAXE\w*'                   # HORS TAXE, HORS-TAXE
            r'|IMPORT\w*'                             # IMPORT, IMPORTATION
            r'|EXPORT\w*'                             # EXPORT, EXPORTATION
            r'|FRANCE'                                # FRANCE
            r'|FR'                                    # FR
            r'|UE'                                    # UE
            r'|CEE'                                   # CEE
            r'|HT'                                    # HT
            r'|TTC'                                   # TTC
            r')\b[\s\-/\)]*', ' ', s, flags=re.IGNORECASE)
        # 3. Normaliser la ponctuation résiduelle (. , ; :) en espaces
        s = re.sub(r'[.,;:]+', ' ', s)
        # 4. Supprimer les mots résiduels type EX... en fin de chaîne
        s = re.sub(r'[\s\-/]*\bEX\w*\s*$', '', s, flags=re.IGNORECASE)
        # 5. Nettoyer les signes orphelins en début/fin et les espaces multiples
        s = re.sub(r'^[\s\-/\(\)]+|[\s\-/\(\)]+$', '', s)
        return " ".join(s.split()).upper()
    return str(label).strip().upper()

# --- 3. MOTEUR API ---
def get_correct_id(item, endpoint):
    if endpoint in ["purchase-classifications", "sale-classifications"]:
        return item.get('classificationid') or item.get('id')
    if "affectations" in endpoint:
        return item.get('affectationid') or item.get('id')
    if endpoint == "accounts":
        return item.get('accountid') or item.get('id')
    return item.get('id')

def fetch_evoliz_data(endpoint, headers, company_id=None):
    results = {}
    # URL avec prefixe en primaire, sans prefixe en fallback (cas company_users avec cid invalide)
    urls_to_try = []
    if company_id:
        urls_to_try.append(f"https://www.evoliz.io/api/v1/companies/{company_id}/{endpoint}")
    urls_to_try.append(f"https://www.evoliz.io/api/v1/{endpoint}")

    # Tester chaque URL jusqu'a en trouver une qui repond 200
    url = None
    for _u in urls_to_try:
        try:
            _r_test = requests.get(_u, headers=headers, params={"per_page": 1, "page": 1}, timeout=10)
            if _r_test.status_code == 200:
                url = _u; break
        except Exception:
            continue
    if url is None:
        return results  # aucune URL accessible

    params = {"per_page": 100, "page": 1}
    is_account = endpoint == "accounts"
    try:
        while True:
            r = requests.get(url, headers=headers, params=params, timeout=15)
            if r.status_code != 200:
                break
            data = r.json()
            for item in data.get('data', []):
                item_id = get_correct_id(item, endpoint)
                code = to_clean_str(item.get('code'))
                label = str(item.get('label') or item.get('category_name') or code).strip().upper()
                # accountid : dans account imbriqué (account.accountid)
                acc_id = item.get('accountid')
                if acc_id is None and isinstance(item.get('account'), dict):
                    acc_id = item['account'].get('accountid')
                # vataccount : le compte TVA est dans vataccount.accountid (pas vataccountid !)
                vat_id = None
                if isinstance(item.get('vataccount'), dict):
                    vat_id = item['vataccount'].get('accountid')
                entry = {"id": item_id, "label": label, "code": code, "acc_id": acc_id, "vat_id": vat_id, "_raw_keys": list(item.keys())}
                pivot_code = norm_piv(code) if code else None
                pivot_label = norm_piv(label)
                if is_account:
                    # Comptes : indexer uniquement par code (pas de doublon)
                    if pivot_code:
                        results[pivot_code] = entry
                else:
                    # Flux : indexer par code ET label pour permettre le matching
                    if pivot_code:
                        results[pivot_code] = entry
                    if pivot_label and pivot_label != pivot_code:
                        results[pivot_label] = entry
            if params['page'] >= data.get('meta', {}).get('last_page', 1):
                break
            params['page'] += 1
    except Exception as e:
        st.warning(f"Erreur API ({endpoint}): {e}")
    return results


# =========================================================
# EVOLIZ SYNC V10.5
# =========================================================

FLUX_ENDPOINTS = {
    "ACHAT": "purchase-classifications",
    "VENTE": "sale-classifications",
    "ENTRÉE BQ": "sale-affectations",
    "SORTIE BQ": "purchase-affectations",
}

def count_unique(data_dict):
    """Compte les éléments uniques par id dans un dict potentiellement double-indexé"""
    return len({d['id'] for d in data_dict.values()})

def has_movement(row, flux_cols):
    for c in flux_cols:
        val = pd.to_numeric(row[c], errors='coerce')
        if pd.notna(val) and abs(val) > 0:
            return True
    return False

def has_movement_debit_credit(row, flux_cols):
    """Vérifie le mouvement uniquement sur les colonnes Débit/Crédit (ignore Solde)."""
    for c in flux_cols:
        if "SOLDE" in str(c).upper():
            continue
        val = pd.to_numeric(row[c], errors='coerce')
        if pd.notna(val) and abs(val) > 0:
            return True
    return False

def inject_account(code, label, headers):
    try:
        r = requests.post("https://www.evoliz.io/api/v1/accounts",
                          headers=headers, json={"code": code, "label": label}, timeout=15)
        if r.status_code in (200, 201):
            return True, r.json()
        # Si le compte existe déjà, considérer comme succès
        if r.status_code == 400 and "already been taken" in r.text:
            return True, f"Déjà existant (code {code})"
        return False, f"HTTP {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, str(e)

def patch_evoliz_item(category, item_id, payload, headers, company_id=None):
    endpoint = ERAZ_ENDPOINTS[category]
    # Essai 1 : URL company-scoped
    # Essai 2 : URL sans prefixe (cas company_users avec cid != vrai companyid)
    urls_to_try = []
    if company_id:
        urls_to_try.append(f"https://www.evoliz.io/api/v1/companies/{company_id}/{endpoint}/{item_id}")
    urls_to_try.append(f"https://www.evoliz.io/api/v1/{endpoint}/{item_id}")
    last_err = None
    for url in urls_to_try:
        try:
            r = requests.patch(url, headers=headers, json=payload, timeout=15)
            if r.status_code in (200, 204):
                return True, r.json() if r.status_code == 200 else "OK"
            last_err = f"HTTP {r.status_code}: {r.text[:200]}"
            if r.status_code not in (403, 404):
                break  # erreur definitive
        except Exception as e:
            last_err = str(e)
    return False, last_err or "Erreur inconnue"

def inject_flux(flux_type, code, label, headers, vat_id=None, acc_id=None, vat_rate=None, company_id=None):
    endpoint = FLUX_ENDPOINTS[flux_type]
    # Le code d'une catégorie/affectation est son libellé (tronqué à 50 car pour l'API)
    payload = {"code": label[:50], "label": label}
    if acc_id and not (isinstance(acc_id, float) and pd.isna(acc_id)):
        payload["accountid"] = int(acc_id)
    if vat_id and not (isinstance(vat_id, float) and pd.isna(vat_id)):
        # L'API Evoliz attend le champ "vataccountid" pour lier un compte TVA
        payload["vataccountid"] = int(vat_id)
    if vat_rate is not None and flux_type == "ACHAT":
        payload["vat_rate"] = float(vat_rate)
    # Debug : log du payload envoyé
    _payload_debug = {k: v for k, v in payload.items() if k not in ('code', 'label')}
    # Fallback : recuperer le company_id depuis la session si non fourni
    if company_id is None:
        company_id = st.session_state.get('company_id_105')
    # URL company-scoped si possible (requis pour les tokens prescriber_users)
    if company_id:
        url = f"https://www.evoliz.io/api/v1/companies/{company_id}/{endpoint}"
    else:
        url = f"https://www.evoliz.io/api/v1/{endpoint}"
    try:
        r = requests.post(url, headers=headers, json=payload, timeout=15)
        if r.status_code in (200, 201):
            resp = r.json()
            resp['_sent'] = _payload_debug
            return True, resp
        # Fallback : si 403/404 en company-scoped, re-essayer sans le prefixe
        if r.status_code in (403, 404) and company_id:
            r2 = requests.post(f"https://www.evoliz.io/api/v1/{endpoint}",
                                headers=headers, json=payload, timeout=15)
            if r2.status_code in (200, 201):
                resp = r2.json()
                resp['_sent'] = _payload_debug
                return True, resp
        return False, f"HTTP {r.status_code} [sent:{_payload_debug}]: {r.text[:150]}"
    except Exception as e:
        return False, str(e)

ERAZ_ENDPOINTS = {
    "COMPTE": "accounts",
    "ACHAT": "purchase-classifications",
    "VENTE": "sale-classifications",
    "ENTRÉE BQ": "sale-affectations",
    "SORTIE BQ": "purchase-affectations",
}

def delete_evoliz_item(category, item_id, headers, company_id=None):
    endpoint = ERAZ_ENDPOINTS[category]
    # Pattern URL : primaire avec prefixe /companies/{cid}/, fallback sans prefixe
    urls = []
    if company_id:
        urls.append(f"https://www.evoliz.io/api/v1/companies/{company_id}/{endpoint}/{item_id}")
    urls.append(f"https://www.evoliz.io/api/v1/{endpoint}/{item_id}")

    # Etape 1 : tenter DELETE sur chaque URL
    last_status = None; last_body = ""
    for url in urls:
        try:
            r = requests.delete(url, headers=headers, timeout=15)
            last_status = r.status_code; last_body = r.text
            if r.status_code in (200, 204):
                return True, "Supprimé"
            if r.status_code not in (403, 404):
                break  # erreur non liee aux droits -> inutile de retenter
        except Exception as e:
            last_status = -1; last_body = str(e)

    # Etape 2 : DELETE a echoue -> tenter PATCH enabled=false pour "desactiver"
    # (sauf pour COMPTE qui n'a pas de champ enabled)
    if category != "COMPTE":
        for url in urls:
            try:
                r2 = requests.patch(url, headers=headers,
                                    json={"enabled": False}, timeout=15)
                if r2.status_code in (200, 204):
                    return True, "Désactivé (enabled=false)"
                if r2.status_code not in (403, 404):
                    break
            except Exception:
                pass

    return False, f"HTTP {last_status}: {last_body[:200]}"

st.title("🍌 Banana Import Club")
st.caption("Version **v2026.04.15-auth-v10** — dossier mono OK, import fichier OK meme sans nom.")

for key, default in [('nr_v62', pd.DataFrame()), ('audit_matrix_105', pd.DataFrame()),
                         ('rejets_105', pd.DataFrame()), ('prot_105', set()), ('sync_log', []),
                         ('ev_acc_105', {}), ('ev_data_105', {"ACHAT": {}, "VENTE": {}, "ENTRÉE BQ": {}, "SORTIE BQ": {}}),
                         ('token_headers_105', {}), ('company_id_105', None), ('companies_list', []), ('eraz_log', []),
                         ('eraz_counts', {"COMPTE": 0, "ACHAT": 0, "VENTE": 0, "ENTRÉE BQ": 0, "SORTIE BQ": 0}),
                         ('eraz_items', {})]:
    if key not in st.session_state:
        st.session_state[key] = default

# --- Constantes & helpers requis par la sidebar (definis avant pour eviter NameError) ---
APP_DIR = os.path.dirname(os.path.abspath(__file__))
CREDS_PATH = os.path.join(APP_DIR, ".evoliz_creds.json")
PARAM_PATH = os.path.join(APP_DIR, "param_local.csv")
BALANCE_PATH_FILE = os.path.join(APP_DIR, ".last_balance_path.txt")

def save_param_local(df):
    df.to_csv(PARAM_PATH, index=False)

def load_param_local():
    if os.path.exists(PARAM_PATH):
        return pd.read_csv(PARAM_PATH)
    return pd.DataFrame()

# --- Configuration du perimetre ---
with st.sidebar:
    st.header("⚙️ Perimetre d'integration")
    scope = st.radio("Type de parametrage", ["Parametrage complet", "Ventes seules"], key="scope_mode")
    st.divider()
    st.subheader("Modules a activer")
    if scope == "Parametrage complet":
        mod_compta = st.checkbox("📂 Comptabilite (Balance, Param, Synchro)", value=True, key="mod_compta")
    else:
        mod_compta = False
    mod_clients = st.checkbox("👥 Injection Clients", value=True, key="mod_clients")
    mod_fournisseurs = st.checkbox("🏭 Injection Fournisseurs", value=True, key="mod_fournisseurs")
    mod_articles = st.checkbox("📦 Articles", value=True, key="mod_articles")
    mod_factures = st.checkbox("🧾 Factures & Avoirs", value=scope == "Parametrage complet", key="mod_factures")
    if mod_compta:
        with st.expander("📂 Parametres comptables", expanded=False):
            _f_param_sb = st.file_uploader("Importer un fichier parametres", type=["xlsm", "xlsx", "xls"], key="imp_param_sb", label_visibility="collapsed")
            if _f_param_sb: st.session_state["imp_file_param"] = _f_param_sb
            show_param = st.checkbox("Voir / éditer les paramètres", value=False, key="show_param")
            if show_param and not st.session_state.nr_v62.empty:
                _edited_sb = st.data_editor(
                    st.session_state.nr_v62, num_rows="dynamic", use_container_width=True, key="param_editor_sb"
                )
                if st.button("💾 Sauvegarder", key="btn_save_param_sb"):
                    st.session_state.nr_v62 = _edited_sb
                    save_param_local(_edited_sb)
                    st.success("Sauvegardé")
                st.download_button(
                    "📥 CSV", data=st.session_state.nr_v62.to_csv(index=False),
                    file_name="param_local.csv", mime="text/csv", key="btn_dl_param_sb"
                )
            fusion_tva = st.checkbox("Fusionner classif. par taux TVA", value=True, key="fusion_tva",
                                      help="Si coche, les classifications 6xx/7xx qui ne different que par le taux de TVA sont regroupees en une seule.")
            tva_achat_rate = st.number_input("Taux TVA achats (%)", value=20.0, step=0.5, min_value=0.0, max_value=100.0, key="tva_achat_rate",
                                              help="Applique aux classifications d'achat (comptes 2xx/6xx).")
    else:
        show_param = False
        fusion_tva = True
        tva_achat_rate = 20.0

# Élargir la sidebar quand on édite les paramètres comptables
if show_param:
    st.markdown("""<style>
        [data-testid="stSidebar"] { min-width: 600px; max-width: 800px; }
    </style>""", unsafe_allow_html=True)

def save_creds(pk, sk):
    """No-op : les cles API ne sont pas persistees entre rechargements (securite)."""
    pass

def load_creds():
    """Supprime tout fichier de cles residuel et retourne des champs vides.
    Les cles doivent etre saisies a nouveau a chaque rechargement."""
    try:
        if os.path.exists(CREDS_PATH):
            os.remove(CREDS_PATH)
    except Exception:
        pass
    return "", ""

def save_balance_path(path):
    if os.path.isfile(path):
        with open(BALANCE_PATH_FILE, 'w') as f:
            f.write(path)

def load_balance_path():
    if os.path.exists(BALANCE_PATH_FILE):
        with open(BALANCE_PATH_FILE) as f:
            p = f.read().strip()
            if os.path.isfile(p):
                return p
    return ""

# Construction dynamique des onglets - uniquement Connexion API tant qu'aucun dossier n'est connecte
_tab_names = []
_tab_keys = []
_tab_names.append("🔑 Connexion API"); _tab_keys.append("api")

_connected = bool(st.session_state.get('company_id_105')) and bool(st.session_state.get('token_headers_105'))
if _connected:
    _tab_names.append("📁 Import fichiers"); _tab_keys.append("import")
    if mod_compta:
        _tab_names.append("🔍 Matrice comptable"); _tab_keys.append("matrice")
    if mod_clients:
        _tab_names.append("👥 Injection Clients"); _tab_keys.append("clients")
    if mod_fournisseurs:
        _tab_names.append("🏭 Injection Fournisseurs"); _tab_keys.append("fournisseurs")
    if mod_articles:
        _tab_names.append("📦 Articles"); _tab_keys.append("articles")
    if mod_factures:
        _tab_names.append("🧾 Factures & Avoirs"); _tab_keys.append("factures")

_tabs = st.tabs(_tab_names)
_tab_map = {k: t for k, t in zip(_tab_keys, _tabs)}

# Aliases pour compatibilite avec le code existant
m2 = _tab_map.get("api", st.container())
m_import = _tab_map.get("import", st.container())
m1 = st.container()  # Onglet Param supprimé, paramètres dans la sidebar
m4 = _tab_map.get("matrice", st.container())
m6 = _tab_map.get("synthese", st.container())
m7 = _tab_map.get("synchro", st.container())
m_cli = _tab_map.get("clients", st.container())
m_four = _tab_map.get("fournisseurs", st.container())
m_fac = _tab_map.get("factures", st.container())
m_art = _tab_map.get("articles", st.container())
# Flags pour conditionner l'execution du contenu
_has_param_tab = False
_has_matrice_tab = "matrice" in _tab_map
_has_synthese_tab = "synthese" in _tab_map
_has_synchro_tab = "synchro" in _tab_map
_has_clients_tab = "clients" in _tab_map
_has_fournisseurs_tab = "fournisseurs" in _tab_map
_has_factures_tab = "factures" in _tab_map
_has_articles_tab = "articles" in _tab_map

def _get_sheet_names(f):
    """Retourne la liste des onglets d'un fichier Excel, ou None si CSV/mono-feuille HTML."""
    if hasattr(f, 'name') and f.name.lower().endswith('.csv'):
        return None
    fname = getattr(f, 'name', '').lower()
    if fname.endswith('.xls'):
        engines = ["xlrd", None]
    elif fname.endswith(('.xlsx', '.xlsm')):
        engines = ["openpyxl", None]
    else:
        engines = [None, "openpyxl", "xlrd"]
    for engine in engines:
        try:
            f.seek(0)
            xl = pd.ExcelFile(f, engine=engine)
            return xl.sheet_names
        except Exception:
            pass
    # Fallback : certains .xls sont du HTML déguisé
    try:
        f.seek(0)
        raw = f.read() if hasattr(f, 'read') else open(f, 'rb').read()
        f.seek(0)
        dfs = pd.read_html(raw, header=0)
        if dfs and len(dfs) > 1:
            return [f"Feuille {i+1}" for i in range(len(dfs))]
    except Exception:
        pass
    return None

def _sheet_selector(f, label, key):
    """Si le fichier a plusieurs onglets, affiche un selectbox et retourne le nom choisi."""
    sheets = _get_sheet_names(f)
    if sheets and len(sheets) > 1:
        chosen = st.selectbox(f"📑 Onglet à utiliser ({label})", sheets, key=key)
        if chosen.startswith("Feuille "):
            try: return int(chosen.split(" ")[1]) - 1
            except Exception: pass
        return chosen
    return 0

# --- Onglet Import fichiers centralise ---
if _connected:
 with m_import:
    st.subheader("📁 Import des fichiers sources")
    st.caption("Centralisez ici tous vos fichiers. Ils seront utilises dans les onglets correspondants.")

    # Gate logique : pas d'import tant qu'aucun dossier Evoliz n'est identifie
    _gate_import = bool(st.session_state.get('company_id_105')) and bool(st.session_state.get('token_headers_105'))
    if not _gate_import:
        st.warning("⛔ **Import bloque** — Connectez-vous a l'API Evoliz puis selectionnez un dossier dans l'onglet **🔑 Connexion API** avant d'importer des fichiers.")

    # Layout compact : label | browse | statut sur la même ligne
    # + sélection d'onglet automatique si fichier multi-feuilles
    def _file_row(label, types, session_key, uploader_key):
        _f_uploaded = st.session_state.get(session_key)
        # Détecter les onglets du fichier déjà chargé pour dimensionner la ligne
        _sheets = _get_sheet_names(_f_uploaded) if _f_uploaded else None
        _has_sheets = _sheets and len(_sheets) > 1
        if _has_sheets:
            c1, c2, c3, c4 = st.columns([1.2, 3, 1.8, 1.2])
        else:
            c1, c2, c3 = st.columns([1.5, 4, 1.5])
        c1.markdown(f"**{label}**")
        _f = c2.file_uploader(label, type=types, key=uploader_key, label_visibility="collapsed")
        if _f: st.session_state[session_key] = _f; _f_uploaded = _f
        if _has_sheets:
            _sheet_key = f"_sheet_{session_key}"
            _chosen = c3.selectbox(f"Onglet", _sheets, key=_sheet_key, label_visibility="collapsed")
            if _chosen and str(_chosen).startswith("Feuille "):
                try: _chosen = int(_chosen.split(" ")[1]) - 1
                except Exception: pass
            st.session_state[f"{session_key}_sheet"] = _chosen
            c4.markdown(f"✅ {_f_uploaded.name}" if _f_uploaded else "❌ —")
        else:
            c3.markdown(f"✅ {_f_uploaded.name}" if _f_uploaded else "❌ —")
            if _f_uploaded:
                st.session_state[f"{session_key}_sheet"] = 0

    if _gate_import:
        if mod_compta:
            _file_row("📂 Balance", ["xlsm", "xlsx", "xls"], "imp_file_balance", "imp_balance")
        if mod_clients:
            _file_row("👥 Clients", ["xlsx", "xls", "csv"], "imp_file_clients", "imp_clients")
        if mod_fournisseurs:
            _file_row("🏭 Fournisseurs", ["xlsx", "xls", "csv"], "imp_file_fournisseurs", "imp_fournisseurs")
        if mod_factures:
            _file_row("🧾 Factures", ["xlsx", "xls"], "imp_file_factures", "imp_factures")
        if mod_articles:
            _file_row("📦 Articles (Excel)", ["xlsx", "xls"], "imp_file_articles", "imp_articles")
            # Uploader PDF multi-fichiers pour les factures fournisseur (complete l'Excel)
            _c1, _c2, _c3 = st.columns([1.5, 4, 1.5])
            _c1.markdown("**📸 Factures PDF**")
            _pdfs_art = _c2.file_uploader(
                "Factures PDF articles", type=["pdf"], accept_multiple_files=True,
                key="imp_articles_pdfs", label_visibility="collapsed",
                help="Jusqu'a 100 factures PDF. Les lignes d'article seront extraites et fusionnees avec l'import Excel.",
            )
            if _pdfs_art:
                st.session_state["imp_file_articles_pdfs"] = _pdfs_art
                _c3.markdown(f"✅ {len(_pdfs_art)} PDF")
            else:
                _existing_pdfs = st.session_state.get("imp_file_articles_pdfs", [])
                _c3.markdown(f"✅ {len(_existing_pdfs)} PDF" if _existing_pdfs else "❌ —")


def load_param_from_excel(file_obj):
    raw_p = pd.read_excel(file_obj, sheet_name="Param", header=None)
    df_p = pd.DataFrame()
    df_p['Racine'] = raw_p.iloc[1:, 0].apply(to_clean_str)
    df_p['Libellé'] = raw_p.iloc[1:, 1]
    df_p['_tags'] = raw_p.iloc[1:].apply(
        lambda r: [norm_piv(r.iloc[i]) for i in range(7, 17) if not pd.isna(r.iloc[i])],
        axis=1
    )
    for flux, tag in [("ACHAT", "ACHAT"), ("VENTE", "VENTE"), ("ENTRÉE BQ", "ENTREE"), ("SORTIE BQ", "SORTIE")]:
        df_p[flux] = df_p['_tags'].apply(lambda tags, t=tag: t in tags)
    df_p = df_p.drop(columns=['_tags'])
    return df_p.dropna(subset=['Racine']).reset_index(drop=True)

# Chargement auto des params (necessaire pour la matrice, meme sans onglet Param visible)
if st.session_state.nr_v62.empty and os.path.exists(PARAM_PATH):
    st.session_state.nr_v62 = load_param_local()
_f_param_loaded = st.session_state.get("imp_file_param")
if _f_param_loaded:
    try:
        _xl_p = pd.ExcelFile(_f_param_loaded)
        if "Param" in _xl_p.sheet_names:
            st.session_state.nr_v62 = load_param_from_excel(_f_param_loaded)
            save_param_local(st.session_state.nr_v62)
    except Exception:
        pass

with m2:
    # --- Clés API ---
    st.subheader("🔑 Connexion Evoliz")
    saved_pk, saved_sk = load_creds()
    col_pk, col_sk = st.columns(2)
    st.caption("🔐 L'app detecte automatiquement le type de cle : **client** (company_users — mono-dossier) ou **plateforme** (prescriber_users — multi-dossier).")
    st.caption("🛡️ **Securite** : aucune cle API n'est sauvegardee. Les cles sont effacees de la memoire des que le token est obtenu, et les champs sont vides a chaque rafraichissement.")
    pk_105 = col_pk.text_input("Public Key", value=saved_pk, key="pk_105")
    sk_105 = col_sk.text_input("Secret Key", type="password", value=saved_sk, key="sk_105")

    # --- Detection du changement de cle : vider toutes les donnees API precedentes ---
    _prev_pk = st.session_state.get("_prev_pk_105", "")
    _prev_sk = st.session_state.get("_prev_sk_105", "")
    _keys_changed = (pk_105 != _prev_pk or sk_105 != _prev_sk) and (_prev_pk or _prev_sk)
    if _keys_changed:
        # Vider toutes les donnees API/session relatives a l'ancien token
        for _k in (
            'token_headers_105', 'company_id_105', 'companies_list',
            'ev_acc_105', 'ev_data_105', 'ev_clients_raw', 'ev_articles_raw', 'ev_invoices_raw',
            '_key_mode', 'eraz_log', 'sync_log',
            '_art_consol', '_art_consol_stats', '_art_sale_cl', '_art_consol_id',
            '_art_pdf_rows', '_art_pdf_raw', '_art_pdf_extracted_rows',
            '_art_pending_classifs', '_art_hidden_cols',
            '_four_consol', '_four_consol_stats', '_four_sirene_cells', '_four_sirene_result',
            '_2eme_lame_four_props', '_2eme_lame_four_result',
            'meg_df_clients', 'meg_sirene_cells', 'meg_sirene_info', 'meg_sirene_log',
            'meg_sirene_stats', 'meg_enrichir_flags', 'meg_sirene_suggestions',
            '_2eme_lame_props', '_2eme_lame_result',
            '_bal_analysed_id', '_synth_modif', '_skip_delete', '_skip_create',
        ):
            if _k in st.session_state:
                _v = st.session_state[_k]
                if isinstance(_v, dict): st.session_state[_k] = {}
                elif isinstance(_v, list): st.session_state[_k] = []
                elif isinstance(_v, set): st.session_state[_k] = set()
                elif isinstance(_v, bool): st.session_state[_k] = False
                elif isinstance(_v, int): st.session_state[_k] = 0
                else: st.session_state[_k] = None
        # Reset specifique aux types particuliers (DataFrames, dict structures)
        st.session_state.ev_data_105 = {"ACHAT": {}, "VENTE": {}, "ENTRÉE BQ": {}, "SORTIE BQ": {}}
        st.session_state.token_headers_105 = {}
        st.session_state.audit_matrix_105 = pd.DataFrame()
        st.session_state.rejets_105 = pd.DataFrame()
        st.session_state.prot_105 = set()
        st.session_state.eraz_counts = {"COMPTE": 0, "ACHAT": 0, "VENTE": 0, "ENTRÉE BQ": 0, "SORTIE BQ": 0}
        st.session_state.eraz_items = {}
        st.info("🔄 Cle API modifiee — donnees precedentes effacees. Reconnectez-vous.")
        st.session_state["_prev_pk_105"] = pk_105
        st.session_state["_prev_sk_105"] = sk_105
        st.rerun()
    st.session_state["_prev_pk_105"] = pk_105
    st.session_state["_prev_sk_105"] = sk_105

    # --- Bloc 1 : Login (récupère le token + liste des dossiers accessibles) ---
    auto_connect = not st.session_state.token_headers_105 and saved_pk and saved_sk
    # Toujours rendre le bouton (même si auto_connect va le déclencher)
    btn_connect = st.button("🔗 CONNECTER", type="primary", use_container_width=True, key="btn_connect_105")
    # Bouton déconnexion si déjà connecté
    if st.session_state.token_headers_105:
        if st.button("🔌 Se déconnecter / changer de clés", key="btn_disconnect"):
            for _k in ('token_headers_105', 'company_id_105', 'companies_list',
                        'ev_acc_105', 'ev_data_105', 'ev_clients_raw', 'ev_articles_raw', 'ev_invoices_raw'):
                if _k in st.session_state:
                    st.session_state[_k] = type(st.session_state[_k])() if isinstance(st.session_state[_k], (dict, list, set)) else None
            st.session_state.token_headers_105 = {}
            st.rerun()
    if auto_connect or btn_connect:
        if pk_105 and sk_105:
            save_creds(pk_105, sk_105)
            r_log = None
            with st.spinner("Connexion à Evoliz en cours..."):
                try:
                    r_log = requests.post("https://www.evoliz.io/api/login",
                                          json={"public_key": pk_105, "secret_key": sk_105}, timeout=15)
                except Exception as e:
                    st.error(f"Erreur de connexion : {e}")
            if r_log is None:
                st.error("Aucune réponse du serveur Evoliz (timeout ou erreur réseau).")
            elif r_log.status_code in (429, 500, 502, 503, 504):
                st.warning(f"API Evoliz temporairement indisponible (HTTP {r_log.status_code}). Réessayez dans quelques instants.")
            elif r_log.status_code == 200:
                login_data = r_log.json()
                h = {"Authorization": f"Bearer {login_data.get('access_token')}", "Accept": "application/json"}
                st.session_state.token_headers_105 = h

                # --- SECURITE : effacer immediatement les cles API de la memoire ---
                # PK + SK ne sont plus necessaires apres login (seul le token est utilise)
                pk_105 = ""; sk_105 = ""
                for _k in ("pk_105", "sk_105", "_prev_pk_105", "_prev_sk_105"):
                    if _k in st.session_state:
                        try: del st.session_state[_k]
                        except Exception: pass
                # Effacer aussi de l'objet login_data pour eviter qu'il traine
                if "access_token" in login_data:
                    # On garde une copie minimale dans une variable locale, pas en session
                    pass

                _scopes = login_data.get("scopes", []) or []
                if isinstance(_scopes, str):
                    _scopes = [s.strip() for s in _scopes.split(",") if s.strip()]
                _is_multi = "prescriber_users" in _scopes
                st.session_state["_key_mode"] = "multi" if _is_multi else "mono"

                # --- MODE MULTI (prescriber_users) : GET /companies liste tous les dossiers ---
                if _is_multi:
                    _companies_all = []
                    try:
                        _pg = 1
                        while True:
                            r_co = requests.get("https://www.evoliz.io/api/v1/companies", headers=h,
                                                params={"per_page": 100, "page": _pg}, timeout=15)
                            if r_co.status_code != 200: break
                            _d = r_co.json()
                            _companies_all.extend(_d.get("data", []))
                            if _pg >= _d.get("meta", {}).get("last_page", 1): break
                            _pg += 1
                    except Exception as e:
                        st.error(f"Erreur GET /companies : {e}")
                    # Normaliser name
                    for _c in _companies_all:
                        if "name" not in _c:
                            _c["name"] = _c.get("company_name") or f"Dossier {_c.get('companyid', '?')}"
                    st.session_state.companies_list = _companies_all
                    if len(_companies_all) == 1:
                        _c0 = _companies_all[0]
                        st.session_state.company_id_105 = _c0.get("companyid")
                        st.success(f"🔑 **Cle plateforme (prescriber_users)** — 1 dossier : **{_c0.get('company_name', 'N/C')}** (ID: {_c0.get('companyid')})")
                        st.rerun()
                    elif len(_companies_all) > 1:
                        st.session_state.company_id_105 = None
                        st.success(f"🔑 **Cle plateforme (prescriber_users)** — {len(_companies_all)} dossiers accessibles. Selectionnez un dossier ci-dessous.")
                    else:
                        st.error("❌ Aucun dossier accessible avec cette cle.")
                else:
                    # --- MODE MONO (company_users) ---
                    # L'API Evoliz ne permet PAS aux company_users de recuperer leur company_name :
                    # /companies requiert prescriber_users, /companies/{cid} renvoie 403 meme
                    # avec le vrai cid. Les ressources scopees (/clients, /articles, ...) fonctionnent
                    # mais n'exposent pas le companyid dans leurs reponses.
                    # -> On se passe du cid : les URLs /api/v1/{endpoint} (sans prefixe) fonctionnent
                    #    nativement pour company_users.
                    cid = None
                    _co_name = None
                    _diag_steps = []

                    # Decoder JWT pour sub (= user_id, utilise comme placeholder)
                    try:
                        import base64 as _b64, json as _jsn
                        _tok_str = login_data.get("access_token", "")
                        _parts = _tok_str.split(".")
                        if len(_parts) >= 2:
                            _pad = _parts[1] + "=" * (-len(_parts[1]) % 4)
                            _payload = _jsn.loads(_b64.urlsafe_b64decode(_pad))
                            _sub = _payload.get("sub")
                            if _sub:
                                try: cid = int(_sub)
                                except (TypeError, ValueError): cid = _sub
                    except Exception:
                        pass

                    if not cid:
                        st.error("❌ Impossible d'extraire le user_id depuis le token.")
                    else:
                        _co_name = _co_name or "Mon dossier"
                        st.session_state.company_id_105 = cid
                        st.session_state.companies_list = [{
                            "companyid": cid,
                            "company_name": _co_name,
                            "name": _co_name,
                        }]
                        # Mode no-prefix : les URLs /api/v1/{endpoint} fonctionnent nativement
                        st.session_state["_api_no_prefix"] = True
                        st.success(f"🔑 **Cle client (company_users)** — dossier : **{_co_name}** (ID: {cid})")
                        st.rerun()
            elif r_log.status_code == 401:
                st.error(f"❌ Echec login : HTTP 401 — **credentials invalides**")
                with st.expander("🔍 Diagnostic", expanded=True):
                    st.markdown("""
                    **Causes possibles** :
                    - La **Public Key** ou la **Secret Key** est incorrecte (verifier la copie depuis Evoliz)
                    - La paire de cles a ete **revoque/regeneree** depuis Evoliz → recreer une nouvelle paire
                    - Les cles sont expirees (duree de vie configuree cote Evoliz)
                    - Espaces parasites en debut/fin de cle

                    **Verification** :
                    1. Connecte-toi sur ton compte Evoliz (fiteco.evoliz.com ou equivalent)
                    2. Va dans **Parametres > API** (ou **Mon compte > Cles API**)
                    3. Cree une **nouvelle paire de cles de type Client** (company_users)
                    4. Copie-colle integralement sans espaces
                    5. Utilise cette paire ici
                    """)
                    st.caption(f"Reponse brute : `{r_log.text[:200]}`")
            else:
                st.error(f"Échec login : HTTP {r_log.status_code} — {r_log.text[:300]}")
        else:
            st.warning("Renseignez la Public Key et la Secret Key pour vous connecter.")

    # --- Bloc 2 : Sélecteur de dossier (multi-company) ---
    _companies = st.session_state.get('companies_list', [])
    if st.session_state.token_headers_105 and len(_companies) > 1:
        st.divider()
        st.subheader("📁 Sélection du dossier")

        # Filtre par site (home_site.home_site dans le détail company)
        def _get_site(c):
            hs = c.get('home_site')
            if isinstance(hs, dict):
                return hs.get('home_site', '')
            return ''
        _sites = sorted({_get_site(c) for c in _companies})
        _sites = [s for s in _sites if s]

        _filtered = _companies
        if _sites:
            _site_filter = st.selectbox("Filtrer par site", ["— Tous les sites —"] + _sites, key="site_filter")
            if _site_filter != "— Tous les sites —":
                _filtered = [c for c in _companies if _get_site(c) == _site_filter]

        def _get_company_label(c):
            return c.get('company_name') or c.get('name') or f"Company {c.get('companyid')}"
        _company_names = [f"{_get_company_label(c)}" for c in _filtered]
        _company_ids = [c.get('companyid') or c.get('id') for c in _filtered]
        _prev_cid = st.session_state.company_id_105
        _sel_idx = st.selectbox("Dossier Evoliz", range(len(_company_names)),
                                format_func=lambda i: _company_names[i],
                                index=None, placeholder="— Sélectionnez un dossier —",
                                key="company_select")
        if _sel_idx is not None:
            _sel_cid = _company_ids[_sel_idx]
            if _sel_cid != _prev_cid:
                st.session_state.company_id_105 = _sel_cid
                # Réinitialiser les données du précédent dossier
                st.session_state.ev_acc_105 = {}
                st.session_state.ev_data_105 = {"ACHAT": {}, "VENTE": {}, "ENTRÉE BQ": {}, "SORTIE BQ": {}}
                for _k in ("ev_clients_raw", "ev_articles_raw", "ev_invoices_raw"):
                    st.session_state[_k] = []
                st.rerun()


    # --- Affichage du dossier actif ---
    def _get_company_label(c):
        return c.get('company_name') or c.get('name') or f"Company {c.get('companyid')}"
    _cid = st.session_state.company_id_105
    if _cid and _companies:
        _current = next((c for c in _companies if (c.get('companyid') or c.get('id')) == _cid), None)
        if _current:
            st.info(f"📂 Dossier actif : **{_get_company_label(_current)}** (ID: {_cid})")

    # --- Bloc 3 : Chargement des données ---
    _is_multi = len(st.session_state.get('companies_list', [])) > 1
    _cid = st.session_state.company_id_105
    _h = st.session_state.token_headers_105
    _data_empty = not (st.session_state.ev_acc_105 or st.session_state.get("ev_clients_raw") or st.session_state.get("ev_articles_raw") or st.session_state.get("ev_invoices_raw"))
    # Chargement auto dès qu'un dossier est sélectionné et que les données sont vides
    _should_load = _h and _cid and _data_empty
    if _should_load:
        # Pour company_users (mono) : certains endpoints refusent le prefixe /companies/{cid}/
        #   -> on utilise /api/v1/{endpoint} (prefixe optionnel selon doc Evoliz)
        # Pour prescriber_users (multi) : le prefixe est requis.
        _mode = st.session_state.get("_key_mode", "multi")
        if _mode == "mono":
            _base = "https://www.evoliz.io/api/v1"
            _cid_for_fetch = None  # fetch_evoliz_data utilisera /api/v1/{endpoint} sans prefixe
        else:
            _base = f"https://www.evoliz.io/api/v1/companies/{_cid}"
            _cid_for_fetch = _cid

        def _fetch_paginated(endpoint, headers, params_extra=None):
            """Lecture paginée d'un endpoint Evoliz (thread-safe)."""
            out = []; pg = 1
            while True:
                p = {"per_page": 100, "page": pg}
                if params_extra: p.update(params_extra)
                r = requests.get(endpoint, headers=headers, params=p, timeout=15)
                if r.status_code != 200: break
                d = r.json(); out.extend(d.get("data", []))
                if pg >= d.get("meta", {}).get("last_page", 1): break
                pg += 1
            return out

        with st.spinner("Lecture des données Evoliz (parallèle)..."):
            from datetime import timedelta
            _date_from = (dt_datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

            # 8 tâches en parallèle : 5 compta + clients + articles + factures
            # Pour les endpoints compta (accounts, classifications, affectations) : fetch_evoliz_data
            # gere lui-meme le prefixe /companies/{cid}/ (ou absence si cid=None).
            # Pour clients/articles/invoices : URL complete construite a partir de _base.
            _load_tasks = {
                "accounts": ("accounts", _h, _cid_for_fetch, "compta", None),
                "purchase-classifications": ("purchase-classifications", _h, _cid_for_fetch, "compta", None),
                "sale-classifications": ("sale-classifications", _h, _cid_for_fetch, "compta", None),
                "sale-affectations": ("sale-affectations", _h, _cid_for_fetch, "compta", None),
                "purchase-affectations": ("purchase-affectations", _h, _cid_for_fetch, "compta", None),
                "clients": (f"{_base}/clients", _h, None, "url", None),
                "articles": (f"{_base}/articles", _h, None, "url", None),
                "invoices": (f"{_base}/invoices", _h, None, "url", None),
            }

            def _run_fetch(key, args):
                endpoint, headers, cid, kind, extra = args
                if kind == "compta":
                    return key, fetch_evoliz_data(endpoint, headers, company_id=cid)
                else:
                    return key, _fetch_paginated(endpoint, headers, params_extra=extra)

            _results = {}
            with ThreadPoolExecutor(max_workers=8) as pool:
                futures = {pool.submit(_run_fetch, k, v): k for k, v in _load_tasks.items()}
                for f in as_completed(futures):
                    k, data = f.result()
                    _results[k] = data

            st.session_state.ev_acc_105 = _results.get("accounts", {})
            st.session_state.ev_data_105 = {
                "ACHAT": _results.get("purchase-classifications", {}),
                "VENTE": _results.get("sale-classifications", {}),
                "ENTRÉE BQ": _results.get("sale-affectations", {}),
                "SORTIE BQ": _results.get("purchase-affectations", {}),
            }
            st.session_state["ev_clients_raw"] = _results.get("clients", [])
            st.session_state["ev_articles_raw"] = _results.get("articles", [])
            st.session_state["ev_invoices_raw"] = _results.get("invoices", [])
        _company_name = ""
        for _c in st.session_state.get('companies_list', []):
            if (_c.get('companyid') or _c.get('id')) == _cid:
                _company_name = _c.get('name', '')
                break
        st.success(f"Dossier « {_company_name or _cid} » chargé — {len(st.session_state.get('ev_clients_raw', []))} clients, {len(st.session_state.get('ev_articles_raw', []))} articles, {len(st.session_state.get('ev_invoices_raw', []))} factures (30j)")
    elif _h and not _cid and len(_companies) > 1:
        st.info("👆 Sélectionnez un dossier ci-dessus pour charger les données.")

    _has_any_data = (st.session_state.ev_acc_105
                      or st.session_state.get("ev_clients_raw")
                      or st.session_state.get("ev_articles_raw")
                      or st.session_state.get("ev_invoices_raw"))
    if _has_any_data:
        # Nom du dossier actif affiche dans le titre de l'expander
        _cid_disp = st.session_state.company_id_105
        _co_list_disp = st.session_state.get("companies_list", [])
        _co_disp = next((c for c in _co_list_disp if (c.get('companyid') or c.get('id')) == _cid_disp), None)
        _co_nm = (_co_disp or {}).get("company_name") or (_co_disp or {}).get("name") or f"Dossier {_cid_disp}"
        with st.expander(f"📊 Données lues depuis Evoliz — **{_co_nm}** (ID: {_cid_disp})", expanded=True):
            cc1, cc2, cc3 = st.columns(3)
            cc1.metric("👥 Clients", len(st.session_state.get("ev_clients_raw", [])))
            cc2.metric("📦 Articles", len(st.session_state.get("ev_articles_raw", [])))
            cc3.metric("🧾 Factures (30j)", len(st.session_state.get("ev_invoices_raw", [])))
            if mod_compta:
                st.divider()
                st.caption("Pre-comptabilite")
                cp1, cp2, cp3, cp4, cp5 = st.columns(5)
                cp1.metric("Comptes", len(st.session_state.get("ev_acc_105", {})))
                cp2.metric("Achats", count_unique(st.session_state.get("ev_data_105", {}).get("ACHAT", {})))
                cp3.metric("Ventes", count_unique(st.session_state.get("ev_data_105", {}).get("VENTE", {})))
                cp4.metric("Entrees BQ", count_unique(st.session_state.get("ev_data_105", {}).get("ENTRÉE BQ", {})))
                cp5.metric("Sorties BQ", count_unique(st.session_state.get("ev_data_105", {}).get("SORTIE BQ", {})))

# Le code Balance s'execute dans l'onglet Import fichiers (traitement silencieux)
if _connected:
 with m_import:
  if mod_compta and st.session_state.get('company_id_105') and st.session_state.get('token_headers_105'):
    f105 = st.session_state.get("imp_file_balance")
    if not f105:
        # Fallback : chemin local
        last_bal_path = load_balance_path()
        if last_bal_path:
            f105 = last_bal_path
    if not f105:
        st.info("Chargez vos fichiers ci-dessus.")

    if f105:
        xl = None
        # Essayer plusieurs engines pour supporter tous les formats Excel
        for engine, kwargs in [(None, {}), ('openpyxl', {}), ('xlrd', {}), ('xlrd', {'engine_kwargs': {'ignore_workbook_corruption': True}})]:
            try:
                if hasattr(f105, 'seek'):
                    f105.seek(0)
                xl = pd.ExcelFile(f105, engine=engine, **kwargs)
                break
            except Exception:
                pass
        # Fallback : lire comme HTML (certains .xls sont du HTML déguisé)
        if xl is None:
            try:
                if hasattr(f105, 'seek'):
                    f105.seek(0)
                raw = f105.read() if hasattr(f105, 'read') else open(f105, 'rb').read()
                dfs = pd.read_html(raw, header=0)
                if dfs:
                    st.session_state._bal_html_fallback = dfs
                    xl = "HTML"
            except Exception:
                pass
        if xl is None:
            st.error("Impossible de lire ce fichier. Essayez de le ré-enregistrer en .xlsx depuis Excel.")
            st.stop()
        # Utiliser l'onglet choisi à l'import (ou le premier par défaut)
        _bal_sheet = st.session_state.get("imp_file_balance_sheet", 0)
        if xl == "HTML":
            sheet_idx = _bal_sheet if isinstance(_bal_sheet, int) else 0
            df_bal_preview = st.session_state._bal_html_fallback[sheet_idx]
        else:
            sheet_bal = _bal_sheet if _bal_sheet != 0 and _bal_sheet in xl.sheet_names else xl.sheet_names[0]
            df_bal_preview = xl.parse(sheet_name=sheet_bal)
        cols_preview = df_bal_preview.columns.tolist()
        comptes_4456 = []
        for _, row in df_bal_preview.iterrows():
            code = to_clean_str(row[cols_preview[0]])
            if code.startswith('4456'):
                label = str(row[cols_preview[1]]).strip().upper()
                comptes_4456.append({"code": code, "label": f"{code} - {label}"})

        st.divider()
        st.subheader("🏷️ Comptes TVA pour catégories d'achat")
        options_4456 = ["— Aucun"] + [c['label'] for c in comptes_4456]

        # Présélection par défaut : 1er 44566xxx pour achats, 1er 44562xxx pour immos
        _default_6 = 0  # "— Aucun"
        _default_2 = 0
        for _i, _c in enumerate(comptes_4456):
            if _c['code'].startswith('44566') and _default_6 == 0:
                _default_6 = _i + 1  # +1 car "— Aucun" est en position 0
            if _c['code'].startswith('44562') and _default_2 == 0:
                _default_2 = _i + 1

        col_tva1, col_tva2 = st.columns(2)
        with col_tva1:
            sel_tva_6 = st.selectbox("TVA Deductible (comptes 6xx)", options_4456, index=_default_6, key="sel_tva_6")
        with col_tva2:
            sel_tva_2 = st.selectbox("TVA deductible sur immos (comptes 2xx)", options_4456, index=_default_2, key="sel_tva_2")

        vat_label_6, vat_label_2 = None, None
        vat_api_id_6, vat_api_id_2 = None, None
        if sel_tva_6 != "— Aucun":
            vat_label_6 = sel_tva_6
            code_tva_6 = sel_tva_6.split(" - ")[0]
            ev_tva_6 = st.session_state.ev_acc_105.get(norm_piv(code_tva_6))
            vat_api_id_6 = ev_tva_6['id'] if ev_tva_6 else None
        if sel_tva_2 != "— Aucun":
            vat_label_2 = sel_tva_2
            code_tva_2 = sel_tva_2.split(" - ")[0]
            ev_tva_2 = st.session_state.ev_acc_105.get(norm_piv(code_tva_2))
            vat_api_id_2 = ev_tva_2['id'] if ev_tva_2 else None

        st.divider()
        # Auto-analyse si fichier charge et pas encore analyse, ou bouton manuel
        _bal_file_id = str(f105) if isinstance(f105, str) else f105.name + str(f105.size)
        _bal_already = st.session_state.get("_bal_analysed_id") == _bal_file_id
        _bal_auto = not _bal_already and not st.session_state.nr_v62.empty
        _bal_manual = st.button("🔍 Re-ANALYSER", use_container_width=True, key="btn_analyse_105")
        if _bal_auto or _bal_manual:
            st.session_state["_bal_analysed_id"] = _bal_file_id
            st.session_state.vat_label_6 = vat_label_6
            st.session_state.vat_label_2 = vat_label_2
            st.session_state.vat_api_id_6 = vat_api_id_6
            st.session_state.vat_api_id_2 = vat_api_id_2
            df_bal = df_bal_preview
            cols = cols_preview
            flux_cols = [c for c in cols if any(x in str(c).upper() for x in ["DEBIT", "CREDIT", "SOLDE"])]

            rules_df = st.session_state.nr_v62
            sorted_roots = sorted(
                [str(r) for r in rules_df['Racine'].dropna() if str(r).strip()],
                key=len, reverse=True
            )

            seen_pivots, results, rejets = set(), [], []
            seen_codes = set()     # déduplication comptes par code
            prot_comptes = set()   # protections comptes : par numéro de compte
            prot_flux = set()      # protections flux : par label normalisé
            prot_flux_ids = set()  # id des flux reconnus par la matrice

            for _, row in df_bal.iterrows():
                code = to_clean_str(row[cols[0]])
                label = str(row[cols[1]]).strip().upper()
                label_flux = clean_label_tva(label, code, fusion_tva)
                pivot_flux = norm_piv(label_flux)
                mvt = has_movement_debit_credit(row, flux_cols) if flux_cols else True

                match_root = next((r for r in sorted_roots if code.startswith(r)), None)

                if match_root is None or code.startswith(('401', '411')):
                    raison = "Tiers 401/411" if match_root else "Pas de racine"
                    rejets.append({"N°": code, "Libellé": label, "Raison": raison})
                    continue

                # Déduplication : si ce code a déjà été traité, ignorer (doublon balance)
                code_norm = norm_piv(code)
                if code_norm in seen_codes:
                    rejets.append({"N°": code, "Libellé": label, "Raison": "Doublon (code déjà traité)"})
                    continue
                seen_codes.add(code_norm)

                match_row = rules_df[rules_df['Racine'].apply(to_clean_str) == match_root].iloc[0]

                ev_account = st.session_state.ev_acc_105.get(code_norm)
                vat_id = ev_account['vat_id'] if ev_account else None
                account_exists = ev_account is not None

                # TVA : libellé pour affichage, id API pour injection
                tva_display = "—"
                tva_api_id = vat_id
                if match_row.get("ACHAT", False):
                    if code.startswith('6') and st.session_state.get('vat_label_6'):
                        tva_display = st.session_state.vat_label_6
                        if not tva_api_id:
                            tva_api_id = st.session_state.get('vat_api_id_6')
                    elif code.startswith('2') and st.session_state.get('vat_label_2'):
                        tva_display = st.session_state.vat_label_2
                        if not tva_api_id:
                            tva_api_id = st.session_state.get('vat_api_id_2')

                # Détection des écarts sur COMPTE existant
                compte_status = "➕"
                compte_patch = None
                if account_exists:
                    compte_status = "✅"
                    api_label_norm = norm_piv(ev_account['label'])
                    bal_label_norm = norm_piv(label)
                    if api_label_norm != bal_label_norm:
                        compte_status = "🔄"
                        compte_patch = {"_patch_cat": "COMPTE", "_patch_id": ev_account['id'],
                                        "_patch_payload": {"label": label},
                                        "_patch_detail": f"Libellé: {ev_account['label']} → {label}"}

                item = {
                    "Sync": compte_status in ("➕", "🔄"),
                    "N°": code, "Libellé": label, "LibFlux": label_flux,
                    "TVA": tva_display,
                    "_vat_id": tva_api_id,
                    "COMPTE": compte_status,
                }
                if compte_patch:
                    item.update(compte_patch)

                prot_comptes.add(norm_piv(code))

                # Résolution de l'accountid attendu pour les flux
                expected_acc_id = ev_account['id'] if ev_account else None

                is_first_pivot = pivot_flux not in seen_pivots
                for flux in ["ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
                    if match_row.get(flux, False) and mvt and is_first_pivot:
                        # --- RECHERCHE DU FLUX PAR SON CHAMP "code" UNIQUEMENT ---
                        # Le code d'un flux = son libellé. On cherche un flux dont
                        # norm_piv(code) == norm_piv(label_flux attendu par la matrice)
                        flux_store = st.session_state.ev_data_105[flux]
                        ev_flux = None
                        seen_flux_ids = set()
                        for d in flux_store.values():
                            if d['id'] in seen_flux_ids:
                                continue
                            seen_flux_ids.add(d['id'])
                            if norm_piv(d.get('code', '')) == pivot_flux:
                                ev_flux = d
                                break

                        if ev_flux:
                            prot_flux_ids.add(ev_flux['id'])
                            # --- COMPARAISON CHAMP PAR CHAMP ---
                            patch_payload = {}
                            patch_details = []
                            # 1. Code (doit être = label_flux)
                            api_code = ev_flux.get('code', '')
                            if norm_piv(api_code) != norm_piv(label_flux):
                                patch_payload["code"] = label_flux[:50]
                                patch_details.append(f"code: {api_code} → {label_flux}")
                            # 2. Label (doit être = label_flux)
                            api_label = ev_flux.get('label', '')
                            if norm_piv(api_label) != norm_piv(label_flux):
                                patch_payload["label"] = label_flux
                                patch_details.append(f"label: {api_label} → {label_flux}")
                            # 3. Accountid (doit pointer vers le bon compte)
                            current_acc = ev_flux.get('acc_id')
                            if expected_acc_id and current_acc != expected_acc_id:
                                patch_payload["accountid"] = expected_acc_id
                                patch_details.append(f"accountid: {current_acc} → {expected_acc_id}")
                            # 4. Vataccountid (ACHAT uniquement)
                            if flux == "ACHAT" and tva_api_id:
                                current_vat = ev_flux.get('vat_id')
                                if current_vat != tva_api_id:
                                    patch_payload["vataccountid"] = tva_api_id
                                    patch_details.append(f"vataccountid: {current_vat} → {tva_api_id}")

                            if patch_payload:
                                item[flux] = "🔄"
                                item[f"_patch_{flux}"] = {"cat": flux, "id": ev_flux['id'],
                                                          "payload": patch_payload, "detail": " | ".join(patch_details)}
                                item["Sync"] = True
                            else:
                                item[flux] = "✅"
                        else:
                            item[flux] = "➕"
                            item["Sync"] = True
                    else:
                        item[flux] = "—"

                if is_first_pivot and pivot_flux:
                    seen_pivots.add(pivot_flux)
                results.append(item)

            st.session_state.audit_matrix_105 = pd.DataFrame(results)
            st.session_state.rejets_105 = pd.DataFrame(rejets)
            st.session_state.prot_105 = {"comptes": prot_comptes, "flux": prot_flux, "flux_ids": prot_flux_ids}
            st.success(f"Analyse terminée : {len(results)} lignes traitées, {len(rejets)} rejetées → voir onglets Matrice / Rejetées")

if _connected and mod_compta:
 with m4:
    # Gate : necessite API + dossier + fichier Balance
    _gate_matrice = bool(st.session_state.get('company_id_105')) and bool(st.session_state.get('token_headers_105')) and bool(st.session_state.get('imp_file_balance'))
    if not _gate_matrice:
        if not st.session_state.get('company_id_105'):
            st.warning("⛔ Connectez-vous a l'API et selectionnez un dossier (onglet **🔑 Connexion API**).")
        elif not st.session_state.get('imp_file_balance'):
            st.warning("⛔ Importez le fichier **Balance** dans l'onglet **📁 Import fichiers**.")
    sub_rejets, sub_eraz, sub_audit, sub_matrice, sub_synthese, sub_synchro = st.tabs([
        "🚫 Rejetees", "🧹 Suppressions", "🔎 Mises a jour", "📋 Matrice", "📊 Synthese injection", "🚀 Injection"
    ])
    # Reassigner m6/m7 pour que le code existant ecrive dans les bons sous-onglets
    m6 = sub_synthese
    m7 = sub_synchro

    with sub_matrice:
        if not st.session_state.audit_matrix_105.empty:
            _status_options = ["✅", "➕", "🔄", "🗑️", "—"]
            st.session_state.audit_matrix_105 = st.data_editor(
                st.session_state.audit_matrix_105,
                use_container_width=True, hide_index=True,
                disabled=["N°", "Libellé", "TVA",
                           "_vat_id", "_patch_cat", "_patch_id", "_patch_payload", "_patch_detail",
                           "_patch_ACHAT", "_patch_VENTE", "_patch_ENTRÉE BQ", "_patch_SORTIE BQ"],
                column_config={
                    "Libellé": st.column_config.TextColumn("Libellé Compte", help="Libellé du compte (balance) — non modifiable"),
                    "LibFlux": st.column_config.TextColumn("Libellé Flux", help="Libellé catégorie/affectation — modifiable"),
                    "COMPTE": st.column_config.SelectboxColumn("COMPTE", options=_status_options, help="✅ existant, ➕ a creer, 🔄 a MAJ, 🗑️ supprimer, — ignorer"),
                    "ACHAT": st.column_config.SelectboxColumn("ACHAT", options=_status_options, help="✅ existant, ➕ a creer, 🔄 a MAJ, 🗑️ supprimer, — ignorer"),
                    "VENTE": st.column_config.SelectboxColumn("VENTE", options=_status_options, help="✅ existant, ➕ a creer, 🔄 a MAJ, 🗑️ supprimer, — ignorer"),
                    "ENTRÉE BQ": st.column_config.SelectboxColumn("ENTRÉE BQ", options=_status_options, help="✅ existant, ➕ a creer, 🔄 a MAJ, 🗑️ supprimer, — ignorer"),
                    "SORTIE BQ": st.column_config.SelectboxColumn("SORTIE BQ", options=_status_options, help="✅ existant, ➕ a creer, 🔄 a MAJ, 🗑️ supprimer, — ignorer"),
                    "_vat_id": None, "_patch_cat": None, "_patch_id": None,
                    "_patch_payload": None, "_patch_detail": None,
                    "_patch_ACHAT": None, "_patch_VENTE": None,
                    "_patch_ENTRÉE BQ": None, "_patch_SORTIE BQ": None,
                },
            )

    with sub_audit:
        if not st.session_state.audit_matrix_105.empty:
            df_audit = st.session_state.audit_matrix_105
            api_acc = st.session_state.ev_acc_105
            api_data = st.session_state.ev_data_105
            audit_rows = []

            for idx, row in df_audit.iterrows():
                code = row['N°']

                if row.get('COMPTE') == '🔄':
                    ev_account = api_acc.get(norm_piv(code))
                    target_label = str(row['Libellé']).strip().upper()
                    api_label = ev_account['label'] if ev_account else "—"
                    audit_rows.append({
                        "Catégorie": "COMPTE", "N°": code,
                        "Champ": "label",
                        "Valeur cible": target_label,
                        "Valeur API": api_label,
                    })

                for flux in ["ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
                    if row.get(flux) == '🔄':
                        patch_info = row.get(f'_patch_{flux}')
                        if isinstance(patch_info, str):
                            try: patch_info = ast.literal_eval(patch_info)
                            except: patch_info = None

                        if isinstance(patch_info, dict):
                            flux_data = api_data.get(flux, {})
                            ev_flux = None
                            fid = patch_info.get('id')
                            if fid:
                                for d in flux_data.values():
                                    if d['id'] == fid:
                                        ev_flux = d
                                        break

                            payload = patch_info.get('payload', {})
                            if isinstance(payload, str):
                                try: payload = ast.literal_eval(payload)
                                except: payload = {}

                            for field, target_val in payload.items():
                                if ev_flux:
                                    if field == "accountid":
                                        api_val = ev_flux.get('acc_id')
                                    elif field == "vataccountid":
                                        api_val = ev_flux.get('vat_id')
                                    elif field == "code":
                                        api_val = ev_flux.get('code')
                                    elif field == "label":
                                        api_val = ev_flux.get('label')
                                    else:
                                        api_val = ev_flux.get(field)
                                else:
                                    api_val = "—"

                                audit_rows.append({
                                    "Catégorie": flux, "N°": code,
                                    "Champ": field,
                                    "Valeur cible": str(target_val),
                                    "Valeur API": str(api_val) if api_val is not None else "—",
                                })

            if audit_rows:
                df_audit_display = pd.DataFrame(audit_rows)
                st.metric("Champs à mettre à jour", len(df_audit_display))
                st.dataframe(df_audit_display, use_container_width=True, hide_index=True)
            else:
                st.info("Aucune mise à jour (🔄) détectée dans la matrice")
        else:
            st.info("Lancez l'analyse d'abord (onglet Balance)")

    with sub_rejets:
        if isinstance(st.session_state.rejets_105, pd.DataFrame) and not st.session_state.rejets_105.empty:
            st.dataframe(st.session_state.rejets_105, use_container_width=True)
        else:
            st.info("Aucun rejet")

    with sub_eraz:
        has_headers = bool(st.session_state.token_headers_105)
        prot = st.session_state.prot_105

        if not has_headers:
            st.warning("Connectez-vous à l'API Evoliz d'abord")
        elif not prot:
            st.info("Lancez l'analyse d'abord (onglet Balance)")
        else:
            prot_comptes = prot.get("comptes", set()) if isinstance(prot, dict) else prot
            prot_flux_ids = prot.get("flux_ids", set()) if isinstance(prot, dict) else set()
            orphans_by_cat = {}

            orphans_by_cat["COMPTE"] = [
                {"pivot": p, "id": d['id'], "Code": d['code'], "Libellé": d['label']}
                for p, d in st.session_state.ev_acc_105.items() if p not in prot_comptes
            ]
            for flux in ["ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
                seen_ids = set()
                orphans = []
                for p, d in st.session_state.ev_data_105.get(flux, {}).items():
                    if d['id'] not in prot_flux_ids and d['id'] not in seen_ids:
                        orphans.append({"pivot": p, "id": d['id'], "Code": d['code'], "Libellé": d['label']})
                        seen_ids.add(d['id'])
                orphans_by_cat[flux] = orphans

            # Partage avec la synthese
            st.session_state["_orphans_total_by_cat"] = {c: len(v) for c, v in orphans_by_cat.items()}

            # Init des exclusions de suppression
            if "_skip_delete" not in st.session_state:
                st.session_state._skip_delete = set()

            total_orphans = sum(len(v) for v in orphans_by_cat.values())
            st.subheader(f"🧹 {total_orphans} éléments orphelins")

            for cat, items in orphans_by_cat.items():
                if items:
                    with st.expander(f"{cat} — {len(items)} orphelins", expanded=False):
                        for it in items:
                            _key = (cat, it['id'])
                            _cb = st.checkbox(
                                f"Supprimer `{it['Code']}` — {it['Libellé']}",
                                value=True,
                                key=f"cb_del_{cat}_{it['id']}",
                            )
                            if not _cb:
                                st.session_state._skip_delete.add(_key)
                            else:
                                st.session_state._skip_delete.discard(_key)

            # Compter APRÈS les checkboxes pour avoir l'état à jour
            _filtered_eraz = {}
            for cat, items in orphans_by_cat.items():
                _filtered_eraz[cat] = [it for it in items if (cat, it['id']) not in st.session_state._skip_delete]
            st.session_state.eraz_items = _filtered_eraz

            _n_effective = sum(len(v) for v in _filtered_eraz.values())
            _n_skipped = total_orphans - _n_effective

            # --- Reintegrer les orphelins gardes dans la matrice ---
            if _n_skipped > 0 and not st.session_state.audit_matrix_105.empty:
                _df_m = st.session_state.audit_matrix_105
                _existing_codes = set(_df_m['N°'].astype(str).str.strip())
                _added_to_matrice = 0
                for cat, items in orphans_by_cat.items():
                    for it in items:
                        if (cat, it['id']) in st.session_state._skip_delete:
                            # Cet orphelin est garde -> l'ajouter a la matrice s'il n'y est pas
                            _code = str(it['Code']).strip()
                            if _code not in _existing_codes:
                                _new_row = {
                                    "Sync": False, "N°": _code, "Libellé": it['Libellé'],
                                    "LibFlux": it['Libellé'], "TVA": "—", "_vat_id": None,
                                    "COMPTE": "✅" if cat == "COMPTE" else "—",
                                    "ACHAT": "✅" if cat == "ACHAT" else "—",
                                    "VENTE": "✅" if cat == "VENTE" else "—",
                                    "ENTRÉE BQ": "✅" if cat == "ENTRÉE BQ" else "—",
                                    "SORTIE BQ": "✅" if cat == "SORTIE BQ" else "—",
                                }
                                _df_m = pd.concat([_df_m, pd.DataFrame([_new_row])], ignore_index=True)
                                _existing_codes.add(_code)
                                _added_to_matrice += 1
                            else:
                                # Le code existe deja -> juste mettre le statut ✅ pour cette categorie
                                _idx = _df_m[_df_m['N°'].astype(str).str.strip() == _code].index
                                if len(_idx) > 0 and _df_m.at[_idx[0], cat] in ("—", ""):
                                    _df_m.at[_idx[0], cat] = "✅"
                                    _added_to_matrice += 1
                if _added_to_matrice > 0:
                    st.session_state.audit_matrix_105 = _df_m

            if _n_skipped:
                st.info(f"⏭️ {_n_skipped} élément(s) conserve(s) et reintegre(s) dans la matrice — **{_n_effective}** suppression(s) effectives.")
            else:
                st.success(f"✅ {_n_effective} suppression(s) prévues.")

            if not total_orphans:
                st.success("Aucun orphelin détecté")

# --- m6 : Synthèse (vérification avant synchro) ---
with m6:
    if not st.session_state.audit_matrix_105.empty:
        df_m = st.session_state.audit_matrix_105
        eraz = st.session_state.eraz_counts
        eraz_items = st.session_state.eraz_items

        # Récupérer les ajustements manuels précédents (persistés en session)
        if "_synth_modif" not in st.session_state:
            st.session_state._synth_modif = {c: 0 for c in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]}

        # --- Decompte automatique des decisions manuelles ---
        # 1) Orphelins gardes (decoches dans l'onglet Suppressions) -> + (n'a pas ete supprime)
        _skip_del = st.session_state.get("_skip_delete", set())
        orph_kept_par_cat = {c: 0 for c in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]}
        for _key in _skip_del:
            if isinstance(_key, tuple) and len(_key) == 2:
                _cat = _key[0]
                if _cat in orph_kept_par_cat:
                    orph_kept_par_cat[_cat] += 1

        # 2) Creations desactivees dans le detail synthese -> - (n'a pas ete cree)
        _skip_cr = st.session_state.get("_skip_create", set())
        skip_create_par_cat = {c: 0 for c in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]}
        for _key in _skip_cr:
            if isinstance(_key, tuple) and len(_key) >= 1:
                _cat = _key[0]
                if _cat in skip_create_par_cat:
                    skip_create_par_cat[_cat] += 1

        # 3) Total orphelins par categorie (peuple par l'onglet Suppressions, fallback 0)
        _orph_total_par_cat = st.session_state.get("_orphans_total_by_cat", {})

        stats = []
        for cat in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
            lus = len(st.session_state.ev_acc_105) if cat == "COMPTE" else count_unique(st.session_state.ev_data_105.get(cat, {}))
            crees = len(df_m[df_m[cat] == '➕'])
            maj = len(df_m[df_m[cat] == '🔄'])
            # "A supprimer" = TOUS les orphelins (independamment de leur sort)
            a_supprimer = _orph_total_par_cat.get(cat, len(eraz_items.get(cat, [])))
            supprimes = eraz.get(cat, 0)
            en_matrice = len(df_m[df_m[cat].isin(['✅', '➕', '🔄'])])

            # Decisions manuelles signees
            plus_n = orph_kept_par_cat.get(cat, 0)        # + : items conserves (orphelins gardes)
            minus_n = skip_create_par_cat.get(cat, 0)     # - : creations desactivees
            modif_input = st.session_state._synth_modif.get(cat, 0)
            net_modif = plus_n - minus_n + modif_input

            modif_str = f"+{plus_n} / -{minus_n}"
            if modif_input:
                modif_str += f" ({'+' if modif_input>0 else ''}{modif_input})"

            # Etat predit de l'API apres synchro
            attendu = lus + crees - a_supprimer + net_modif
            # Etat intentionnel (matrice + decisions manuelles)
            cible = en_matrice + plus_n - minus_n
            coherent = attendu == cible

            stats.append({
                "Catégorie": cat,
                "📖 Lus API": lus,
                "➕ À créer": crees,
                "🔄 À maj": maj,
                "🗑️ À supprimer": a_supprimer,
                "✅ Supprimés": supprimes,
                "🛡️ Modifications manuelles": modif_str,
                "🔧 Ajust.": modif_input,
                "= Attendu": attendu,
                "🎯 Cible": cible,
                "📊 Matrice": en_matrice,
                "✔️ Cohérent": "✅" if coherent else "❌",
            })

        df_stats = pd.DataFrame(stats)
        edited_stats = st.data_editor(
            df_stats,
            use_container_width=True, hide_index=True,
            disabled=["Catégorie", "📖 Lus API", "➕ À créer", "🔄 À maj",
                       "🗑️ À supprimer", "✅ Supprimés", "🛡️ Modifications manuelles",
                       "= Attendu", "🎯 Cible", "📊 Matrice", "✔️ Cohérent"],
            column_config={
                "🛡️ Modifications manuelles": st.column_config.TextColumn(
                    "🛡️ Modifications manuelles",
                    help="+X : orphelins conserves (decoches dans Suppressions). -Y : creations desactivees (decochees dans le detail). (Z) : ajustement manuel.",
                ),
                "🔧 Ajust.": st.column_config.NumberColumn(
                    "🔧 Ajust.",
                    help="Ajustement manuel additionnel (modifications hors app non detectees)",
                    step=1, format="%d",
                ),
                "🎯 Cible": st.column_config.NumberColumn(
                    "🎯 Cible",
                    help="Etat intentionnel = Matrice + decisions manuelles",
                    format="%d",
                ),
            },
        )
        # Persister les ajustements et recalculer la cohérence en temps réel
        _modif_changed = False
        for _, row in edited_stats.iterrows():
            cat = row["Catégorie"]
            new_modif = int(row["🔧 Ajust."])
            if st.session_state._synth_modif.get(cat, 0) != new_modif:
                st.session_state._synth_modif[cat] = new_modif
                _modif_changed = True
        if _modif_changed:
            st.rerun()

        # --- Visualisation par catégorie : existants + à créer (avec possibilité de désactiver) ---
        st.divider()
        st.subheader("🔍 Détail par catégorie")

        # Init session state pour les créations désactivées
        if "_skip_create" not in st.session_state:
            st.session_state._skip_create = set()  # ensemble de (cat, code) à ne pas créer

        _cat_labels = {
            "COMPTE": "📖 Comptes comptables",
            "ACHAT": "📥 Classifications achat",
            "VENTE": "📤 Classifications vente",
            "ENTRÉE BQ": "📤 Affectations entrées bancaires",
            "SORTIE BQ": "📥 Affectations sorties bancaires",
        }
        for _cat, _cat_label in _cat_labels.items():
            # Existants dans Evoliz
            if _cat == "COMPTE":
                _existing = []
                _seen_ids = set()
                for _v in st.session_state.ev_acc_105.values():
                    if _v['id'] in _seen_ids: continue
                    _seen_ids.add(_v['id'])
                    _existing.append({"Code": _v.get('code', ''), "Libellé": _v.get('label', '')})
            else:
                _flux_data = st.session_state.ev_data_105.get(_cat, {})
                _seen_ids = set()
                _existing = []
                for _v in _flux_data.values():
                    if _v['id'] in _seen_ids: continue
                    _seen_ids.add(_v['id'])
                    _existing.append({"Code": _v.get('code', ''), "Libellé": _v.get('label', ''), "Compte": _v.get('acc_id', '')})

            # À créer (➕) depuis la matrice
            _to_create = []
            for _idx, _row in df_m.iterrows():
                if _row.get(_cat) == '➕':
                    _code = _row.get('N°', '')
                    _label = _row.get('LibFlux', _row.get('Libellé', ''))
                    _key = (_cat, _code, _label)
                    _to_create.append({"Code": _code, "Libellé": _label, "_key": _key, "_idx": _idx})

            # À mettre à jour (🔄) depuis la matrice
            _to_update = []
            for _idx, _row in df_m.iterrows():
                if _row.get(_cat) == '🔄':
                    _to_update.append({"Code": _row.get('N°', ''), "Libellé": _row.get('LibFlux', _row.get('Libellé', ''))})

            _n_exist = len(_existing)
            _n_create = len(_to_create)
            _n_update = len(_to_update)
            _summary = f"{_n_exist} existant(s)"
            if _n_create: _summary += f", {_n_create} à créer"
            if _n_update: _summary += f", {_n_update} à MAJ"

            with st.expander(f"{_cat_label} — {_summary}", expanded=False):
                # Existants
                if _existing:
                    st.caption(f"✅ **{_n_exist} existant(s) dans Evoliz**")
                    st.dataframe(pd.DataFrame(_existing).sort_values("Code"), use_container_width=True, hide_index=True)

                # MAJ
                if _to_update:
                    st.caption(f"🔄 **{_n_update} à mettre à jour**")
                    st.dataframe(pd.DataFrame(_to_update).sort_values("Code"), use_container_width=True, hide_index=True)

                # Créations avec cases à cocher
                if _to_create:
                    st.caption(f"➕ **{_n_create} à créer** — décochez pour désactiver la création")
                    for _item in _to_create:
                        _checked = _item["_key"] not in st.session_state._skip_create
                        _cb = st.checkbox(
                            f"`{_item['Code']}` — {_item['Libellé']}",
                            value=_checked,
                            key=f"cb_create_{_cat}_{_item['_idx']}",
                        )
                        if not _cb and _item["_key"] not in st.session_state._skip_create:
                            st.session_state._skip_create.add(_item["_key"])
                        elif _cb and _item["_key"] in st.session_state._skip_create:
                            st.session_state._skip_create.discard(_item["_key"])

                if not _existing and not _to_create and not _to_update:
                    st.caption("Aucune donnée.")

        # Résumé des créations désactivées
        _n_skipped = len(st.session_state._skip_create)
        if _n_skipped:
            st.warning(f"⚠️ {_n_skipped} création(s) désactivée(s). Ces éléments ne seront pas injectés lors de la synchro.")

    else:
        st.info("Lancez l'analyse d'abord (onglet Balance)")

# --- m7 : Synchro (DELETE + PATCH + CREATE, dernier onglet) ---
with m7:
    df_sync = st.session_state.audit_matrix_105
    has_headers = bool(st.session_state.token_headers_105)

    if not df_sync.empty and has_headers:
        # Option de filtrage
        sync_scope = st.radio("Périmètre de synchronisation",
                              ["🌐 Tout", "🏷️ Ventes uniquement (comptes 7xx)"],
                              horizontal=True, key="sync_scope")
        ventes_only = sync_scope.startswith("🏷️")

        to_sync = df_sync[df_sync['Sync'] == True]
        if ventes_only:
            to_sync = to_sync[to_sync['N°'].astype(str).str.startswith('7')]

        new_accounts = to_sync[to_sync['COMPTE'] == '➕']
        patch_accounts = to_sync[to_sync['COMPTE'] == '🔄']
        # Suppressions demandees directement dans la matrice (🗑️)
        del_accounts_matrice = to_sync[to_sync['COMPTE'] == '🗑️']
        if ventes_only:
            new_flux = {f: (to_sync[to_sync[f] == '➕'] if f == "VENTE" else pd.DataFrame()) for f in FLUX_ENDPOINTS}
            patch_flux = {f: (to_sync[to_sync[f] == '🔄'] if f == "VENTE" else pd.DataFrame()) for f in FLUX_ENDPOINTS}
            del_flux_matrice = {f: (to_sync[to_sync[f] == '🗑️'] if f == "VENTE" else pd.DataFrame()) for f in FLUX_ENDPOINTS}
        else:
            new_flux = {f: to_sync[to_sync[f] == '➕'] for f in FLUX_ENDPOINTS}
            patch_flux = {f: to_sync[to_sync[f] == '🔄'] for f in FLUX_ENDPOINTS}
            del_flux_matrice = {f: to_sync[to_sync[f] == '🗑️'] for f in FLUX_ENDPOINTS}

        # Filtrer les créations désactivées par l'utilisateur dans la Synthèse
        _skip = st.session_state.get("_skip_create", set())
        if _skip:
            new_accounts = new_accounts[~new_accounts.apply(
                lambda r: ("COMPTE", r['N°'], r.get('LibFlux', r.get('Libellé', ''))) in _skip, axis=1)]
            for _f in new_flux:
                if not new_flux[_f].empty:
                    new_flux[_f] = new_flux[_f][~new_flux[_f].apply(
                        lambda r, _cat=_f: (_cat, r['N°'], r.get('LibFlux', r.get('Libellé', ''))) in _skip, axis=1)]

        total_create = len(new_accounts) + sum(len(v) for v in new_flux.values())
        total_patch = len(patch_accounts) + sum(len(v) for v in patch_flux.values())

        eraz_items = st.session_state.eraz_items
        if ventes_only:
            # ERAZ : ne supprimer que les comptes 7xx et les flux VENTE
            eraz_items = {
                "COMPTE": [i for i in eraz_items.get("COMPTE", []) if str(i.get('Code', '')).startswith('7')],
                "VENTE": eraz_items.get("VENTE", []),
                "ACHAT": [], "ENTRÉE BQ": [], "SORTIE BQ": [],
            }
        total_delete = sum(len(v) for v in eraz_items.values()) if eraz_items else 0
        # Ajouter les suppressions demandees dans la matrice (🗑️)
        _del_matrice_count = len(del_accounts_matrice) + sum(len(v) for v in del_flux_matrice.values())
        total_delete += _del_matrice_count

        st.subheader("📋 Résumé des opérations")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("🗑️ Suppressions", total_delete)
        c2.metric("🔄 Mises à jour", total_patch)
        c3.metric("➕ Créations", total_create)
        c4.metric("Total", total_delete + total_patch + total_create)

        total_all = total_delete + total_patch + total_create
        if total_all > 0 and st.button("⚡ LANCER LA SYNCHRONISATION", type="primary", use_container_width=True):
            headers = st.session_state.token_headers_105
            cid = st.session_state.company_id_105
            log = []
            progress = st.progress(0)
            done = [0]  # list pour mutation dans les threads
            eraz_counts = {"COMPTE": 0, "ACHAT": 0, "VENTE": 0, "ENTRÉE BQ": 0, "SORTIE BQ": 0}
            MAX_WORKERS = 5  # 5 workers pour rester sous 100 req/min avec marge

            # Rate limiter : max 90 requêtes par fenêtre de 60s (marge de sécurité)
            _rate_lock = threading.Lock()
            _rate_timestamps = []
            RATE_LIMIT = 90
            RATE_WINDOW = 60

            def rate_wait():
                """Attend si nécessaire pour ne pas dépasser le rate limit API."""
                with _rate_lock:
                    now = time.time()
                    # Purger les timestamps hors fenêtre
                    while _rate_timestamps and _rate_timestamps[0] < now - RATE_WINDOW:
                        _rate_timestamps.pop(0)
                    if len(_rate_timestamps) >= RATE_LIMIT:
                        wait = _rate_timestamps[0] + RATE_WINDOW - now + 0.1
                        if wait > 0:
                            time.sleep(wait)
                        # Re-purger après l'attente
                        now = time.time()
                        while _rate_timestamps and _rate_timestamps[0] < now - RATE_WINDOW:
                            _rate_timestamps.pop(0)
                    _rate_timestamps.append(time.time())

            def safe_dict(val):
                if isinstance(val, dict): return val
                if isinstance(val, str):
                    try: return ast.literal_eval(val)
                    except: return None
                return None

            def update_progress():
                done[0] += 1
                progress.progress(min(done[0] / total_all, 1.0))

            # Wrappers avec rate limiting
            def _del_rate(cat, item_id, hdrs):
                rate_wait()
                return delete_evoliz_item(cat, item_id, hdrs, company_id=cid)

            def _patch_rate(cat, item_id, payload, hdrs, company_id=None):
                rate_wait()
                return patch_evoliz_item(cat, item_id, payload, hdrs, company_id=company_id)

            def _add_account_rate(code, label, hdrs):
                rate_wait()
                return inject_account(code, label, hdrs)

            def _add_flux_rate(flux_type, code, label, hdrs, vat_id=None, acc_id=None, compte_code=None):
                rate_wait()
                # Appliquer le taux de TVA pour les classifications d'achat (comptes 2xx/6xx)
                _vr = None
                if flux_type == "ACHAT" and compte_code and str(compte_code).startswith(('2', '6')):
                    _vr = tva_achat_rate
                return inject_flux(flux_type, code, label, hdrs, vat_id=vat_id, acc_id=acc_id, vat_rate=_vr, company_id=cid)

            # --- 1. DELETE orphelins (parallèle par paquets) ---
            if eraz_items:
                del_tasks = []
                for cat, items in eraz_items.items():
                    for item in items:
                        del_tasks.append((cat, item))
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                    futures = {pool.submit(_del_rate, t[0], t[1]['id'], headers): t for t in del_tasks}
                    for f in as_completed(futures):
                        cat, item = futures[f]
                        ok, msg = f.result()
                        log.append({"Action": "🗑️ DEL", "Type": cat, "Code": item['Code'],
                                    "Libellé": item['Libellé'], "Résultat": "✅" if ok else "❌", "Détail": msg})
                        if ok:
                            eraz_counts[cat] += 1
                        update_progress()

            st.session_state.eraz_counts = eraz_counts

            # --- 1bis. DELETE depuis la matrice (🗑️ selectionne par l'utilisateur) ---
            # Comptes marques 🗑️
            for _, row in del_accounts_matrice.iterrows():
                _code = row['N°']
                _ev_a = st.session_state.ev_acc_105.get(norm_piv(_code))
                if _ev_a:
                    rate_wait()
                    ok, msg = delete_evoliz_item("COMPTE", _ev_a['id'], headers, company_id=cid)
                    log.append({"Action": "🗑️ DEL", "Type": "COMPTE", "Code": _code,
                                "Libellé": row.get('Libellé', ''), "Résultat": "✅" if ok else "❌", "Détail": msg})
                update_progress()
            # Flux marques 🗑️
            for flux, df_del in del_flux_matrice.items():
                for _, row in df_del.iterrows():
                    _code = row['N°']
                    _label = row.get('LibFlux', row.get('Libellé', ''))
                    _flux_data = st.session_state.ev_data_105.get(flux, {})
                    # Chercher l'id dans les donnees API
                    _item_id = None
                    for _p, _d in _flux_data.items():
                        if norm_piv(_d.get('label', '')) == norm_piv(_label) or norm_piv(_d.get('code', '')) == norm_piv(_code):
                            _item_id = _d['id']; break
                    if _item_id:
                        rate_wait()
                        ok, msg = delete_evoliz_item(flux, _item_id, headers, company_id=cid)
                        log.append({"Action": "🗑️ DEL", "Type": flux, "Code": _code,
                                    "Libellé": _label, "Résultat": "✅" if ok else "❌", "Détail": msg})
                    else:
                        log.append({"Action": "🗑️ DEL", "Type": flux, "Code": _code,
                                    "Libellé": _label, "Résultat": "❌", "Détail": "ID non trouve dans API"})
                    update_progress()

            # --- 2. UPDATE comptes (DELETE + POST, séquentiel) ---
            patch_accounts = to_sync[to_sync['COMPTE'] == '🔄'] if 'COMPTE' in to_sync.columns else pd.DataFrame()
            for _, row in patch_accounts.iterrows():
                p_id = row.get('_patch_id')
                p_detail = row.get('_patch_detail', '')
                if p_id:
                    rate_wait()
                    ok_del, _ = delete_evoliz_item("COMPTE", p_id, headers)
                    if ok_del:
                        rate_wait()
                        ok_add, resp = inject_account(row['N°'], row['Libellé'], headers)
                        log.append({"Action": "🔄 UPD", "Type": "COMPTE", "Code": row['N°'],
                                    "Libellé": p_detail, "Résultat": "✅" if ok_add else "❌",
                                    "Détail": f"DEL+POST: {str(resp)[:100]}"})
                    else:
                        log.append({"Action": "🔄 UPD", "Type": "COMPTE", "Code": row['N°'],
                                    "Libellé": p_detail, "Résultat": "❌", "Détail": "Échec DELETE"})
                update_progress()

            # Rafraîchir les comptes après DEL+POST pour avoir les nouveaux id
            if not patch_accounts.empty:
                time.sleep(1)
                st.session_state.ev_acc_105 = fetch_evoliz_data("accounts", headers, company_id=cid)

            # --- 3. PATCH flux (parallèle) ---
            patch_tasks = []
            acc_lookup = st.session_state.ev_acc_105
            for flux in FLUX_ENDPOINTS:
                for _, row in patch_flux[flux].iterrows():
                    patch_info = safe_dict(row.get(f'_patch_{flux}'))
                    if patch_info:
                        payload = safe_dict(patch_info.get('payload')) or patch_info.get('payload')
                        # Résoudre l'accountid à jour (le compte a pu être DEL+POST)
                        if isinstance(payload, dict) and 'accountid' in payload:
                            ev_a = acc_lookup.get(norm_piv(row['N°']))
                            if ev_a:
                                payload['accountid'] = ev_a['id']
                        # Ajouter vat_rate pour les classifications d'achat (comptes 2xx/6xx)
                        if isinstance(payload, dict) and flux == "ACHAT" and str(row['N°']).startswith(('2', '6')):
                            payload['vat_rate'] = float(tva_achat_rate)
                        patch_tasks.append((flux, row['N°'], patch_info, payload))
            if patch_tasks:
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                    futures = {pool.submit(_patch_rate, t[0], t[2]['id'], t[3], headers, company_id=cid): t for t in patch_tasks}
                    for f in as_completed(futures):
                        flux, code, pinfo, _ = futures[f]
                        ok, resp = f.result()
                        log.append({"Action": "🔄 UPD", "Type": flux, "Code": code,
                                    "Libellé": pinfo.get('detail', ''), "Résultat": "✅" if ok else "❌", "Détail": str(resp)[:120]})
                        update_progress()

            # --- 4. CREATE comptes (parallèle) ---
            if not new_accounts.empty:
                acc_tasks = [(row['N°'], row['Libellé']) for _, row in new_accounts.iterrows()]
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                    futures = {pool.submit(_add_account_rate, t[0], t[1], headers): t for t in acc_tasks}
                    for f in as_completed(futures):
                        code, label = futures[f]
                        ok, resp = f.result()
                        log.append({"Action": "➕ ADD", "Type": "COMPTE", "Code": code,
                                    "Libellé": label, "Résultat": "✅" if ok else "❌", "Détail": str(resp)[:120]})
                        update_progress()

            # Rafraîchir les comptes pour résoudre les accountid des flux à créer
            time.sleep(2)  # Pause pour laisser l'API digérer les créations
            with st.spinner("Rafraîchissement des comptes..."):
                st.session_state.ev_acc_105 = fetch_evoliz_data("accounts", headers, company_id=cid)

            # --- 5. CREATE flux (parallèle) ---
            flux_tasks = []
            acc_lookup = st.session_state.ev_acc_105
            for flux, df_f in new_flux.items():
                for _, row in df_f.iterrows():
                    label_for_flux = row.get('LibFlux', row['Libellé'])
                    row_vat = row.get('_vat_id') if flux == "ACHAT" else None
                    # Résoudre accountid depuis les comptes rafraîchis
                    row_acc_id = None
                    ev_a = acc_lookup.get(norm_piv(row['N°']))
                    if ev_a:
                        row_acc_id = ev_a['id']
                    else:
                        # Fallback : chercher par code brut dans toutes les valeurs
                        code_str = str(row['N°']).strip()
                        for a in acc_lookup.values():
                            if a.get('code', '').strip() == code_str:
                                row_acc_id = a['id']
                                break
                    flux_tasks.append((flux, row['N°'], label_for_flux, row_vat, row_acc_id, row['N°']))
            if flux_tasks:
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                    futures = {pool.submit(_add_flux_rate, t[0], t[1], t[2], headers, vat_id=t[3], acc_id=t[4], compte_code=t[5]): t for t in flux_tasks}
                    for f in as_completed(futures):
                        flux, code, lbl, _, _, _ = futures[f]
                        ok, resp = f.result()
                        log.append({"Action": "➕ ADD", "Type": flux, "Code": code,
                                    "Libellé": lbl, "Résultat": "✅" if ok else "❌", "Détail": str(resp)[:120]})
                        update_progress()

            st.session_state.sync_log = log
            ok_count = sum(1 for l in log if l['Résultat'] == '✅')
            st.success(f"Terminé : {ok_count}/{len(log)} opérations réussies")

            with st.spinner("Rafraîchissement des données API..."):
                st.session_state.ev_acc_105 = fetch_evoliz_data("accounts", headers, company_id=cid)
                st.session_state.ev_data_105 = {
                    "ACHAT": fetch_evoliz_data("purchase-classifications", headers, company_id=cid),
                    "VENTE": fetch_evoliz_data("sale-classifications", headers, company_id=cid),
                    "ENTRÉE BQ": fetch_evoliz_data("sale-affectations", headers, company_id=cid),
                    "SORTIE BQ": fetch_evoliz_data("purchase-affectations", headers, company_id=cid),
                }

        if st.session_state.sync_log:
            df_log = pd.DataFrame(st.session_state.sync_log)

            # --- CR synthétique ---
            st.subheader("📊 Compte-rendu d'exécution")
            ok_total = len(df_log[df_log['Résultat'] == '✅'])
            ko_total = len(df_log[df_log['Résultat'] == '❌'])
            c1, c2, c3 = st.columns(3)
            c1.metric("Total opérations", len(df_log))
            c2.metric("✅ Réussies", ok_total)
            c3.metric("❌ Échouées", ko_total)

            # Tableau croisé Action x Type
            cr_data = []
            for action in ['🗑️ DEL', '🔄 UPD', '➕ ADD']:
                row_cr = {"Action": action}
                for cat in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
                    mask = (df_log['Action'] == action) & (df_log['Type'] == cat)
                    ok = len(df_log[mask & (df_log['Résultat'] == '✅')])
                    ko = len(df_log[mask & (df_log['Résultat'] == '❌')])
                    if ok + ko > 0:
                        row_cr[cat] = f"{ok}✅ {ko}❌" if ko else f"{ok}✅"
                    else:
                        row_cr[cat] = "—"
                cr_data.append(row_cr)
            st.table(pd.DataFrame(cr_data))

            # --- Vérification post-exécution ---
            st.subheader("🔎 Vérification post-exécution (GET vs Matrice)")
            # Refresh des donnees API apres sync : lecture a nouveau des comptes + flux
            _h_refresh = st.session_state.get("token_headers_105", {})
            _cid_refresh = st.session_state.get("company_id_105")
            if _h_refresh and _cid_refresh:
                with st.spinner("Relecture des donnees API post-sync..."):
                    st.session_state.ev_acc_105 = fetch_evoliz_data("accounts", _h_refresh, company_id=_cid_refresh)
                    for _flux_key, _endpoint in FLUX_ENDPOINTS.items():
                        st.session_state.ev_data_105[_flux_key] = fetch_evoliz_data(_endpoint, _h_refresh, company_id=_cid_refresh)
            df_m = st.session_state.audit_matrix_105
            api_acc = st.session_state.ev_acc_105
            api_data = st.session_state.ev_data_105

            verif = []
            for cat in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
                api_src = api_acc if cat == "COMPTE" else api_data.get(cat, {})
                api_count = len(api_src) if cat == "COMPTE" else count_unique(api_src)
                matrice_attendus = len(df_m[df_m[cat].isin(['✅', '➕', '🔄'])])

                if cat == "COMPTE":
                    # Comptes : comparaison par pivot code
                    api_by_id = {d['id']: norm_piv(d['code']) for d in api_src.values()}
                    api_pivots_unique = set(api_by_id.values())
                    matrice_pivots = set(norm_piv(c) for c in df_m[df_m[cat].isin(['✅', '➕', '🔄'])]['N°'])
                else:
                    # Flux : dédupliquer par id, comparer par pivot label (comme la matrice)
                    seen_ids = set()
                    api_pivots_unique = set()
                    for p, d in api_src.items():
                        if d['id'] not in seen_ids:
                            seen_ids.add(d['id'])
                            # Utiliser le pivot label pour comparer (même logique que la matrice)
                            pivot_label = norm_piv(d['label'])
                            api_pivots_unique.add(pivot_label)
                    matrice_pivots = set(
                        norm_piv(clean_label_tva(r['Libellé'], r['N°'], fusion_tva))
                        for _, r in df_m[df_m[cat].isin(['✅', '➕', '🔄'])].iterrows()
                    )

                presents = matrice_pivots & api_pivots_unique
                manquants = matrice_pivots - api_pivots_unique
                excedents = api_pivots_unique - matrice_pivots

                status = "✅" if not manquants and not excedents else "❌"
                verif.append({
                    "Catégorie": cat,
                    "API (GET)": api_count,
                    "Matrice": matrice_attendus,
                    "Présents": len(presents),
                    "Manquants": len(manquants),
                    "Excédents": len(excedents),
                    "Statut": status,
                })
            st.table(pd.DataFrame(verif))

            # Détail des écarts si présents
            all_manquants = []
            all_excedents = []
            for cat in ["COMPTE", "ACHAT", "VENTE", "ENTRÉE BQ", "SORTIE BQ"]:
                api_src = api_acc if cat == "COMPTE" else api_data.get(cat, {})

                if cat == "COMPTE":
                    api_labels_by_pivot = {norm_piv(d['code']): d for d in api_src.values()}
                    matrice_pivots = {norm_piv(c): c for c in df_m[df_m[cat].isin(['✅', '➕', '🔄'])]['N°']}
                else:
                    # Dédupliquer par id, indexer par pivot label
                    seen_ids = set()
                    api_labels_by_pivot = {}
                    for d in api_src.values():
                        if d['id'] not in seen_ids:
                            seen_ids.add(d['id'])
                            api_labels_by_pivot[norm_piv(d['label'])] = d
                    matrice_pivots = {
                        norm_piv(clean_label_tva(r['Libellé'], r['N°'], fusion_tva)): r.get('LibFlux', r['Libellé'])
                        for _, r in df_m[df_m[cat].isin(['✅', '➕', '🔄'])].iterrows()
                    }

                for p, ref in matrice_pivots.items():
                    if p and p not in api_labels_by_pivot:
                        all_manquants.append({"Catégorie": cat, "Pivot": p, "Référence": ref})
                for p, d in api_labels_by_pivot.items():
                    if p not in matrice_pivots:
                        all_excedents.append({"Catégorie": cat, "Code": d.get('code', ''), "Libellé": d.get('label', '')})

            if all_manquants:
                with st.expander(f"⚠️ {len(all_manquants)} éléments manquants dans l'API"):
                    st.dataframe(pd.DataFrame(all_manquants), use_container_width=True)
            if all_excedents:
                with st.expander(f"⚠️ {len(all_excedents)} éléments excédentaires dans l'API"):
                    st.dataframe(pd.DataFrame(all_excedents), use_container_width=True)
            if not all_manquants and not all_excedents:
                st.success("Parfaite concordance entre l'API et la matrice")

            # --- Logs détaillés ---
            st.subheader("📋 Logs détaillés")
            filtre = st.radio("Filtrer", ["Tout", "✅ Réussies", "❌ Échouées"], horizontal=True, key="log_filter")
            if filtre == "✅ Réussies":
                df_show = df_log[df_log['Résultat'] == '✅']
            elif filtre == "❌ Échouées":
                df_show = df_log[df_log['Résultat'] == '❌']
            else:
                df_show = df_log
            st.dataframe(df_show, use_container_width=True)

            # Export CSV
            csv = df_log.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Exporter les logs (CSV)", csv, "sync_log.csv", "text/csv")
    elif not has_headers:
        st.warning("Connectez-vous à l'API Evoliz d'abord (onglet Connexion API)")
    else:
        st.info("Lancez l'analyse d'abord (onglet Balance)")


# =========================================================
# BASCULE MEG -> EVOLIZ
# =========================================================

ISO2_MAP = {
    "AFGHANISTAN": "AF", "ALBANIE": "AL", "ALGERIE": "DZ", "ALLEMAGNE": "DE",
    "ANDORRE": "AD", "ANGOLA": "AO", "ARABIE SAOUDITE": "SA", "ARGENTINE": "AR",
    "AUSTRALIE": "AU", "AUTRICHE": "AT", "BELGIQUE": "BE", "BRESIL": "BR",
    "BULGARIE": "BG", "BURKINA FASO": "BF", "CAMEROUN": "CM", "CANADA": "CA",
    "CHILI": "CL", "CHINE": "CN", "COLOMBIE": "CO", "COREE DU SUD": "KR",
    "COSTA RICA": "CR", "COTE D'IVOIRE": "CI", "CROATIE": "HR", "CUBA": "CU",
    "DANEMARK": "DK", "EGYPTE": "EG", "EMIRATS ARABES UNIS": "AE", "ESPAGNE": "ES",
    "ESTONIE": "EE", "ETATS-UNIS": "US", "FINLANDE": "FI", "FRANCE": "FR",
    "GABON": "GA", "GEORGIE": "GE", "GHANA": "GH", "GRECE": "GR",
    "GUADELOUPE": "GP", "GUATEMALA": "GT", "GUINEE": "GN", "GUYANE FRANCAISE": "GF",
    "HAITI": "HT", "HONDURAS": "HN", "HONG KONG": "HK", "HONGRIE": "HU",
    "INDE": "IN", "INDONESIE": "ID", "IRAK": "IQ", "IRAN": "IR", "IRLANDE": "IE",
    "ISLANDE": "IS", "ISRAEL": "IL", "ITALIE": "IT", "JAMAIQUE": "JM", "JAPON": "JP",
    "JORDANIE": "JO", "KAZAKHSTAN": "KZ", "KENYA": "KE", "KOWEIT": "KW",
    "LAOS": "LA", "LETTONIE": "LV", "LIBAN": "LB", "LIBYE": "LY",
    "LIECHTENSTEIN": "LI", "LITUANIE": "LT", "LUXEMBOURG": "LU", "MACEDOINE": "MK",
    "MADAGASCAR": "MG", "MALAISIE": "MY", "MALI": "ML", "MALTE": "MT",
    "MAROC": "MA", "MARTINIQUE": "MQ", "MAURICE": "MU", "MAURITANIE": "MR",
    "MAYOTTE": "YT", "MEXIQUE": "MX", "MOLDAVIE": "MD", "MONACO": "MC",
    "MONGOLIE": "MN", "MONTENEGRO": "ME", "MOZAMBIQUE": "MZ", "NAMIBIE": "NA",
    "NEPAL": "NP", "NICARAGUA": "NI", "NIGER": "NE", "NIGERIA": "NG",
    "NORVEGE": "NO", "NOUVELLE-ZELANDE": "NZ", "OMAN": "OM", "OUGANDA": "UG",
    "PAKISTAN": "PK", "PALESTINE": "PS", "PANAMA": "PA", "PARAGUAY": "PY",
    "PAYS-BAS": "NL", "PEROU": "PE", "PHILIPPINES": "PH", "POLOGNE": "PL",
    "PORTUGAL": "PT", "QATAR": "QA", "REPUBLIQUE TCHEQUE": "CZ", "REUNION": "RE",
    "ROUMANIE": "RO", "ROYAUME-UNI": "GB", "RUSSIE": "RU", "SENEGAL": "SN",
    "SERBIE": "RS", "SINGAPOUR": "SG", "SLOVAQUIE": "SK", "SLOVENIE": "SI",
    "SOUDAN": "SD", "SRI LANKA": "LK", "SUEDE": "SE", "SUISSE": "CH",
    "SYRIE": "SY", "TAIWAN": "TW", "TANZANIE": "TZ", "TCHAD": "TD",
    "THAILANDE": "TH", "TOGO": "TG", "TUNISIE": "TN", "TURQUIE": "TR",
    "UKRAINE": "UA", "URUGUAY": "UY", "VENEZUELA": "VE", "VIET NAM": "VN",
    "YEMEN": "YE", "ZAMBIE": "ZM", "ZIMBABWE": "ZW",
    "GERMANY": "DE", "AUSTRALIA": "AU", "AUSTRIA": "AT", "BELGIUM": "BE",
    "BRAZIL": "BR", "CANADA": "CA", "CHINA": "CN", "DENMARK": "DK",
    "SPAIN": "ES", "FINLAND": "FI", "FRANCE": "FR", "GREECE": "GR",
    "INDIA": "IN", "IRELAND": "IE", "ITALY": "IT", "JAPAN": "JP",
    "LUXEMBOURG": "LU", "MOROCCO": "MA", "MEXICO": "MX", "NORWAY": "NO",
    "NETHERLANDS": "NL", "POLAND": "PL", "PORTUGAL": "PT",
    "UNITED KINGDOM": "GB", "ROMANIA": "RO", "SWEDEN": "SE",
    "SWITZERLAND": "CH", "TURKEY": "TR", "UNITED STATES": "US", "ENGLAND": "GB",
}

# Table NAF rev.2 : code -> libelle
_NAF_PATH = os.path.join(APP_DIR, "naf_rev2.json")
if os.path.exists(_NAF_PATH):
    with open(_NAF_PATH, "r", encoding="utf-8") as _f:
        NAF_LABELS = json.load(_f)
else:
    NAF_LABELS = {}

def _naf_label(code):
    """Retourne le libelle NAF pour un code (ex: '58.29C' -> 'Edition de logiciels applicatifs')."""
    if not code: return ""
    c = str(code).strip()
    if c in NAF_LABELS: return NAF_LABELS[c]
    # Essayer sans la lettre finale (XX.XX)
    if len(c) >= 5 and c[:-1] in NAF_LABELS: return NAF_LABELS[c[:-1]]
    return c  # fallback : retourner le code tel quel

# --- Normalisation forme juridique ---
# Codes Evoliz : https://evoliz.io/documentation#section/Legal-status-list
EVOLIZ_LEGAL = {
    "1": "Association", "2": "Auto entrepreneur", "3": "EIRL",
    "4": "Entreprise Individuelle", "5": "EURL", "6": "GIE",
    "7": "Independant - AGESSA", "8": "Independant - Maison des Artistes",
    "9": "SA", "10": "SARL", "11": "SARL a capital variable",
    "12": "SAS", "13": "SASU", "14": "SNC", "15": "SELARL",
    "16": "Profession liberale", "17": "SCI", "18": "SCS",
    "19": "Societe Civile", "20": "SCM", "21": "SEP", "22": "SCP",
    "23": "SCOP SA", "24": "SCOP SARL", "25": "SCIC SA", "26": "SCIC SARL",
    "27": "SELAFA", "28": "SELAS", "29": "SELCA", "30": "SPRL",
}

# Code INSEE nature_juridique -> label Evoliz
# Ref: https://www.insee.fr/fr/information/2028129
_INSEE_TO_EVOLIZ = {
    # Entrepreneur individuel
    "1000": "Entreprise Individuelle",
    # Auto-entrepreneur / micro
    "1300": "Auto entrepreneur",
    # EIRL
    "1400": "EIRL",
    # GIE
    "6100": "GIE",
    # SA a conseil d'admin
    "5505": "SA", "5510": "SA", "5515": "SA", "5520": "SA",
    "5522": "SA", "5525": "SA", "5530": "SA", "5531": "SA",
    "5532": "SA", "5535": "SA", "5538": "SA",
    # SA a directoire
    "5605": "SA", "5610": "SA", "5615": "SA", "5620": "SA",
    "5622": "SA", "5625": "SA", "5630": "SA", "5631": "SA",
    "5632": "SA", "5635": "SA", "5638": "SA",
    # SAS
    "5710": "SAS",
    # SASU
    "5720": "SASU",
    # SARL
    "5499": "SARL", "5498": "SARL",
    # EURL
    "5498": "EURL",
    # SARL
    "5410": "SARL", "5415": "SARL", "5422": "SARL",
    "5426": "SARL", "5430": "SARL", "5431": "SARL",
    "5432": "SARL", "5442": "SARL", "5443": "SARL",
    "5451": "SARL", "5453": "SARL", "5454": "SARL",
    "5455": "SARL", "5458": "SARL", "5459": "SARL",
    "5460": "SARL",
    # SNC
    "5202": "SNC", "5203": "SNC",
    # SCS
    "5306": "SCS", "5307": "SCS", "5308": "SCS",
    # SCI
    "6540": "SCI",
    # SCM
    "6542": "SCM",
    # SCP
    "6543": "SCP",
    # Societe civile
    "6553": "Societe Civile", "6554": "Societe Civile",
    "6558": "Societe Civile", "6560": "Societe Civile",
    # SEP
    "6539": "SEP",
    # SELARL
    "5485": "SELARL",
    # SELAS
    "5785": "SELAS",
    # SELAFA
    "5585": "SELAFA",
    # SELCA
    "5385": "SELCA",
    # SCOP
    "5451": "SCOP SARL", "5547": "SCOP SA",
    # Association
    "9210": "Association", "9220": "Association", "9221": "Association",
    "9222": "Association", "9223": "Association", "9224": "Association",
    "9230": "Association", "9240": "Association", "9260": "Association",
    # Profession liberale
    "0000": "Profession liberale",
}

def _normalize_forme_juridique(value):
    """Normalise une forme juridique (code INSEE, texte libre) vers un label Evoliz."""
    if not value:
        return ""
    s = str(value).strip()

    # Si c'est un code INSEE numerique
    if s.isdigit():
        # Match exact
        if s in _INSEE_TO_EVOLIZ:
            return _INSEE_TO_EVOLIZ[s]
        # Match par prefixe (ex: 5710 -> 57xx)
        for prefix_len in [4, 3, 2]:
            prefix = s[:prefix_len]
            for code, label in _INSEE_TO_EVOLIZ.items():
                if code.startswith(prefix):
                    return label
        return ""  # code INSEE inconnu, vide pour ne pas bloquer l'import

    # Si c'est du texte, essayer de matcher vers un label Evoliz
    up = s.upper().strip()
    # Mapping texte courant -> Evoliz
    _TEXT_MAP = {
        "SARL": "SARL", "S.A.R.L.": "SARL", "S.A.R.L": "SARL",
        "SAS": "SAS", "S.A.S.": "SAS", "S.A.S": "SAS",
        "SASU": "SASU", "S.A.S.U.": "SASU",
        "SA": "SA", "S.A.": "SA", "S.A": "SA",
        "EURL": "EURL", "E.U.R.L.": "EURL",
        "EIRL": "EIRL", "E.I.R.L.": "EIRL",
        "EI": "Entreprise Individuelle", "ENTREPRISE INDIVIDUELLE": "Entreprise Individuelle",
        "SNC": "SNC", "S.N.C.": "SNC",
        "SCI": "SCI", "S.C.I.": "SCI",
        "SCM": "SCM", "S.C.M.": "SCM",
        "SCP": "SCP", "S.C.P.": "SCP",
        "SCS": "SCS", "S.C.S.": "SCS",
        "SEP": "SEP", "S.E.P.": "SEP",
        "GIE": "GIE", "G.I.E.": "GIE",
        "SELARL": "SELARL", "SELAS": "SELAS", "SELAFA": "SELAFA", "SELCA": "SELCA",
        "SCOP": "SCOP SARL", "SCIC": "SCIC SARL",
        "SPRL": "SPRL",
        "ASSOCIATION": "Association", "ASSOC": "Association", "ASSO": "Association",
        "ASSOCIATION LOI 1901": "Association", "ASSOCIATION DECLAREE": "Association",
        "AUTO ENTREPRENEUR": "Auto entrepreneur", "AUTO-ENTREPRENEUR": "Auto entrepreneur",
        "AUTOENTREPRENEUR": "Auto entrepreneur",
        "MICRO ENTREPRISE": "Auto entrepreneur", "MICRO-ENTREPRISE": "Auto entrepreneur",
        "MICROENTREPRISE": "Auto entrepreneur",
        "PROFESSION LIBERALE": "Profession liberale", "PROF. LIBERALE": "Profession liberale",
        "SOCIETE CIVILE": "Societe Civile",
        "SOCIETE PAR ACTIONS SIMPLIFIEE": "SAS", "SOCIETE PAR ACTIONS SIMPLIFIEES": "SAS",
        "SOCIETE A RESPONSABILITE LIMITEE": "SARL",
        "SOCIETE ANONYME": "SA",
        "SOCIETE EN NOM COLLECTIF": "SNC",
        "SOCIETE CIVILE IMMOBILIERE": "SCI",
        "SOCIETE CIVILE DE MOYENS": "SCM",
        "SOCIETE CIVILE PROFESSIONNELLE": "SCP",
        "SOCIETE EN COMMANDITE SIMPLE": "SCS",
        "GROUPEMENT D'INTERET ECONOMIQUE": "GIE",
        "ENTREPRISE UNIPERSONNELLE": "EURL",
    }
    if up in _TEXT_MAP:
        return _TEXT_MAP[up]
    # Recherche partielle : contient un des mots cles
    for key, val in _TEXT_MAP.items():
        if key in up:
            return val
    # Rien trouve : vide pour ne pas bloquer l'import Evoliz
    return ""

def _lookup_iso2(country):
    if not country: return "FR"
    n = norm_piv(country)
    for name, code in ISO2_MAP.items():
        if norm_piv(name) == n: return code
    for name, code in ISO2_MAP.items():
        if norm_piv(name) in n or n in norm_piv(name): return code
    return "FR"

def _make_wb(headers):
    wb = Workbook(); ws = wb.active; ws.append(headers); return wb, ws

def _wb_bytes(wb):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def _read_meg(f, sheet_name=0):
    for engine in [None, "openpyxl", "xlrd"]:
        try:
            f.seek(0); return pd.read_excel(f, header=0, engine=engine, sheet_name=sheet_name)
        except Exception: pass
    # Fallback HTML (certains .xls sont du HTML déguisé)
    try:
        f.seek(0)
        raw = f.read() if hasattr(f, 'read') else open(f, 'rb').read()
        dfs = pd.read_html(raw, header=0)
        if dfs:
            idx = sheet_name if isinstance(sheet_name, int) else 0
            return dfs[idx] if idx < len(dfs) else dfs[0]
    except Exception: pass
    raise ValueError("Impossible de lire ce fichier.")

def _safe_float(v):
    try: return float(v)
    except (ValueError, TypeError): return 0.0

def _detect_type_from_name(name):
    """Analyse un nom pour determiner le type client Evoliz :
    'Particulier', 'Professionnel' ou 'Administration publique'."""
    if not name:
        return "Professionnel"
    s = name.strip()
    up = s.upper()

    # --- Administration publique ---
    admin_markers = [
        "MAIRIE", "COMMUNE ", "COMMUNAUTE ", "CONSEIL DEPARTEMENTAL",
        "CONSEIL REGIONAL", "CONSEIL GENERAL", "CONSEIL MUNICIPAL",
        "REGION ", "DEPARTEMENT ", "PREFECTURE", "SOUS-PREFECTURE",
        "MINISTERE", "TRESOR PUBLIC", "TRESORERIE", "DGFIP",
        "LYCEE", "COLLEGE", "ECOLE PUBLIQUE", "UNIVERSITE",
        "CENTRE HOSPITALIER", "CHU ", "CHR ", "HOPITAL PUBLIC",
        "CNRS", "INRA", "INSERM", "INSEE", "POLE EMPLOI",
        "CHAMBRE DE COMMERCE", "CCI ", "CHAMBRE DES METIERS",
        "CHAMBRE D'AGRICULTURE", "OFFICE PUBLIC", "OPH ",
        "ETABLISSEMENT PUBLIC", "EPIC ", "EPA ",
        "CAISSE DES DEPOTS", "CDC ", "CPAM ", "CAF ",
        "URSSAF", "MSA ", "CARSAT",
    ]
    for marker in admin_markers:
        if marker in up:
            return "Administration publique"

    # --- Professionnel (societes, structures, commerces) ---
    pro_markers = [
        "SARL", "SAS", "SASU", "SA ", "SCI", "SNC", "EURL", "SELARL", "SELAS",
        "EARL", "GIE", "SCOP", "SCA", "SCS", "SEP", "GAEC", "SELAFA", "SELCA",
        "EI ", "MICRO", "AUTO-ENTREPRENEUR",
        "ASSOCIATION", "ASSOC", "FONDATION",
        "CABINET", "OFFICE", "GROUPE", "HOLDING", "SOCIETE",
        "PHARMACIE", "CLINIQUE", "LABORATOIRE", "LABO",
        "BOULANGERIE", "RESTAURANT", "HOTEL", "GARAGE", "PRESSING",
        "INSTITUT", "AGENCE", "SYNDIC", "COPROPRIETE",
        "INTERNATIONAL", "SERVICES", "CONSULTING", "CONSEIL",
        "TRANSPORT", "LOGISTIQUE", "IMMOBILIER",
        "MUTUELLE", "BANQUE", "CREDIT", "ASSURANCE",
        "&", " ET FILS", " ET CIE", " ET ASSOCIES", " ET COMPAGNIE",
        "CHEZ ", "C/O ",
    ]
    for marker in pro_markers:
        if marker in up:
            return "Professionnel"

    # --- Heuristiques pour Particulier ---
    # Chiffres dans le nom -> probablement pas une personne
    if re.search(r'\d', s):
        return "Professionnel"
    # Un seul mot -> probablement un nom de societe
    words = s.split()
    if len(words) == 1:
        return "Professionnel"
    # Plus de 3 mots -> probablement un nom de societe
    if len(words) > 3:
        return "Professionnel"
    # Plus de 35 caracteres
    if len(s) > 35:
        return "Professionnel"
    # 2-3 mots, uniquement des lettres/tirets -> probablement Prenom Nom
    if all(re.match(r'^[A-Za-z\u00C0-\u00FF\-\']+$', w) for w in words):
        return "Particulier"
    return "Professionnel"

def _normalize_type(type_val, nom_societe=""):
    """Normalise un champ type source vers une valeur Evoliz.
    Si le champ type n'est pas exploitable, analyse le nom."""
    if type_val:
        tv = str(type_val).upper().strip()
        # Valeurs directement reconnues
        if tv in ("PRO", "PROFESSIONNEL", "ENTREPRISE", "SOCIETE", "B2B", "PROFESSIONAL"):
            return "Professionnel"
        if tv in ("PART", "PARTICULIER", "INDIVIDU", "B2C", "PERSONNE PHYSIQUE", "INDIVIDUAL", "PERSONAL"):
            return "Particulier"
        if tv in ("ADMIN", "ADMINISTRATION", "ADMINISTRATION PUBLIQUE", "PUBLIC",
                   "COLLECTIVITE", "COLLECTIVITE TERRITORIALE", "ETAT"):
            return "Administration publique"
        # Valeur non reconnue -> essayer d'analyser le contenu
        # Certains fichiers mettent le type juridique dans le champ type
        for marker in ["SARL", "SAS", "SASU", "SA", "SCI", "EURL", "SNC", "GIE"]:
            if marker in tv:
                return "Professionnel"
        for marker in ["MAIRIE", "COMMUNE", "PREFECTURE", "MINISTERE", "LYCEE", "COLLEGE"]:
            if marker in tv:
                return "Administration publique"
    # Type vide ou non exploitable -> ne pas deviner, laisser vide
    # (sera déterminé par l'enrichissement Sirene via nature_juridique)
    if nom_societe:
        return _detect_type_from_name(nom_societe)
    return ""

def _parse_date(v):
    if isinstance(v, dt_datetime): return v
    if pd.isna(v): return ""
    for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
        try: return dt_datetime.strptime(str(v).strip(), fmt)
        except (ValueError, TypeError): pass
    return v

H_CLIENT = ["Code *","Date de creation","Societe / Nom *","Type *","Civilite","Forme juridique","Siren","APE / NAF","TVA intracommunautaire","Numero Immatriculation","Banque","RIB","IBAN","BIC","Adresse","Complement d'adresse","Complement d'adresse (suite)","Code postal *","Ville *","Pays","Code pays (ISO 2) *","Siret","Nb adresses livraison","Telephone","Portable","Fax","Site web","Montant de l'encours garanti","Commentaires","Desactive","Taux de penalite","Aucun Taux de penalite","Frais de recouvrement","Taux d'escompte","Aucun Taux d'escompte","Conditions de reglement","Mode de paiement","Duree de validite","Mode de saisie des prix","Taux de TVA","Remise globale","Article d'exoneration","ZRR","Code ZRR","Devise"]
H_FOURNISSEUR = ["Code *","Date de creation","Raison sociale *","Forme juridique","Siret","APE / NAF","TVA intracommunautaire","RIB","IBAN","BIC","Adresse","Adresse (suite)","Code postal","Ville","Pays","Code pays (ISO 2)","Telephone","Portable","Fax","Site web","Classification","Code classification","Commentaires","Desactive","Conditions de reglement","Mode de paiement","Axe 1","Code Axe 1"]
H_CONTACT = ["Nom Client","Code client *","Civilite","Nom *","Prenom","E-mail","Metier/Fonction","Consentement *","Libelle Telephone","Telephone","Libelle Telephone 2","Telephone 2","Libelle Telephone 3","Telephone 3","Desactive"]
H_FACTURE = ["N facture externe *","Date facture *","Client","Code client *","Nom adresse de livraison","Code adresse de livraison","Total TVA","Total HT","Total TTC","Total regle","Etat","Date Etat","Date de creation","Objet","Date d'echeance","Date d'execution","Taux de penalite","Frais de recouvrement","Taux d'escompte","Conditions de reglement *","Mode de paiement","Remise globale","Acompte","Nombre de relance","Commentaires","N facture","Annule","Catalogue","Ref.","Designation *","Qte *","Unite","PU HT *","Remise","TVA","Total TVA","Total HT","Classification vente","Code Classification vente","Prix d'achat HT","Createur"]
H_AVOIR = ["N avoir externe *","Date avoir *","Client","Code client *","Nom adresse de livraison","Code adresse de livraison","Total TVA","Total HT","Total TTC","Etat","Date de creation avoir","Objet","Date d'echeance","Taux de penalite","Frais de recouvrement","Taux d'escompte","Conditions de reglement *","Mode de paiement","Remise globale","Acompte","Commentaires","Annule","Catalogue","Ref.","Designation *","Qte *","Unite","PU HT *","Remise","TVA","Total TVA","Total HT","Classification vente","Code Classification vente","Createur"]
H_PAIEMENT = ["Facture n *","Date paiement *","Date de creation","Client","Code client","Libelle *","Mode de paiement *","Montant *","Commentaires","Createur"]
H_ARTICLE = ["Reference *","Nature","Classification vente","Code Classification vente","Designation *","Quantite","Poids par unite","Unite","PU HT","PU TTC","TVA","Saisie en TTC","Prix d'achat HT","Classification achat","Code Classification achat","Coefficient multiplicateur","% Marque","% Marge","Marge brute","Fournisseur","Code fournisseur","Ref. Fournisseur","Article stocke","Qte stockee","Desactive","Createur"]

GABARIT_CLIENT_PATH = os.path.join(APP_DIR, ".gabarit_client_meg.xlsx")

# --- Auto-mapping colonnes ---
# Champs Evoliz client -> mots-cles pour detection automatique
_EVOLIZ_FIELD_KEYWORDS = {
    "Code":                 ["code", "ref", "id client", "identifiant", "numero client", "n client", "code client"],
    "Societe / Nom":        ["societe", "raison sociale", "nom", "denomination", "entreprise", "client", "name", "company"],
    "Type":                 ["type", "nature", "categorie"],
    "Civilite":             ["civilite", "titre", "civ"],
    "Forme juridique":      ["forme juridique", "statut juridique", "forme legale", "forme jur"],
    "Siren":                ["siren"],
    "APE / NAF":            ["ape", "naf", "code activite"],
    "TVA intracommunautaire": ["tva intra", "tva", "n tva", "vat", "num tva"],
    "Adresse":              ["adresse", "adresse 1", "rue", "voie", "address", "adresse facturation", "auto adresse rue"],
    "Complement d'adresse": ["complement", "adresse 2", "adresse2", "complement adresse"],
    "Code postal":          ["code postal", "cp", "postal", "zip", "auto code postal"],
    "Ville":                ["ville", "commune", "city", "localite", "auto ville"],
    "Pays":                 ["pays", "country"],
    "Code pays (ISO 2)":    ["iso", "code pays", "country code", "iso2", "auto pays iso2"],
    "Siret":                ["siret"],
    "Telephone":            ["telephone", "tel", "phone", "tel fixe", "tel bureau"],
    "Portable":             ["portable", "mobile", "gsm", "cell"],
    "Fax":                  ["fax", "telecopie"],
    "Site web":             ["site", "web", "url", "site web", "website"],
    "Commentaires":         ["commentaire", "note", "observation", "memo", "comment"],
    "E-mail":               ["email", "e-mail", "mail", "courriel", "adresse mail"],
    "Prenom":               ["prenom", "first name", "firstname"],
    "Nom contact":          ["nom contact", "nom", "last name", "lastname", "nom de famille"],
}

def _auto_map_columns(src_columns):
    """Propose un mapping automatique des colonnes source vers les champs Evoliz."""
    mapping = {}

    for evoliz_field, keywords in _EVOLIZ_FIELD_KEYWORDS.items():
        best_match = None
        best_score = 0
        for src_col in src_columns:
            sn = norm_piv(str(src_col))
            if not sn:
                continue
            for kw in keywords:
                kn = norm_piv(kw)
                if not kn:
                    continue
                score = 0
                if kn == sn:
                    # Match exact normalise
                    score = 100
                elif kn in sn and len(kn) >= len(sn) * 0.6:
                    # Keyword contenu dans source, mais assez long pour eviter faux positifs
                    score = 85
                elif sn in kn and len(sn) >= len(kn) * 0.6:
                    # Source contenu dans keyword, assez long
                    score = 80
                elif kn in sn:
                    # Keyword court contenu dans source plus longue
                    score = 50 + len(kn) * 2
                if score > best_score:
                    best_score = score
                    best_match = src_col
        if best_match and best_score >= 50:
            mapping[evoliz_field] = best_match
    return mapping

# --- Onglet Injection Clients ---
if _connected and mod_clients:
 with m_cli:
    _gate_cli = bool(st.session_state.get('company_id_105')) and bool(st.session_state.get('token_headers_105'))
    if not _gate_cli:
        st.warning("⛔ Connectez-vous a l'API et selectionnez un dossier (onglet **🔑 Connexion API**) avant d'utiliser cet onglet.")
    _is_supplier = False
    _entity_label = "clients"
    _entity_api = "clients"
    _entity_id_field = "clientid"
    _entity_name_field = "Societe / Nom"
    _H_ENTITY = H_CLIENT
    st.subheader("👥 Injection Clients")
    st.caption("1. Importez un fichier clients  2. Consolidation avec Evoliz  3. Enrichissement Sirene  4. Injection")

    # --- Etape 1 : Upload fichier ---
    f_meg_cli = st.session_state.get("imp_file_clients")
    if not f_meg_cli:
        st.info("Importez d'abord un fichier clients dans l'onglet 📁 Import fichiers.")

    if f_meg_cli:
        # Lecture du fichier
        if f_meg_cli.name.lower().endswith(".csv"):
            # Essayer plusieurs encodages (utf-8, latin-1, cp1252)
            df_src = None
            for _enc in ["utf-8", "latin-1", "cp1252"]:
                for _sep in [None, ";", ","]:
                    try:
                        f_meg_cli.seek(0)
                        _kw = {"header": 0, "encoding": _enc}
                        if _sep: _kw["sep"] = _sep
                        else: _kw["sep"] = None; _kw["engine"] = "python"
                        df_src = pd.read_csv(f_meg_cli, **_kw)
                        if len(df_src.columns) > 1: break
                        df_src = None
                    except Exception:
                        df_src = None
                if df_src is not None: break
            if df_src is None:
                st.error("Impossible de lire ce fichier CSV. Verifiez l'encodage.")
                f_meg_cli = None
        else:
            _cli_sheet = st.session_state.get("imp_file_clients_sheet", 0)
            df_src = _read_meg(f_meg_cli, sheet_name=_cli_sheet)

        st.caption(f"Fichier lu : **{len(df_src)} lignes**, **{len(df_src.columns)} colonnes**")

        # --- Detection et parsing adresse -> colonnes virtuelles ---
        # Si une colonne ressemble a une adresse mais qu'il n'y a pas de colonne CP/Ville,
        # on parse l'adresse pour creer des colonnes synthetiques
        fwd_test = _auto_map_columns(df_src.columns.tolist())
        has_adr_col = "Adresse" in fwd_test
        has_cp_col = "Code postal" in fwd_test
        has_ville_col = "Ville" in fwd_test
        has_pays_col = "Pays" in fwd_test or "Code pays (ISO 2)" in fwd_test

        if has_adr_col and (not has_cp_col or not has_ville_col):
            adr_src_col = fwd_test["Adresse"]
            parsed_rues, parsed_cps, parsed_villes, parsed_pays = [], [], [], []
            for _, row in df_src.iterrows():
                adr = to_clean_str(row.get(adr_src_col, ""))
                m_adr = re.search(r'^(.*?)\s+(\d{5})\s+(.+)$', adr) if adr else None
                if m_adr:
                    parsed_rues.append(m_adr.group(1).strip())
                    parsed_cps.append(m_adr.group(2))
                    reste = m_adr.group(3).strip().upper()
                    ville_p, pays_p = reste, ""
                    if "FRANCE" in ville_p:
                        ville_p = ville_p.replace("FRANCE", "").strip()
                        pays_p = "FR"
                    else:
                        for cn, cc in ISO2_MAP.items():
                            if cn in ville_p:
                                ville_p = ville_p.replace(cn, "").strip()
                                pays_p = cc
                                break
                    parsed_villes.append(ville_p)
                    parsed_pays.append(pays_p if pays_p else "FR")
                else:
                    parsed_rues.append(adr)
                    parsed_cps.append("")
                    parsed_villes.append("")
                    parsed_pays.append("")

            # Ajouter les colonnes virtuelles au DataFrame
            added_cols = []
            if not has_cp_col and any(parsed_cps):
                df_src["[Auto] Code postal"] = parsed_cps
                added_cols.append("Code postal")
            if not has_ville_col and any(parsed_villes):
                df_src["[Auto] Ville"] = parsed_villes
                added_cols.append("Ville")
            if not has_pays_col and any(parsed_pays):
                df_src["[Auto] Pays ISO2"] = parsed_pays
                added_cols.append("Pays ISO2")
            # Remplacer l'adresse originale par la rue seule (in-place)
            df_src[adr_src_col] = parsed_rues
            added_cols.append(f"Adresse nettoyee (rue seule dans '{adr_src_col}')")

            if added_cols:
                st.info(f"🔄 Colonnes generees depuis l'adresse : **{', '.join(added_cols)}**")

        st.dataframe(df_src.head(5), use_container_width=True, hide_index=True)

        # --- Auto-mapping inverse ---
        if _is_supplier:
            all_evoliz_fields = ["— Ignorer",
                "Raison sociale", "Code", "Code postal", "Ville", "Code pays (ISO 2)",
                "Forme juridique", "Siren", "Siret", "APE / NAF", "TVA intracommunautaire",
                "Adresse", "Adresse (suite)", "Pays",
                "Telephone", "Portable", "Fax", "E-mail", "Site web",
                "Classification", "Code classification", "Commentaires"]
            required_set = {"Raison sociale", "Code"}
        else:
            all_evoliz_fields = ["— Ignorer",
                "Societe / Nom", "Code", "Type", "Code postal", "Ville", "Code pays (ISO 2)",
                "Civilite", "Prenom", "Nom contact",
                "Forme juridique", "Siren", "Siret", "APE / NAF", "TVA intracommunautaire",
                "Adresse", "Complement d'adresse", "Pays",
                "Telephone", "Portable", "Fax", "E-mail", "Site web", "Commentaires"]
            required_set = {"Societe / Nom", "Code", "Code postal", "Ville", "Code pays (ISO 2)"}

        # Recalculer le mapping si fichier change ou si colonnes ont change (ajout auto)
        current_cols_sig = ",".join(df_src.columns.tolist())
        if (st.session_state.get("meg_last_file") != f_meg_cli.name
                or st.session_state.get("meg_last_cols_sig") != current_cols_sig):
            # Auto-detect : pour chaque colonne source, quel champ Evoliz ?
            fwd = _auto_map_columns(df_src.columns.tolist())  # evoliz_field -> src_col
            rev = {}  # src_col -> evoliz_field
            for ef, sc in fwd.items():
                if sc not in rev:  # premiere correspondance gagne
                    rev[sc] = ef
            st.session_state["meg_col_mapping_rev"] = rev
            st.session_state["meg_last_file"] = f_meg_cli.name
            st.session_state["meg_last_cols_sig"] = current_cols_sig

        rev_map = st.session_state["meg_col_mapping_rev"]

        st.divider()
        st.subheader("🔗 Mapping des colonnes du fichier")
        st.caption("Pour chaque colonne de votre fichier, indiquez a quel champ Evoliz elle correspond.")

        # Affichage en 2 colonnes : colonne source | champ Evoliz cible
        mapping_rev = {}  # src_col -> evoliz_field
        n_cols = len(df_src.columns)
        cols_per_row = 2
        for i in range(0, n_cols, cols_per_row):
            ui_cols = st.columns(cols_per_row)
            for j in range(cols_per_row):
                if i + j >= n_cols:
                    break
                src_col = df_src.columns[i + j]
                default_ev = rev_map.get(src_col, "— Ignorer")
                idx = all_evoliz_fields.index(default_ev) if default_ev in all_evoliz_fields else 0
                with ui_cols[j]:
                    chosen = st.selectbox(
                        f"📄 **{src_col}**",
                        all_evoliz_fields,
                        index=idx,
                        key=f"rmap_{i+j}",
                    )
                    mapping_rev[src_col] = chosen

        # Inverser : evoliz_field -> src_col
        mapping = {}
        for sc, ef in mapping_rev.items():
            if ef != "— Ignorer" and ef not in mapping:
                mapping[ef] = sc

        # --- Controle des champs obligatoires ---
        st.divider()
        st.subheader("📋 Controle des champs obligatoires")

        required_fields_ordered = list(required_set)
        ctrl_rows = []
        for ef in required_fields_ordered:
            served = ef in mapping
            # Determiner la source si auto-genere
            if served:
                statut = "✅ Mappe"
                source = mapping[ef]
            elif ef == "Code":
                statut = "🔄 Auto-genere"
                source = "Genere depuis le nom"
            elif ef == "Code pays (ISO 2)":
                statut = "🔄 Auto-genere"
                source = "FR par defaut"
            else:
                statut = "❌ Manquant"
                source = "Completable via Sirene"
            ctrl_rows.append({
                "Champ obligatoire": f"🔴 {ef}",
                "Statut": statut,
                "Source": source,
            })
        df_ctrl = pd.DataFrame(ctrl_rows)

        def _color_ctrl(row):
            s = row["Statut"]
            if "Mappe" in s: return ["background-color: #d4edda"] * len(row)
            if "Auto" in s: return ["background-color: #e8f4fd"] * len(row)
            if "Manquant" in s: return ["background-color: #f8d7da"] * len(row)
            return [""] * len(row)

        st.dataframe(df_ctrl.style.apply(_color_ctrl, axis=1), use_container_width=True, hide_index=True)

        real_missing = [ef for ef in required_fields_ordered
                        if ef not in mapping and ef not in ("Code", "Code pays (ISO 2)")]
        if real_missing:
            st.warning(f"Champs manquants : **{', '.join(real_missing)}** — completables via l'enrichissement Sirene.")
        else:
            st.success("Tous les champs obligatoires sont couverts.")


        st.session_state["meg_col_mapping_final"] = mapping

    # --- Etape 2 : Consolidation fichier (+ Evoliz si connecté) ---
    has_api = bool(st.session_state.get("token_headers_105"))
    if f_meg_cli:
        _cli_file_id = f_meg_cli.name + str(f_meg_cli.size)
        _no_data = st.session_state.get("meg_df_clients") is None
        _new_file = st.session_state.get("meg_consol_file_id") != _cli_file_id
        # Auto-consolidation uniquement au 1er chargement d'un nouveau fichier
        _auto_run = _no_data or _new_file
        _manual_run = st.button("🔄 Re-consolider" + (" avec Evoliz" if has_api else ""), use_container_width=True, key="btn_consolider")
        if _auto_run or _manual_run:
            st.session_state["meg_consol_file_id"] = _cli_file_id
            with st.spinner("Consolidation en cours..."):
                mapping = st.session_state.get("meg_col_mapping_final", {})
                headers = st.session_state.token_headers_105 if has_api else None
                cid = st.session_state.company_id_105 if has_api else None

                def _get(row, field):
                    col = mapping.get(field)
                    if not col or col == "— Ignorer" or col not in df_src.columns: return ""
                    return to_clean_str(row.get(col, ""))

                # Lire les entités Evoliz (clients ou fournisseurs) — seulement si API connectée
                ev_clients = []; ev_by_name = {}; ev_by_siren = {}
                if headers and cid:
                    # URL avec fallback sans prefixe (cas company_users avec cid invalide)
                    _url_primary = f"https://www.evoliz.io/api/v1/companies/{cid}/{_entity_api}"
                    _url_fallback = f"https://www.evoliz.io/api/v1/{_entity_api}"
                    url_cli = _url_primary
                    # Tester une fois pour determiner l'URL qui fonctionne
                    try:
                        _r_test = requests.get(_url_primary, headers=headers, params={"per_page": 1, "page": 1}, timeout=10)
                        if _r_test.status_code in (403, 404):
                            url_cli = _url_fallback
                    except Exception:
                        url_cli = _url_fallback
                    page = 1
                    while True:
                        r = requests.get(url_cli, headers=headers, params={"per_page": 100, "page": page}, timeout=15)
                        if r.status_code != 200: break
                        d = r.json()
                        for it in d.get("data", []):
                            adr = it.get("address") or {}
                            entry = {
                                _entity_id_field: it.get(_entity_id_field), "code": (it.get("code") or "").strip(),
                                "name": (it.get("name") or "").strip(), "type": (it.get("type") or ""),
                                "vat_number": (it.get("vat_number") or ""), "business_number": (it.get("business_number") or ""),
                                "business_identification_number": (it.get("business_identification_number") or ""),
                                "legalform": (it.get("legal_status") or {}).get("label", "") if isinstance(it.get("legal_status"), dict) else "",
                                "activity_number": (it.get("activity_number") or ""),
                                "phone": (it.get("phone") or ""), "mobile": (it.get("mobile") or ""),
                                "fax": (it.get("fax") or ""), "website": (it.get("website") or ""),
                                "addr": (adr.get("addr") or ""), "postcode": (adr.get("postcode") or ""),
                                "town": (adr.get("town") or ""), "iso2": (adr.get("iso2") or ""),
                            }
                            ev_clients.append(entry)
                            n = norm_piv(entry["name"])
                            if n: ev_by_name[n] = entry
                            s = entry.get("business_identification_number", "").strip()
                            if s and s != "N/C": ev_by_siren[s] = entry
                        if page >= d.get("meta", {}).get("last_page", 1): break
                        page += 1

                # Construire la liste consolidee
                ci = {h.split(" *")[0]: i for i, h in enumerate(_H_ENTITY)}
                consol_rows = []; seen_ev_ids = set()

                for _, row in df_src.iterrows():
                    nom = _get(row, _entity_name_field)
                    prenom = _get(row, "Prenom"); nom_contact = _get(row, "Nom contact")
                    if not nom and not _is_supplier and (prenom or nom_contact): nom = f"{prenom} {nom_contact}".strip()
                    if not nom: continue
                    _raw_code = _get(row, "Code")
                    # Si le code est un UUID ou trop long, generer depuis le nom
                    if _raw_code and len(_raw_code) <= 20 and "-" not in _raw_code:
                        code = _raw_code
                    else:
                        code = norm_piv(nom)[:15]
                    # Type : garder la valeur du fichier si mappée, sinon vide (sera déterminé par l'enrichissement Sirene)
                    _raw_type = _get(row, "Type")
                    if _raw_type:
                        type_c = _normalize_type(_raw_type, "")  # normaliser la valeur fichier sans deviner par le nom
                    else:
                        type_c = ""
                    cp = _get(row, "Code postal"); ville = _get(row, "Ville")
                    pays = _get(row, "Pays"); iso2 = _get(row, "Code pays (ISO 2)")
                    adresse = _get(row, "Adresse")
                    if adresse and (not cp or not ville):
                        m_adr = re.search(r'^(.*?)\s+(\d{5})\s+(.+)$', adresse)
                        if m_adr:
                            adresse = m_adr.group(1).strip()
                            if not cp: cp = m_adr.group(2)
                            if not ville:
                                v = m_adr.group(3).strip().upper()
                                if "FRANCE" in v: v = v.replace("FRANCE","").strip(); iso2 = iso2 or "FR"
                                ville = v
                    if iso2: iso2 = iso2.upper()
                    elif pays: iso2 = _lookup_iso2(pays)
                    else: iso2 = "FR"

                    entry = {
                        "Code": code, _entity_name_field: nom, "Type": type_c if not _is_supplier else "",
                        "Siren": _get(row, "Siren"), "Siret": _get(row, "Siret"),
                        "TVA intracommunautaire": _get(row, "TVA intracommunautaire") or ("NC" if type_c != "Particulier" else ""),
                        "Forme juridique": _normalize_forme_juridique(_get(row, "Forme juridique")),
                        "APE / NAF": _get(row, "APE / NAF"),
                        "Adresse": adresse, "Complement d'adresse": _get(row, "Complement d'adresse"),
                        "Code postal": cp or "NC", "Ville": ville or "NC", "Code pays (ISO 2)": iso2,
                        "Telephone": _get(row, "Telephone"), "Portable": _get(row, "Portable"),
                        "Fax": _get(row, "Fax"), "Site web": _get(row, "Site web"),
                        "Commentaires": _get(row, "Commentaires"),
                    }
                    siren_val = entry["Siren"]
                    ev = ev_by_name.get(norm_piv(nom))
                    if not ev and siren_val: ev = ev_by_siren.get(siren_val)
                    if ev:
                        seen_ev_ids.add(ev[_entity_id_field])
                        entry["_source"] = "📄+☁️ Doublon"
                        entry["_entityid"] = ev[_entity_id_field]
                    else:
                        entry["_source"] = "📄 Nouveau"
                        entry["_entityid"] = None
                    consol_rows.append(entry)

                for ev in ev_clients:
                    if ev[_entity_id_field] not in seen_ev_ids:
                        consol_rows.append({
                            "Code": ev["code"], _entity_name_field: ev["name"], "Type": ev.get("type", ""),
                            "Siren": ev["business_identification_number"], "Siret": ev["business_number"],
                            "TVA intracommunautaire": ev["vat_number"], "Forme juridique": ev["legalform"],
                            "APE / NAF": ev["activity_number"],
                            "Adresse": ev["addr"], "Complement d'adresse": "",
                            "Code postal": ev["postcode"], "Ville": ev["town"], "Code pays (ISO 2)": ev["iso2"],
                            "Telephone": ev["phone"], "Portable": ev["mobile"],
                            "Fax": ev["fax"], "Site web": ev["website"], "Commentaires": "",
                            "_source": "☁️ Evoliz seul", "_entityid": ev[_entity_id_field],
                        })

                wb_c, ws_c = _make_wb(_H_ENTITY)
                for cr in consol_rows:
                    ro = [None] * len(_H_ENTITY)
                    for field, idx_c in ci.items():
                        if field in cr: ro[idx_c] = cr[field]
                    ws_c.append(ro)
                cb = _wb_bytes(wb_c)
                with open(GABARIT_CLIENT_PATH, "wb") as f: f.write(cb)
                df_preview_c = pd.read_excel(io.BytesIO(cb), header=0)

                st.session_state["meg_df_clients"] = df_preview_c
                st.session_state["meg_df_clients_original"] = df_preview_c.copy()
                st.session_state["meg_consol_sources"] = {i: cr["_source"] for i, cr in enumerate(consol_rows)}
                st.session_state["meg_consol_ev_ids"] = {i: cr["_entityid"] for i, cr in enumerate(consol_rows)}
                st.session_state["meg_consol_stats"] = {
                    "fichier": sum(1 for cr in consol_rows if "Nouveau" in cr["_source"]),
                    "doublons": sum(1 for cr in consol_rows if "Doublon" in cr["_source"]),
                    "evoliz_seul": sum(1 for cr in consol_rows if "Evoliz seul" in cr["_source"]),
                    "total_evoliz": len(ev_clients),
                }
                for _k in ["meg_sirene_cells","meg_sirene_info","meg_sirene_log","meg_sirene_stats","meg_enrichir_flags","meg_sirene_suggestions"]:
                    st.session_state[_k] = set() if "cells" in _k else ({} if "info" in _k or "flags" in _k else ([] if "log" in _k else None))
                st.rerun()
    if f_meg_cli and not has_api and st.session_state.get("meg_df_clients") is not None:
        st.info("Connectez-vous à l'API Evoliz pour détecter les doublons et injecter.")

    # --- Init session ---
    for _k, _d in [("meg_sirene_cells", set()), ("meg_sirene_log", []), ("meg_sirene_stats", None),
                    ("meg_sirene_info", {}), ("meg_editor_ver", 0), ("meg_consol_sources", {}), ("meg_consol_stats", None)]:
        if _k not in st.session_state: st.session_state[_k] = _d

    # --- Affichage consolide ---
    if "meg_df_clients" in st.session_state and st.session_state["meg_df_clients"] is not None:
        df_preview_c = st.session_state["meg_df_clients"]
        sirene_cells = st.session_state.get("meg_sirene_cells", set())
        sirene_info = st.session_state.get("meg_sirene_info", {})
        sources = st.session_state.get("meg_consol_sources", {})
        consol_stats = st.session_state.get("meg_consol_stats")
        enrichir_flags = st.session_state.get("meg_enrichir_flags", {})

        if consol_stats:
            st.divider()
            st.subheader("📊 Consolidation")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("📄 Nouveaux", consol_stats.get("fichier",0))
            c2.metric("📄+☁️ Doublons", consol_stats.get("doublons",0))
            c3.metric("☁️ Evoliz seul", consol_stats.get("evoliz_seul",0))
            c4.metric("Total", len(df_preview_c))

        cols_show = ["Code *","Societe / Nom *","Type *","Siren","Adresse","Code postal *","Ville *",
                     "Code pays (ISO 2) *","Telephone","TVA intracommunautaire","Forme juridique","APE / NAF","Siret"]
        cols_show = [c for c in cols_show if c in df_preview_c.columns]
        df_show = df_preview_c[cols_show].copy()
        df_show.insert(0, "Source", df_show.index.map(lambda i: sources.get(i, "")))

        # Convertir Siren/Siret en str (Pandas les lit parfois en float)
        for _col_str in ["Siren", "Siret"]:
            if _col_str in df_show.columns:
                df_show[_col_str] = df_show[_col_str].apply(lambda v: to_clean_str(v) if not pd.isna(v) else "")

        has_enriched = sirene_cells and any(any((i, c) in sirene_cells for c in cols_show) for i in sirene_info)
        if has_enriched:
            for idx in df_show.index:
                for col in cols_show:
                    if (idx, col) in sirene_cells:
                        val = df_show.at[idx, col]
                        if val is not None and str(val).strip() and not str(val).startswith("🟢"):
                            df_show.at[idx, col] = f"🟢 {val}"
            df_show.insert(1, "✅", df_show.index.map(
                lambda i: enrichir_flags.get(i, True) if i in sirene_info and any((i, c) in sirene_cells for c in cols_show) else None))
            df_show.insert(2, "🟢 Nom trouve", df_show.index.map(lambda i: f"🟢 {v}" if (v := sirene_info.get(i, {}).get("nom", "")) else ""))
            df_show.insert(3, "🟢 APE trouve", df_show.index.map(lambda i: f"🟢 {v}" if (v := sirene_info.get(i, {}).get("activite", "")) else ""))

        # Filtre : tous / non enrichis uniquement
        _n_non_enrichis = len([i for i in df_show.index if i not in sirene_info])
        _filter_mode = st.radio(
            f"Afficher", [f"Tous ({len(df_show)})", f"Non enrichis uniquement ({_n_non_enrichis})"],
            horizontal=True, key="filter_enrichi")
        if "Non enrichis" in _filter_mode:
            _non_enrichi_idx = [i for i in df_show.index if i not in sirene_info]
            df_show = df_show.loc[_non_enrichi_idx]

        st.subheader(f"👥 {len(df_show)} client(s) affichés")

        # Colonnes éditables : Siren, Type
        _editable_cols = []
        if "Siren" in df_show.columns:
            _editable_cols.append("Siren")
        if "Type *" in df_show.columns:
            _editable_cols.append("Type *")
        _disabled_cols = [c for c in df_show.columns if c not in _editable_cols and c != "✅"]

        if has_enriched:
            st.caption("🟢 = enrichi Sirene. Decochez ✅ pour garder les donnees d'origine. Modifiez le Siren pour relancer la recherche.")
            _col_config = {"✅": st.column_config.CheckboxColumn("✅", default=True, width="small")}
            if "Siren" in df_show.columns:
                _col_config["Siren"] = st.column_config.TextColumn("Siren", width="medium")
            if "Type *" in df_show.columns:
                _col_config["Type *"] = st.column_config.SelectboxColumn("Type *", options=["Particulier", "Professionnel", "Administration publique"], width="medium")
            edited = st.data_editor(df_show, use_container_width=True, hide_index=True,
                disabled=_disabled_cols,
                column_config=_col_config,
                key=f"meg_enrichir_editor_{st.session_state.get('meg_editor_ver', 0)}")
            if "✅" in edited.columns:
                for i in edited.index:
                    v = edited.at[i, "✅"]
                    if v is not None: enrichir_flags[i] = bool(v)
                st.session_state["meg_enrichir_flags"] = enrichir_flags
        else:
            _col_config = {}
            if "Siren" in df_show.columns:
                _col_config["Siren"] = st.column_config.TextColumn("Siren", width="medium")
            if "Type *" in df_show.columns:
                _col_config["Type *"] = st.column_config.SelectboxColumn("Type *", options=["Particulier", "Professionnel", "Administration publique"], width="medium")
            st.caption("Modifiez le Siren pour relancer la recherche sur ce numéro.")
            edited = st.data_editor(df_show, use_container_width=True, hide_index=True,
                disabled=_disabled_cols,
                column_config=_col_config,
                key=f"meg_editor_siren_{st.session_state.get('meg_editor_ver', 0)}")

        # Persister les modifications de Type dans le dataframe source
        if "Type *" in df_show.columns and "Type *" in edited.columns:
            for i in edited.index:
                _old_type = str(df_show.at[i, "Type *"]).replace("🟢 ", "").strip() if i in df_show.index else ""
                _new_type = str(edited.at[i, "Type *"]).strip()
                if _new_type != _old_type and _new_type in ("Particulier", "Professionnel", "Administration publique"):
                    df_preview_c.at[i, "Type *"] = _new_type
                    st.session_state["meg_df_clients"] = df_preview_c

        # Persister les SIREN saisis manuellement dans df_preview_c (sans appel API)
        if "Siren" in edited.columns:
            _siren_pending = []
            for i in edited.index:
                _old_raw = to_clean_str(df_preview_c.at[i, "Siren"]) if "Siren" in df_preview_c.columns and i in df_preview_c.index else ""
                _new = str(edited.at[i, "Siren"]).replace("🟢 ", "").strip()
                if _new and _new != _old_raw and len(_new) >= 9 and _new.isdigit():
                    df_preview_c.at[i, "Siren"] = _new
                    _siren_pending.append((i, _new))
            if _siren_pending:
                st.session_state["meg_df_clients"] = df_preview_c
                st.info(f"💾 {len(_siren_pending)} SIREN saisi(s) — cliquez « 🔍 Enrichissement 2ème lame » ci-dessous pour compléter les données.")

        # --- CR enrichissement ---
        if st.session_state.get("meg_sirene_stats"):
            stats = st.session_state["meg_sirene_stats"]
            _total_recherches = stats["enriched"] + stats["already_complete"] + stats["not_found"]
            _pct_found = round(100 * (stats["enriched"] + stats["already_complete"]) / _total_recherches, 1) if _total_recherches else 0
            st.divider()
            st.subheader(f"📋 Enrichissement Sirene — {_pct_found}% identifiés")
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("✅ Enrichis", stats["enriched"])
            c2.metric("🟰 Déjà complets", stats["already_complete"])
            c3.metric("🔍 Non trouvés", stats["not_found"])
            c4.metric("⏭️ Ignorés", stats["skipped"])
            c5.metric("📊 % trouvés", f"{_pct_found}%")
            with st.expander("📋 Détail ligne par ligne", expanded=False):
                lr = st.session_state.get("meg_sirene_log", [])
                if lr:
                    df_log = pd.DataFrame(lr)
                    def _cl(row):
                        s = row.get("Statut","")
                        if "Enrichi" in s: return ["background-color:#d4edda"]*len(row)
                        if "Non trouve" in s or "Suggestions" in s: return ["background-color:#fff3cd"]*len(row)
                        if "Erreur" in s or "HTTP" in s: return ["background-color:#f8d7da"]*len(row)
                        return [""]*len(row)
                    st.dataframe(df_log.style.apply(_cl, axis=1), use_container_width=True, hide_index=True)

        # --- 1ère lame : enrichissement Sirene par nom ---
        st.divider()
        if st.button("🔍 1ère lame — Enrichir via Sirene", type="primary", use_container_width=True, key="btn_sirene"):
            st.session_state["meg_enrichir_flags"] = {}
            df_e = df_preview_c.copy()
            enriched = skipped = already_complete = not_found_count = 0
            log_rows = []; new_sc = set(sirene_cells); new_si = dict(sirene_info)
            _tasks = []
            for idx in df_e.index:
                nom = str(df_e.at[idx, "Societe / Nom *"]).strip() if "Societe / Nom *" in df_e.columns else ""
                type_c = str(df_e.at[idx, "Type *"]).strip() if "Type *" in df_e.columns else ""
                code = str(df_e.at[idx, "Code *"]).strip() if "Code *" in df_e.columns else ""
                if not nom:
                    skipped += 1; log_rows.append({"Client":code,"Nom":nom,"Statut":"⏭️ Vide","Trouve":"","Detail":""}); continue
                # Skip Particulier UNIQUEMENT si le fichier source contient une colonne
                # mappée sur le champ Type ET que cette colonne dit "Particulier"
                _type_mapped = "Type" in st.session_state.get("meg_col_mapping_final", {})
                if _type_mapped and type_c == "Particulier":
                    skipped += 1; log_rows.append({"Client":code,"Nom":nom,"Statut":"⏭️ Part. (fichier)","Trouve":"","Detail":""}); continue
                _has_siren = bool(to_clean_str(df_e.at[idx, "Siren"]) if "Siren" in df_e.columns else "")
                _has_siret = bool(to_clean_str(df_e.at[idx, "Siret"]) if "Siret" in df_e.columns else "")
                _has_cp = bool(to_clean_str(df_e.at[idx, "Code postal *"]) not in ("", "NC") if "Code postal *" in df_e.columns else False)
                _has_ville = bool(to_clean_str(df_e.at[idx, "Ville *"]) not in ("", "NC") if "Ville *" in df_e.columns else False)
                if _has_siren and _has_siret and _has_cp and _has_ville:
                    already_complete += 1; log_rows.append({"Client":code,"Nom":nom,"Statut":"🟰 Complet","Trouve":"","Detail":""}); continue
                _existing_siren = to_clean_str(df_e.at[idx, "Siren"]) if "Siren" in df_e.columns else ""
                if _existing_siren and len(_existing_siren) >= 9 and _existing_siren.isdigit():
                    _tasks.append((idx, _existing_siren, code, nom, "siren"))
                else:
                    sq = " ".join(nom.replace("nan","").split()).strip()
                    if sq: _tasks.append((idx, sq, code, nom, "nom"))
                    else: skipped += 1; continue

            def _search_sirene(task):
                idx, query, code, nom, mode = task
                params = {"q": query, "per_page": 1 if mode == "siren" else 5, "page": 1}
                try:
                    r = requests.get("https://recherche-entreprises.api.gouv.fr/search", params=params, timeout=10)
                    if r.status_code == 429: time.sleep(2); r = requests.get("https://recherche-entreprises.api.gouv.fr/search", params=params, timeout=10)
                    return (idx, code, nom, mode, query, r.status_code, r.json() if r.status_code == 200 else None)
                except Exception as exc:
                    return (idx, code, nom, mode, query, -1, str(exc))

            progress = st.progress(0, text="1ère lame — Recherche Sirene...")
            _results_list = []
            with ThreadPoolExecutor(max_workers=5) as pool:
                futures = {pool.submit(_search_sirene, t): t for t in _tasks}
                for _n, f in enumerate(as_completed(futures)):
                    _results_list.append(f.result())
                    if _tasks: progress.progress((_n + 1) / len(_tasks), text=f"{_n + 1}/{len(_tasks)}")

            for idx, code, nom, mode, query, status, data in _results_list:
                if status == -1:
                    log_rows.append({"Client":code,"Nom":nom,"Statut":"❌ Erreur","Trouve":"","Detail":str(data)[:60]}); continue
                if status != 200:
                    log_rows.append({"Client":code,"Nom":nom,"Statut":f"❌ HTTP {status}","Trouve":"","Detail":query}); continue
                results = data.get("results", []) if data else []
                if not results:
                    not_found_count += 1
                    if "Type *" in df_e.columns and df_e.at[idx, "Type *"] != "Particulier":
                        df_e.at[idx, "Type *"] = "Particulier"; new_sc.add((idx, "Type *"))
                        log_rows.append({"Client":code,"Nom":nom,"Statut":"🔍 Non trouvé → Particulier","Trouve":"","Detail":query})
                    else:
                        log_rows.append({"Client":code,"Nom":nom,"Statut":"🔍 Non trouvé","Trouve":"","Detail":query})
                    continue
                best = results[0]
                best_name = best.get("nom_complet", best.get("nom_raison_sociale",""))
                if mode == "nom":
                    nom_n = norm_piv(nom); best_n = norm_piv(best_name)
                    auto_match = (nom_n == best_n or nom_n in best_n or best_n in nom_n
                                  or (len(nom_n) > 3 and len(best_n) > 3 and
                                      len(set(nom_n) & set(best_n)) / max(len(set(nom_n)), len(set(best_n))) > 0.6))
                    if not auto_match:
                        not_found_count += 1
                        log_rows.append({"Client":code,"Nom":nom,"Statut":"🔍 Non trouvé (pas de match)","Trouve":best_name,"Detail":query})
                        continue
                ent = best; siege = ent.get("siege",{}); siren = ent.get("siren",""); siret = siege.get("siret","")
                nom_t = ent.get("nom_complet", ent.get("nom_raison_sociale",""))
                new_si[idx] = {"nom": nom_t or "", "activite": _naf_label(siege.get("activite_principale", ent.get("activite_principale",""))), "ville": siege.get("libelle_commune","")}
                champs = []
                def _sc(col, val, lbl, _idx=idx):
                    if col in df_e.columns and val and (not to_clean_str(df_e.at[_idx,col]) or to_clean_str(df_e.at[_idx,col])=="NC"):
                        df_e.at[_idx,col]=val; new_sc.add((_idx,col)); champs.append(f"{lbl}={val}")
                _sc("Siren",siren,"SIREN"); _sc("Siret",siret,"SIRET")
                if siren and "Type *" in df_e.columns:
                    _nj = str(ent.get("nature_juridique", "")).strip()
                    if _nj and _nj[:1] in ("7", "1"): _new_type = "Administration publique"
                    elif _nj and _nj[:1] == "9": _new_type = "Particulier"
                    else: _new_type = "Professionnel"
                    if df_e.at[idx, "Type *"] != _new_type:
                        df_e.at[idx, "Type *"] = _new_type; new_sc.add((idx, "Type *")); champs.append(f"Type={_new_type}")
                _sc("APE / NAF",ent.get("activite_principale",""),"NAF")
                _sc("Forme juridique",_normalize_forme_juridique(ent.get("nature_juridique","")),"Forme")
                if siren and "TVA intracommunautaire" in df_e.columns:
                    cur = to_clean_str(df_e.at[idx,"TVA intracommunautaire"])
                    if not cur or cur=="NC":
                        tv = f"FR{(12+3*(int(siren)%97))%97:02d}{siren}"; df_e.at[idx,"TVA intracommunautaire"]=tv; new_sc.add((idx,"TVA intracommunautaire")); champs.append(f"TVA={tv}")
                if not to_clean_str(df_e.at[idx,"Adresse"]) if "Adresse" in df_e.columns else True:
                    pts = [siege.get("numero_voie",""),siege.get("type_voie",""),siege.get("libelle_voie","")]
                    a = " ".join(p for p in pts if p)
                    if a and "Adresse" in df_e.columns: df_e.at[idx,"Adresse"]=a; new_sc.add((idx,"Adresse")); champs.append(f"Adr={a[:25]}")
                _sc("Code postal *",siege.get("code_postal",""),"CP"); _sc("Ville *",siege.get("libelle_commune",""),"Ville")
                if champs: enriched += 1; log_rows.append({"Client":code,"Nom":nom,"Statut":"✅ Enrichi","Trouve":nom_t,"Detail":", ".join(champs)})
                else: already_complete += 1; log_rows.append({"Client":code,"Nom":nom,"Statut":"🟰 Complet","Trouve":nom_t,"Detail":""})

            progress.empty()
            st.session_state["meg_df_clients"] = df_e; st.session_state["meg_sirene_cells"] = new_sc; st.session_state["meg_sirene_info"] = new_si
            st.session_state["meg_sirene_log"] = log_rows; st.session_state["meg_sirene_stats"] = {"enriched":enriched,"already_complete":already_complete,"not_found":not_found_count,"skipped":skipped}
            wb_n, ws_n = _make_wb(_H_ENTITY)
            for _, rw in df_e.iterrows(): ws_n.append([rw.iloc[i] if not pd.isna(rw.iloc[i]) else None for i in range(len(rw))])
            with open(GABARIT_CLIENT_PATH,"wb") as f: f.write(_wb_bytes(wb_n))
            st.session_state["meg_editor_ver"] = st.session_state.get("meg_editor_ver",0)+1
            st.rerun()

        # --- 2ème lame : propositions Sirene pour les non trouvés ---
        _non_enrichis = []
        if "Societe / Nom *" in df_preview_c.columns:
            for i in df_preview_c.index:
                if i not in sirene_info:
                    _nom = str(df_preview_c.at[i, "Societe / Nom *"]).strip()
                    _code = to_clean_str(df_preview_c.at[i, "Code *"]) if "Code *" in df_preview_c.columns else ""
                    if _nom and _nom != "nan":
                        _non_enrichis.append((i, _nom, _code))
        if _non_enrichis:
            st.divider()
            st.subheader(f"🔍 2ème lame — {len(_non_enrichis)} client(s) non identifié(s)")
            st.caption("Pour chaque client, les 2 résultats Sirene les plus probables sont proposés. Sélectionnez le bon ou ignorez.")
            if st.button(f"🔍 Rechercher les {len(_non_enrichis)} propositions", type="primary", use_container_width=True, key="btn_2eme_lame"):
                def _search_2eme(task):
                    idx, nom, code = task
                    try:
                        sq = " ".join(nom.replace("nan","").split()).strip()
                        if not sq: return (idx, nom, code, [])
                        r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                         params={"q": sq, "per_page": 2, "page": 1}, timeout=10)
                        if r.status_code == 429:
                            time.sleep(2)
                            r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                             params={"q": sq, "per_page": 2, "page": 1}, timeout=10)
                        if r.status_code != 200: return (idx, nom, code, [])
                        results = r.json().get("results", [])
                        props = []
                        for res in results[:2]:
                            rsie = res.get("siege") or {}
                            props.append({
                                "nom": res.get("nom_complet", res.get("nom_raison_sociale", "")),
                                "siren": res.get("siren", ""),
                                "ville": rsie.get("libelle_commune", ""),
                                "activite": _naf_label(rsie.get("activite_principale", "")),
                                "_raw": res,
                            })
                        return (idx, nom, code, props)
                    except Exception:
                        return (idx, nom, code, [])

                _all_props = []
                progress = st.progress(0, text="Recherche propositions Sirene...")
                with ThreadPoolExecutor(max_workers=5) as pool:
                    futures = {pool.submit(_search_2eme, t): t for t in _non_enrichis}
                    for _n, f in enumerate(as_completed(futures)):
                        _all_props.append(f.result())
                        progress.progress((_n + 1) / len(_non_enrichis))
                progress.empty()

                # Stocker les propositions en session state
                _props_dict = {}
                for idx, nom, code, props in _all_props:
                    if props:
                        _props_dict[idx] = {"client": nom, "code": code, "propositions": props}
                st.session_state["_2eme_lame_props"] = _props_dict
                st.session_state["_2eme_lame_result"] = f"2ème lame : {len(_props_dict)} client(s) avec propositions sur {len(_non_enrichis)} recherché(s)"
                st.rerun()

            if st.session_state.get("_2eme_lame_result"):
                st.success(st.session_state["_2eme_lame_result"])

            # Afficher les propositions pour validation
            _props = st.session_state.get("_2eme_lame_props", {})
            if _props:
                st.divider()
                st.subheader(f"📋 {len(_props)} proposition(s) à valider")
                _accepted = []
                for idx, info in sorted(_props.items()):
                    props = info["propositions"]
                    options = ["— Ignorer"] + [
                        f"{p['nom']} | SIREN {p['siren']} | {p['ville']} | {p['activite']}" for p in props
                    ]
                    sel = st.selectbox(f"**{info['client']}** ({info['code']})", options, key=f"prop2_{idx}")
                    if sel != "— Ignorer":
                        _sel_idx = options.index(sel) - 1
                        _accepted.append((idx, props[_sel_idx]))

                if _accepted:
                    if st.button(f"✅ Appliquer {len(_accepted)} sélection(s)", type="primary", use_container_width=True, key="btn_apply_2eme"):
                        df_e = df_preview_c.copy()
                        new_sc = set(sirene_cells); new_si = dict(sirene_info)
                        for idx, prop in _accepted:
                            ent = prop["_raw"]; siege = ent.get("siege", {})
                            siren = ent.get("siren", ""); siret = siege.get("siret", "")
                            nom_t = ent.get("nom_complet", ent.get("nom_raison_sociale", ""))
                            new_si[idx] = {"nom": nom_t, "activite": _naf_label(siege.get("activite_principale", "")),
                                           "ville": siege.get("libelle_commune", "")}
                            def _upd_2l(col, val, _idx=idx):
                                if col in df_e.columns and val:
                                    df_e.at[_idx, col] = val; new_sc.add((_idx, col))
                            _upd_2l("Siren", siren); _upd_2l("Siret", siret)
                            if siren and "Type *" in df_e.columns:
                                _nj = str(ent.get("nature_juridique", "")).strip()
                                if _nj and _nj[:1] in ("7", "1"):
                                    df_e.at[idx, "Type *"] = "Administration publique"; new_sc.add((idx, "Type *"))
                                elif _nj and _nj[:1] == "9":
                                    df_e.at[idx, "Type *"] = "Particulier"; new_sc.add((idx, "Type *"))
                                else:
                                    df_e.at[idx, "Type *"] = "Professionnel"; new_sc.add((idx, "Type *"))
                            _upd_2l("APE / NAF", ent.get("activite_principale", ""))
                            _upd_2l("Forme juridique", _normalize_forme_juridique(ent.get("nature_juridique", "")))
                            if siren and "TVA intracommunautaire" in df_e.columns:
                                try:
                                    tv = f"FR{(12 + 3 * (int(siren) % 97)) % 97:02d}{siren}"
                                    _upd_2l("TVA intracommunautaire", tv)
                                except ValueError: pass
                            pts = [siege.get("numero_voie", ""), siege.get("type_voie", ""), siege.get("libelle_voie", "")]
                            _upd_2l("Adresse", " ".join(p for p in pts if p))
                            _upd_2l("Code postal *", siege.get("code_postal", "")); _upd_2l("Ville *", siege.get("libelle_commune", ""))
                        st.session_state["meg_df_clients"] = df_e
                        st.session_state["meg_sirene_cells"] = new_sc
                        st.session_state["meg_sirene_info"] = new_si
                        st.session_state["meg_editor_ver"] = st.session_state.get("meg_editor_ver", 0) + 1
                        # MAJ stats
                        _prev_stats = st.session_state.get("meg_sirene_stats") or {"enriched": 0, "already_complete": 0, "not_found": 0, "skipped": 0}
                        st.session_state["meg_sirene_stats"] = {
                            "enriched": _prev_stats["enriched"] + len(_accepted),
                            "already_complete": _prev_stats["already_complete"],
                            "not_found": max(0, _prev_stats["not_found"] - len(_accepted)),
                            "skipped": _prev_stats["skipped"],
                        }
                        # Retirer les props appliquées
                        for idx, _ in _accepted:
                            if idx in st.session_state["_2eme_lame_props"]:
                                del st.session_state["_2eme_lame_props"][idx]
                        st.rerun()

        # --- 3ème lame (SIREN saisis manuellement) ---
        _siren_a_traiter = []
        if "Siren" in df_preview_c.columns:
            for i in df_preview_c.index:
                _s = to_clean_str(df_preview_c.at[i, "Siren"])
                if _s and len(_s) >= 9 and _s.isdigit() and i not in sirene_info:
                    _siren_a_traiter.append((i, _s))
        if _siren_a_traiter:
            st.divider()
            st.subheader(f"🔍 3ème lame — {len(_siren_a_traiter)} SIREN saisis à enrichir")
            st.caption("SIREN saisis manuellement et non encore enrichis.")
            if st.button(f"🔍 Enrichir les {len(_siren_a_traiter)} SIREN", type="primary", use_container_width=True, key="btn_sirene_3eme"):
                df_e = df_preview_c.copy()
                new_sc = set(sirene_cells); new_si = dict(sirene_info)
                _ok = _ko = 0
                progress = st.progress(0, text="Enrichissement 2ème lame...")
                for _n, (idx, siren_val) in enumerate(_siren_a_traiter):
                    progress.progress((_n + 1) / len(_siren_a_traiter), text=f"{_n + 1}/{len(_siren_a_traiter)} — SIREN {siren_val}")
                    try:
                        r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                         params={"q": siren_val, "per_page": 1, "page": 1}, timeout=10)
                        if r.status_code == 429: time.sleep(2); r = requests.get("https://recherche-entreprises.api.gouv.fr/search", params={"q": siren_val, "per_page": 1, "page": 1}, timeout=10)
                        if r.status_code != 200: _ko += 1; time.sleep(0.15); continue
                        results = r.json().get("results", [])
                        if not results: _ko += 1; time.sleep(0.15); continue
                        ent = results[0]; siege = ent.get("siege", {})
                        siren = ent.get("siren", ""); siret = siege.get("siret", "")
                        nom_t = ent.get("nom_complet", ent.get("nom_raison_sociale", ""))
                        new_si[idx] = {"nom": nom_t, "activite": _naf_label(siege.get("activite_principale", "")),
                                       "ville": siege.get("libelle_commune", "")}
                        def _upd2(col, val):
                            if col in df_e.columns and val:
                                df_e.at[idx, col] = val; new_sc.add((idx, col))
                        _upd2("Siren", siren); _upd2("Siret", siret)
                        if siren and "Type *" in df_e.columns:
                            _nj = str(ent.get("nature_juridique", "")).strip()
                            _new_t = "Administration publique" if _nj and _nj[:1] in ("7", "1") else ("Particulier" if _nj and _nj[:1] == "9" else "Professionnel")
                            df_e.at[idx, "Type *"] = _new_t; new_sc.add((idx, "Type *"))
                        _upd2("APE / NAF", ent.get("activite_principale", ""))
                        _upd2("Forme juridique", _normalize_forme_juridique(ent.get("nature_juridique", "")))
                        if siren and "TVA intracommunautaire" in df_e.columns:
                            tv = f"FR{(12 + 3 * (int(siren) % 97)) % 97:02d}{siren}"
                            _upd2("TVA intracommunautaire", tv)
                        pts = [siege.get("numero_voie", ""), siege.get("type_voie", ""), siege.get("libelle_voie", "")]
                        _upd2("Adresse", " ".join(p for p in pts if p))
                        _upd2("Code postal *", siege.get("code_postal", "")); _upd2("Ville *", siege.get("libelle_commune", ""))
                        _ok += 1
                        time.sleep(0.15)
                    except Exception:
                        _ko += 1
                progress.empty()
                st.session_state["meg_df_clients"] = df_e
                st.session_state["meg_sirene_cells"] = new_sc
                st.session_state["meg_sirene_info"] = new_si
                st.session_state["meg_editor_ver"] = st.session_state.get("meg_editor_ver", 0) + 1
                # Mettre à jour les stats d'enrichissement globales
                _prev_stats = st.session_state.get("meg_sirene_stats") or {"enriched": 0, "already_complete": 0, "not_found": 0, "skipped": 0}
                st.session_state["meg_sirene_stats"] = {
                    "enriched": _prev_stats["enriched"] + _ok,
                    "already_complete": _prev_stats["already_complete"],
                    "not_found": _prev_stats["not_found"] - _ok + _ko,  # ceux trouvés en 2ème lame ne sont plus "non trouvés"
                    "skipped": _prev_stats["skipped"],
                }
                # Corriger : not_found ne peut pas être négatif
                if st.session_state["meg_sirene_stats"]["not_found"] < 0:
                    st.session_state["meg_sirene_stats"]["not_found"] = _ko
                st.session_state["_3eme_lame_result"] = f"3ème lame terminée : {_ok} enrichi(s), {_ko} non trouvé(s)"
                st.rerun()
            # Afficher le résultat persistant
            if st.session_state.get("_3eme_lame_result"):
                st.success(st.session_state["_3eme_lame_result"])

        # --- Suggestions Sirene (clients non trouves automatiquement) ---
        suggestions = st.session_state.get("meg_sirene_suggestions", {})
        if suggestions:
            st.divider()
            st.subheader(f"🔎 {len(suggestions)} client(s) a confirmer manuellement")
            st.caption("Pour chaque client non identifie automatiquement, selectionnez la bonne entreprise ou ignorez.")

            applied = 0
            for idx, sug_data in sorted(suggestions.items()):
                client_nom = sug_data["client"]
                client_code = sug_data["code"]
                sugs = sug_data["suggestions"]
                options = ["— Ignorer"] + [f"{s['nom']} | SIREN {s['siren']} | {s['ville']} | {s['activite']}" for s in sugs]
                chosen = st.selectbox(f"**{client_nom}** (code: {client_code})", options, key=f"sug_{idx}")

                if chosen != "— Ignorer":
                    chosen_idx = options.index(chosen) - 1
                    if st.button(f"✅ Appliquer pour {client_nom}", key=f"sug_apply_{idx}"):
                        # Appliquer l'enrichissement du choix
                        ent = sugs[chosen_idx]["_raw"]
                        siege = ent.get("siege", {})
                        siren = ent.get("siren", ""); siret = siege.get("siret", "")
                        nom_t = ent.get("nom_complet", ent.get("nom_raison_sociale", ""))

                        sirene_info[idx] = {"nom": nom_t, "activite": _naf_label(siege.get("activite_principale", ent.get("activite_principale", ""))), "ville": siege.get("libelle_commune", "")}
                        sirene_cells_tmp = set(sirene_cells)

                        def _apply(col, val):
                            if col in df_preview_c.columns and val:
                                cur = to_clean_str(df_preview_c.at[idx, col])
                                if not cur or cur == "NC":
                                    df_preview_c.at[idx, col] = val
                                    sirene_cells_tmp.add((idx, col))
                        _apply("Siren", siren); _apply("Siret", siret)
                        if siren and "Type *" in df_preview_c.columns and df_preview_c.at[idx, "Type *"] != "Professionnel":
                            df_preview_c.at[idx, "Type *"] = "Professionnel"; sirene_cells_tmp.add((idx, "Type *"))
                        _apply("APE / NAF", ent.get("activite_principale", ""))
                        _apply("Forme juridique", _normalize_forme_juridique(ent.get("nature_juridique", "")))
                        if siren and "TVA intracommunautaire" in df_preview_c.columns:
                            cur = to_clean_str(df_preview_c.at[idx, "TVA intracommunautaire"])
                            if not cur or cur == "NC":
                                tv = f"FR{(12+3*(int(siren)%97))%97:02d}{siren}"
                                df_preview_c.at[idx, "TVA intracommunautaire"] = tv
                                sirene_cells_tmp.add((idx, "TVA intracommunautaire"))
                        if not to_clean_str(df_preview_c.at[idx, "Adresse"]) if "Adresse" in df_preview_c.columns else True:
                            pts = [siege.get("numero_voie",""), siege.get("type_voie",""), siege.get("libelle_voie","")]
                            a = " ".join(p for p in pts if p)
                            if a and "Adresse" in df_preview_c.columns:
                                df_preview_c.at[idx, "Adresse"] = a; sirene_cells_tmp.add((idx, "Adresse"))
                        _apply("Code postal *", siege.get("code_postal", ""))
                        _apply("Ville *", siege.get("libelle_commune", ""))

                        st.session_state["meg_df_clients"] = df_preview_c
                        st.session_state["meg_sirene_cells"] = sirene_cells_tmp
                        st.session_state["meg_sirene_info"] = sirene_info
                        # Retirer des suggestions
                        del st.session_state["meg_sirene_suggestions"][idx]
                        # Regenerer le workbook
                        wb_n, ws_n = _make_wb(_H_ENTITY)
                        for _, rw in df_preview_c.iterrows():
                            ws_n.append([rw.iloc[i] if not pd.isna(rw.iloc[i]) else None for i in range(len(rw))])
                        with open(GABARIT_CLIENT_PATH, "wb") as f: f.write(_wb_bytes(wb_n))
                        st.session_state["meg_editor_ver"] = st.session_state.get("meg_editor_ver", 0) + 1
                        applied += 1
                        st.rerun()

        # --- Etape 4 : Vérification des champs obligatoires manquants ---
        sirene_cells_per_row = {r for (r,_) in sirene_cells}
        st.divider()
        st.subheader("⚠️ Vérification avant injection")

        # Champs requis par l'API Evoliz pour créer un client
        _required_api = {"Societe / Nom *": "name", "Code *": "code", "Type *": "type",
                         "Code postal *": "postcode", "Ville *": "town", "Code pays (ISO 2) *": "iso2"}
        _missing_rows = []
        for idx, row in df_preview_c.iterrows():
            _manques = []
            for col, api_field in _required_api.items():
                if col in df_preview_c.columns:
                    val = to_clean_str(row[col])
                    if not val or val in ("NC", "nan"):
                        _manques.append(col.replace(" *", ""))
                else:
                    _manques.append(col.replace(" *", ""))
            if _manques:
                _nom = to_clean_str(row.get("Societe / Nom *", "")) if "Societe / Nom *" in df_preview_c.columns else ""
                _code = to_clean_str(row.get("Code *", "")) if "Code *" in df_preview_c.columns else ""
                _missing_rows.append({"idx": idx, "Code": _code, "Nom": _nom, "Champs manquants": ", ".join(_manques)})

        if _missing_rows:
            st.warning(f"⚠️ {len(_missing_rows)} client(s) ont des champs obligatoires manquants — l'injection échouera pour ces lignes.")
            with st.expander(f"📋 {len(_missing_rows)} client(s) incomplets — cliquez pour éditer", expanded=True):
                _df_missing = pd.DataFrame(_missing_rows)
                # Construire un tableau éditable avec les champs manquants
                _edit_rows = []
                for mr in _missing_rows:
                    _row_data = {"Code": mr["Code"], "Nom": mr["Nom"]}
                    _src_row = df_preview_c.loc[mr["idx"]]
                    for col in _required_api:
                        _col_clean = col.replace(" *", "")
                        _row_data[_col_clean] = to_clean_str(_src_row[col]) if col in df_preview_c.columns else ""
                    _row_data["_idx"] = mr["idx"]
                    _edit_rows.append(_row_data)
                _df_edit = pd.DataFrame(_edit_rows)
                _edit_cols = [c for c in _df_edit.columns if c not in ("_idx", "Code", "Nom")]
                _disabled_edit = [c for c in _df_edit.columns if c in ("_idx", "Code", "Nom")]
                _col_cfg = {}
                if "Type" in _df_edit.columns:
                    _col_cfg["Type"] = st.column_config.SelectboxColumn("Type", options=["Particulier", "Professionnel", "Administration publique"])
                _edited_missing = st.data_editor(
                    _df_edit.drop(columns=["_idx"]), use_container_width=True, hide_index=True,
                    disabled=["Code", "Nom"], column_config=_col_cfg,
                    key=f"edit_missing_{st.session_state.get('meg_editor_ver', 0)}")

                _c1_fill, _c2_fill = st.columns([3, 1])
                _type_to_apply = _c2_fill.selectbox("Type à appliquer", ["Particulier", "Professionnel", "Administration publique"], key="fill_type_sel")
                if _c1_fill.button(f"🏷️ Appliquer « {_type_to_apply} » aux clients sans type", use_container_width=True, key="btn_fill_type"):
                    _t_filled = 0
                    if "Type *" in df_preview_c.columns:
                        for idx in df_preview_c.index:
                            _cur_t = to_clean_str(df_preview_c.at[idx, "Type *"])
                            if not _cur_t or _cur_t in ("NC", "nan"):
                                df_preview_c.at[idx, "Type *"] = _type_to_apply
                                _t_filled += 1
                    if _t_filled:
                        st.session_state["meg_df_clients"] = df_preview_c
                        st.session_state["meg_editor_ver"] = st.session_state.get("meg_editor_ver", 0) + 1
                        st.rerun()

                if st.button("🔄 Compléter les vides (CP=00000, Ville=Inconnue, Pays=FR)", use_container_width=True, key="btn_fill_defaults"):
                    _filled = 0
                    for mr in _missing_rows:
                        _orig_idx = mr["idx"]
                        _defaults = {"Code postal *": "00000", "Ville *": "Inconnue", "Code pays (ISO 2) *": "FR"}
                        for col, default_val in _defaults.items():
                            if col in df_preview_c.columns:
                                _cur = to_clean_str(df_preview_c.at[_orig_idx, col])
                                if not _cur or _cur in ("NC", "nan"):
                                    df_preview_c.at[_orig_idx, col] = default_val
                                    _filled += 1
                    if _filled:
                        st.session_state["meg_df_clients"] = df_preview_c
                        st.session_state["meg_editor_ver"] = st.session_state.get("meg_editor_ver", 0) + 1
                        st.rerun()

                # Persistance automatique des modifications (sans bouton)
                _applied = 0
                for _i, _er in enumerate(_edit_rows):
                    _orig_idx = _er["_idx"]
                    for col, api_field in _required_api.items():
                        _col_clean = col.replace(" *", "")
                        if _col_clean in _edited_missing.columns:
                            _new_val = str(_edited_missing.at[_i, _col_clean]).strip()
                            if _new_val and _new_val not in ("NC", "nan", ""):
                                if col in df_preview_c.columns:
                                    _old_val = to_clean_str(df_preview_c.at[_orig_idx, col])
                                    if _new_val != _old_val:
                                        df_preview_c.at[_orig_idx, col] = _new_val
                                        _applied += 1
                if _applied:
                    st.session_state["meg_df_clients"] = df_preview_c
        else:
            st.success("✅ Tous les clients ont les champs obligatoires remplis — prêts pour l'injection.")

        # --- Tableau final : état des données telles qu'elles seront envoyées ---
        st.divider()
        st.subheader(f"📋 Récapitulatif final — {len(df_preview_c)} {_entity_label}")
        _final_cols = ["Source", "Code *", "Societe / Nom *", "Type *", "Code postal *", "Ville *",
                       "Code pays (ISO 2) *", "Siren", "Siret", "TVA intracommunautaire",
                       "Forme juridique", "APE / NAF", "Adresse", "Telephone"]
        _final_cols = [c for c in _final_cols if c in df_preview_c.columns]
        _df_final_show = df_preview_c[_final_cols].copy()
        if "Source" not in _df_final_show.columns:
            _df_final_show.insert(0, "Source", _df_final_show.index.map(lambda i: sources.get(i, "")))
        # Convertir Siren/Siret en str
        for _cs in ["Siren", "Siret"]:
            if _cs in _df_final_show.columns:
                _df_final_show[_cs] = _df_final_show[_cs].apply(lambda v: to_clean_str(v) if not pd.isna(v) else "")
        # Colorer par source
        def _color_final(row):
            src = str(row.get("Source", ""))
            if "Nouveau" in src: return ["background-color: #d4edda"] * len(row)
            if "Doublon" in src: return ["background-color: #fff3cd"] * len(row)
            if "Evoliz seul" in src: return ["background-color: #e8f4fd"] * len(row)
            return [""] * len(row)
        with st.expander(f"👁️ Voir / éditer les {len(df_preview_c)} {_entity_label} tels qu'ils seront envoyés", expanded=False):
            _final_col_config = {}
            if "Type *" in _df_final_show.columns:
                _final_col_config["Type *"] = st.column_config.SelectboxColumn("Type *", options=["Particulier", "Professionnel", "Administration publique"])
            _final_disabled = ["Source"]
            _edited_final = st.data_editor(
                _df_final_show, use_container_width=True, hide_index=True,
                disabled=_final_disabled, column_config=_final_col_config,
                key=f"final_recap_{st.session_state.get('meg_editor_ver', 0)}")
            st.caption("🟢 = Nouveau | 🟡 = Doublon (MAJ) | 🔵 = Evoliz seul (MAJ)")
            # Persister les modifications
            _final_changed = False
            for _i in _edited_final.index:
                for _col in [c for c in _final_cols if c != "Source"]:
                    if _col in _edited_final.columns and _col in df_preview_c.columns:
                        _new_v = str(_edited_final.at[_i, _col]).strip() if not pd.isna(_edited_final.at[_i, _col]) else ""
                        _old_v = to_clean_str(df_preview_c.at[_i, _col]) if not pd.isna(df_preview_c.at[_i, _col]) else ""
                        if _new_v != _old_v and _new_v:
                            df_preview_c.at[_i, _col] = _new_v
                            _final_changed = True
            if _final_changed:
                st.session_state["meg_df_clients"] = df_preview_c

        # --- Injection ---
        st.divider()
        st.subheader(f"🚀 Injection {_entity_label} dans Evoliz")
        inject_btn = st.button(f"🚀 Injecter les {_entity_label}", type="primary", use_container_width=True, key="btn_inject_clients", disabled=not has_api or bool(_missing_rows))
        if inject_btn and has_api:
            headers = st.session_state.token_headers_105; cid = st.session_state.company_id_105
            # URLs avec fallback : prefixe en primaire, sans prefixe en secours (cas company_users avec cid invalide)
            _url_cli_primary = f"https://www.evoliz.io/api/v1/companies/{cid}/{_entity_api}" if cid else f"https://www.evoliz.io/api/v1/{_entity_api}"
            _url_cli_fallback = f"https://www.evoliz.io/api/v1/{_entity_api}"
            url_cli = _url_cli_primary  # pour compat retroactive avec les logs

            def _http_with_fallback(method, suffix, payload):
                """POST/PATCH avec fallback URL sans prefixe si 403/404."""
                for _u in [f"{_url_cli_primary}{suffix}", f"{_url_cli_fallback}{suffix}"]:
                    try:
                        if method == "POST":
                            r = requests.post(_u, headers=headers, json=payload, timeout=15)
                        else:
                            r = requests.patch(_u, headers=headers, json=payload, timeout=15)
                        if r.status_code in (200, 201, 204):
                            return r
                        if r.status_code not in (403, 404):
                            return r
                        if _url_cli_primary == _url_cli_fallback:
                            return r
                    except Exception:
                        pass
                return r
            ev_ids = st.session_state.get("meg_consol_ev_ids", {})
            df_final = df_preview_c.copy(); df_orig = st.session_state.get("meg_df_clients_original")
            for idx in df_final.index:
                if df_orig is not None and idx in sirene_cells_per_row and not enrichir_flags.get(idx, True):
                    if idx < len(df_orig): df_final.iloc[idx] = df_orig.iloc[idx]
            ci_h = {h.split(" *")[0]: i for i, h in enumerate(_H_ENTITY)}
            _name_col = "Raison sociale" if _is_supplier else "Societe / Nom"
            created=updated=up_to_date=skipped_inj=0; errors=[]; inject_log=[]

            # --- Preparation des taches en amont (pas d'IO dans la boucle) ---
            _tasks = []
            for idx, row in df_final.iterrows():
                nom = to_clean_str(row.iloc[ci_h[_name_col]]) if ci_h.get(_name_col) is not None else ""
                if not nom: skipped_inj += 1; continue
                code = to_clean_str(row.iloc[ci_h["Code"]]) if ci_h.get("Code") is not None else ""
                type_c = to_clean_str(row.iloc[ci_h["Type"]]) if ci_h.get("Type") is not None else "Professionnel"
                cp = to_clean_str(row.iloc[ci_h["Code postal"]]) if ci_h.get("Code postal") is not None else ""
                ville = to_clean_str(row.iloc[ci_h["Ville"]]) if ci_h.get("Ville") is not None else ""
                _iso2_raw = to_clean_str(row.iloc[ci_h["Code pays (ISO 2)"]]) if ci_h.get("Code pays (ISO 2)") is not None else ""
                iso2v = _iso2_raw.upper()[:2] if _iso2_raw and len(_iso2_raw) >= 2 and _iso2_raw.isalpha() else "FR"
                payload = {"name":nom,"address":{"postcode":cp if cp and cp!="NC" else "00000","town":ville if ville and ville!="NC" else "NC","iso2":iso2v}}
                if not _is_supplier:
                    payload["type"] = type_c
                if code: payload["code"] = code[:20]
                tva_v = to_clean_str(row.iloc[ci_h["TVA intracommunautaire"]]) if ci_h.get("TVA intracommunautaire") is not None else ""
                payload["vat_number"] = tva_v if tva_v and tva_v != "NC" else "N/C"
                siret_v = to_clean_str(row.iloc[ci_h["Siret"]]) if ci_h.get("Siret") is not None else ""
                payload["business_number"] = siret_v if siret_v and siret_v != "NC" else "N/C"
                for fld, ak in [("Forme juridique","legalform"),("Siren","business_identification_number"),("APE / NAF","activity_number"),("Telephone","phone"),("Portable","mobile"),("Fax","fax"),("Site web","website"),("Commentaires","comment")]:
                    v = to_clean_str(row.iloc[ci_h[fld]]) if ci_h.get(fld) is not None else ""
                    if v and v != "NC": payload[ak] = v
                adr = to_clean_str(row.iloc[ci_h["Adresse"]]) if ci_h.get("Adresse") is not None else ""
                if adr: payload["address"]["addr"] = adr
                client_id = ev_ids.get(idx)
                _tasks.append((idx, code, nom, client_id, payload))

            # --- Execution parallele (5 workers, rate-limit Evoliz 100 req/min) ---
            def _inject_one(task):
                _idx, _code, _nom, _cid_entity, _payload = task
                if _cid_entity:
                    r = _http_with_fallback("PATCH", f"/{_cid_entity}", _payload)
                    action = "🔄 MAJ"
                else:
                    r = _http_with_fallback("POST", "", _payload)
                    action = "➕ Creation"
                return (_idx, _code, _nom, action, r.status_code, r.text[:80])

            progress = st.progress(0.0, text=f"Injection — 0 / {len(_tasks)}")
            _n_done = 0
            with ThreadPoolExecutor(max_workers=5) as pool:
                futures = {pool.submit(_inject_one, t): t for t in _tasks}
                for fut in as_completed(futures):
                    _idx, _code, _nom, action, status, body = fut.result()
                    if action == "🔄 MAJ":
                        if status in (200, 204):
                            updated += 1; inject_log.append({"Code": _code, "Nom": _nom, "Action": action, "Statut": "✅ OK", "Detail": ""})
                        else:
                            errors.append(f"MAJ '{_nom}': HTTP {status}"); inject_log.append({"Code": _code, "Nom": _nom, "Action": action, "Statut": f"❌ {status}", "Detail": body})
                    else:  # Creation
                        if status in (200, 201):
                            created += 1; inject_log.append({"Code": _code, "Nom": _nom, "Action": action, "Statut": "✅ OK", "Detail": ""})
                        elif status == 400 and "already been taken" in body:
                            updated += 1; inject_log.append({"Code": _code, "Nom": _nom, "Action": "🔄 Existant", "Statut": "✅ OK", "Detail": ""})
                        else:
                            errors.append(f"Creation '{_nom}': HTTP {status}"); inject_log.append({"Code": _code, "Nom": _nom, "Action": action, "Statut": f"❌ {status}", "Detail": body})
                    _n_done += 1
                    progress.progress(_n_done / max(len(_tasks), 1), text=f"Injection — {_n_done} / {len(_tasks)} — {_nom[:30]}")
            progress.empty()
            st.divider(); st.subheader("📊 Synthese")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("➕ Crees",created); c2.metric("🔄 MAJ",updated); c3.metric("⏭️ Ignores",skipped_inj); c4.metric("❌ Erreurs",len(errors))
            if created+updated > 0: st.success(f"Injection : {created} cree(s), {updated} mis a jour")
            elif errors: st.error(f"{len(errors)} erreur(s)")
            if inject_log:
                df_il = pd.DataFrame(inject_log)
                def _ci(row):
                    a=str(row.get("Action","")); s=str(row.get("Statut",""))
                    if "OK" in s and "Creation" in a: return ["background-color:#d4edda"]*len(row)
                    if "OK" in s: return ["background-color:#e8f4fd"]*len(row)
                    if "❌" in s: return ["background-color:#f8d7da"]*len(row)
                    return [""]*len(row)
                st.dataframe(df_il.style.apply(_ci, axis=1), use_container_width=True, hide_index=True)


# --- Onglet Injection Fournisseurs ---
if _connected and mod_fournisseurs:
 with m_four:
    if not (st.session_state.get('company_id_105') and st.session_state.get('token_headers_105')):
        st.warning("⛔ Connectez-vous a l'API et selectionnez un dossier (onglet **🔑 Connexion API**) avant d'utiliser cet onglet.")
    st.subheader("🏭 Injection Fournisseurs")
    st.caption("1. Importez un fichier fournisseurs  2. Mapping colonnes  3. Consolidation Evoliz  4. Injection")

    f_meg_four = st.session_state.get("imp_file_fournisseurs")
    if not f_meg_four:
        st.info("Importez d'abord un fichier fournisseurs dans l'onglet 📁 Import fichiers.")

    if f_meg_four:
        # Lecture fichier
        if f_meg_four.name.lower().endswith(".csv"):
            df_four = None
            for _enc in ["utf-8", "latin-1", "cp1252"]:
                for _sep in [None, ";", ","]:
                    try:
                        f_meg_four.seek(0)
                        _kw = {"header": 0, "encoding": _enc}
                        if _sep: _kw["sep"] = _sep
                        else: _kw["sep"] = None; _kw["engine"] = "python"
                        df_four = pd.read_csv(f_meg_four, **_kw)
                        if len(df_four.columns) > 1: break
                        df_four = None
                    except Exception:
                        df_four = None
                if df_four is not None: break
            if df_four is None:
                st.error("Impossible de lire ce fichier CSV.")
                f_meg_four = None
        else:
            _four_sheet = st.session_state.get("imp_file_fournisseurs_sheet", 0)
            df_four = _read_meg(f_meg_four, sheet_name=_four_sheet)

        if f_meg_four and df_four is not None:
            st.caption(f"Fichier lu : **{len(df_four)} lignes**, **{len(df_four.columns)} colonnes**")
            st.dataframe(df_four.head(5), use_container_width=True, hide_index=True)

            # --- Mapping colonnes ---
            all_four_fields = ["— Ignorer",
                "Raison sociale", "Code", "Code postal", "Ville", "Code pays (ISO 2)",
                "Forme juridique", "Siret", "APE / NAF", "TVA intracommunautaire",
                "Adresse", "Adresse (suite)", "Pays",
                "Telephone", "Portable", "Fax", "E-mail", "Site web",
                "Classification", "Code classification", "Commentaires"]

            # Auto-mapping
            _four_file_sig = f_meg_four.name + ",".join(df_four.columns.tolist())
            if st.session_state.get("_four_last_sig") != _four_file_sig:
                fwd_f = _auto_map_columns(df_four.columns.tolist())
                # Adapter les noms : "Societe / Nom" -> "Raison sociale"
                rev_f = {}
                for ef, sc in fwd_f.items():
                    mapped = ef
                    if ef == "Societe / Nom": mapped = "Raison sociale"
                    if ef == "Complement d'adresse": mapped = "Adresse (suite)"
                    if mapped in all_four_fields and sc not in rev_f:
                        rev_f[sc] = mapped
                st.session_state["_four_col_rev"] = rev_f
                st.session_state["_four_last_sig"] = _four_file_sig

            rev_map_f = st.session_state.get("_four_col_rev", {})

            st.divider()
            st.subheader("🔗 Mapping des colonnes")
            mapping_rev_f = {}
            n_cols_f = len(df_four.columns)
            for i in range(0, n_cols_f, 2):
                ui_cols_f = st.columns(2)
                for j in range(2):
                    if i + j >= n_cols_f: break
                    src_col = df_four.columns[i + j]
                    default_ev = rev_map_f.get(src_col, "— Ignorer")
                    idx_f = all_four_fields.index(default_ev) if default_ev in all_four_fields else 0
                    with ui_cols_f[j]:
                        chosen_f = st.selectbox(f"📄 **{src_col}**", all_four_fields, index=idx_f, key=f"fmap_{i+j}")
                        mapping_rev_f[src_col] = chosen_f

            mapping_f = {}
            for sc, ef in mapping_rev_f.items():
                if ef != "— Ignorer" and ef not in mapping_f:
                    mapping_f[ef] = sc

            # Controle champs obligatoires
            st.divider()
            st.subheader("📋 Controle des champs obligatoires")
            _four_required = ["Raison sociale", "Code"]
            ctrl_f = []
            for ef in _four_required:
                if ef in mapping_f:
                    ctrl_f.append({"Champ": ef, "Statut": "✅ Mappé", "Source": mapping_f[ef]})
                elif ef == "Code":
                    ctrl_f.append({"Champ": ef, "Statut": "🔄 Auto-généré", "Source": "Généré depuis le nom"})
                else:
                    ctrl_f.append({"Champ": ef, "Statut": "❌ Manquant", "Source": ""})
            st.dataframe(pd.DataFrame(ctrl_f), use_container_width=True, hide_index=True)

            # --- Consolidation + Injection ---
            has_api_f = bool(st.session_state.get("token_headers_105"))
            _four_file_id = f_meg_four.name + str(f_meg_four.size)
            _four_already = st.session_state.get("_four_consol_id") == _four_file_id
            _four_auto = not _four_already
            _four_manual = st.button("🔄 Consolider avec Evoliz", use_container_width=True, key="btn_consol_four")

            if has_api_f and (_four_auto or _four_manual):
                st.session_state["_four_consol_id"] = _four_file_id
                headers_f = st.session_state.token_headers_105
                cid_f = st.session_state.company_id_105
                if not cid_f:
                    st.error("Aucun dossier sélectionné.")
                    st.stop()

                with st.spinner("Lecture des fournisseurs Evoliz..."):
                    ev_four = []; ev_f_by_name = {}
                    # URL avec fallback (mono/multi indifferent)
                    _url_f_pri = f"https://www.evoliz.io/api/v1/companies/{cid_f}/suppliers"
                    _url_f_fb = "https://www.evoliz.io/api/v1/suppliers"
                    url_four = _url_f_pri
                    try:
                        _rt = requests.get(_url_f_pri, headers=headers_f, params={"per_page": 1, "page": 1}, timeout=10)
                        if _rt.status_code in (403, 404):
                            url_four = _url_f_fb
                    except Exception:
                        url_four = _url_f_fb
                    page_f = 1
                    while True:
                        r_f = requests.get(url_four, headers=headers_f, params={"per_page": 100, "page": page_f}, timeout=15)
                        if r_f.status_code != 200: break
                        d_f = r_f.json()
                        for it in d_f.get("data", []):
                            adr = it.get("address") or {}
                            entry_f = {
                                "supplierid": it.get("supplierid"), "code": (it.get("code") or "").strip(),
                                "name": (it.get("name") or "").strip(),
                                "business_number": (it.get("business_number") or ""),
                                "business_identification_number": (it.get("business_identification_number") or ""),
                                "vat_number": (it.get("vat_number") or ""),
                                "legalform": (it.get("legal_status") or {}).get("label", "") if isinstance(it.get("legal_status"), dict) else "",
                                "activity_number": (it.get("activity_number") or ""),
                                "phone": (it.get("phone") or ""), "mobile": (it.get("mobile") or ""),
                                "fax": (it.get("fax") or ""), "website": (it.get("website") or ""),
                                "addr": (adr.get("addr") or ""), "postcode": (adr.get("postcode") or ""),
                                "town": (adr.get("town") or ""), "iso2": (adr.get("iso2") or ""),
                            }
                            ev_four.append(entry_f)
                            n_f = norm_piv(entry_f["name"])
                            if n_f: ev_f_by_name[n_f] = entry_f
                        if page_f >= d_f.get("meta", {}).get("last_page", 1): break
                        page_f += 1

                # Construction liste consolidée
                consol_four = []; seen_four_ids = set()
                for _, row in df_four.iterrows():
                    nom = to_clean_str(row.get(mapping_f.get("Raison sociale", ""), "")) if "Raison sociale" in mapping_f else ""
                    if not nom: continue
                    _raw_code = to_clean_str(row.get(mapping_f.get("Code", ""), "")) if "Code" in mapping_f else ""
                    code = _raw_code if _raw_code and len(_raw_code) <= 20 else norm_piv(nom)[:15]
                    entry_c = {"Code": code, "Raison sociale": nom, "_source": "📄 Nouveau", "_entityid": None}
                    for fld in ["Siret", "APE / NAF", "TVA intracommunautaire", "Forme juridique",
                                "Adresse", "Adresse (suite)", "Code postal", "Ville", "Code pays (ISO 2)",
                                "Telephone", "Portable", "Fax", "Site web", "Commentaires"]:
                        col = mapping_f.get(fld)
                        entry_c[fld] = to_clean_str(row.get(col, "")) if col and col in df_four.columns else ""
                    if not entry_c.get("Code pays (ISO 2)"): entry_c["Code pays (ISO 2)"] = "FR"
                    ev = ev_f_by_name.get(norm_piv(nom))
                    if ev:
                        seen_four_ids.add(ev["supplierid"])
                        entry_c["_source"] = "📄+☁️ Doublon"
                        entry_c["_entityid"] = ev["supplierid"]
                    consol_four.append(entry_c)

                for ev in ev_four:
                    if ev["supplierid"] not in seen_four_ids:
                        consol_four.append({
                            "Code": ev["code"], "Raison sociale": ev["name"],
                            "Siret": ev["business_number"], "TVA intracommunautaire": ev["vat_number"],
                            "Forme juridique": ev["legalform"], "APE / NAF": ev["activity_number"],
                            "Adresse": ev["addr"], "Adresse (suite)": "",
                            "Code postal": ev["postcode"], "Ville": ev["town"], "Code pays (ISO 2)": ev["iso2"],
                            "Telephone": ev["phone"], "Portable": ev["mobile"],
                            "Fax": ev["fax"], "Site web": ev["website"], "Commentaires": "",
                            "_source": "☁️ Evoliz seul", "_entityid": ev["supplierid"],
                        })

                st.session_state["_four_consol"] = consol_four
                st.session_state["_four_consol_stats"] = {
                    "fichier": sum(1 for c in consol_four if "Nouveau" in c["_source"]),
                    "doublons": sum(1 for c in consol_four if "Doublon" in c["_source"]),
                    "evoliz_seul": sum(1 for c in consol_four if "Evoliz seul" in c["_source"]),
                    "total_evoliz": len(ev_four),
                }
                st.rerun()

            # Affichage consolidation
            consol_four = st.session_state.get("_four_consol", [])
            stats_f = st.session_state.get("_four_consol_stats")
            if stats_f:
                st.divider()
                st.subheader("📊 Consolidation")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("📄 Nouveaux", stats_f["fichier"])
                c2.metric("📄+☁️ Doublons", stats_f["doublons"])
                c3.metric("☁️ Evoliz seul", stats_f["evoliz_seul"])
                c4.metric("☁️ Total Evoliz", stats_f["total_evoliz"])

            if consol_four:
                df_four_preview = pd.DataFrame([{k: v for k, v in c.items() if not k.startswith("_")} for c in consol_four])
                _sources_f = [c["_source"] for c in consol_four]
                df_four_preview.insert(0, "Source", _sources_f)
                _four_ev_ids = {i: c.get("_entityid") for i, c in enumerate(consol_four)}

                # Marquage visuel des cellules enrichies via Sirene (prefixe 🟢)
                _four_cells = st.session_state.get("_four_sirene_cells", set())
                if _four_cells:
                    for idx in df_four_preview.index:
                        for col in df_four_preview.columns:
                            if (idx, col) in _four_cells:
                                val = df_four_preview.at[idx, col]
                                if val is not None and str(val).strip() and not str(val).startswith("🟢"):
                                    df_four_preview.at[idx, col] = f"🟢 {val}"
                st.dataframe(df_four_preview, use_container_width=True, hide_index=True)

                # --- Option : code fournisseur = SIREN ---
                _code_is_siren = st.checkbox(
                    "🔢 Le champ « Code » contient un SIREN — rechercher directement sur ce numero",
                    value=st.session_state.get("_four_code_is_siren", False),
                    key="_four_code_is_siren",
                    help="Active si les codes fournisseurs sont en fait des SIREN (9 chiffres). La recherche Sirene utilisera ce numero au lieu du nom.",
                )

                # --- Enrichissement Sirene fournisseurs ---
                st.divider()
                if st.button("🔍 Enrichir fournisseurs via Sirene", type="primary", use_container_width=True, key="btn_sirene_four"):
                    _f_ok = _f_ko = _f_skip = _f_complete = 0
                    _total_f = len(consol_four)

                    # 1) Preparer les taches (exclure les fournisseurs sans nom ou deja complets)
                    _tasks_f = []
                    for i, cr in enumerate(consol_four):
                        nom = cr.get("Raison sociale", "")
                        code = to_clean_str(cr.get("Code", ""))
                        if not nom and not (_code_is_siren and code): _f_skip += 1; continue
                        _has_all = all(cr.get(f) and cr.get(f) not in ("", "NC") for f in ["Siret", "Code postal", "Ville"])
                        if _has_all: _f_complete += 1; continue
                        # Si l'option SIREN cochee : extraire le SIREN du code (format "F123456789" ou "123456789")
                        _siren_candidat = ""
                        if _code_is_siren and code:
                            _c = code.upper().lstrip("F").strip()
                            if _c.isdigit() and len(_c) == 9:
                                _siren_candidat = _c
                        if _siren_candidat:
                            _tasks_f.append((i, _siren_candidat, nom or code, "siren"))
                        else:
                            sq = " ".join(nom.replace("nan", "").split()).strip()
                            if not sq: _f_skip += 1; continue
                            _tasks_f.append((i, sq, nom, "nom"))

                    # 2) Recherche Sirene parallele (7 workers) - par SIREN ou par nom
                    def _search_four(task):
                        idx, query, nom, mode = task
                        try:
                            r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                             params={"q": query, "per_page": 1, "page": 1}, timeout=10)
                            if r.status_code == 429:
                                time.sleep(2)
                                r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                                 params={"q": query, "per_page": 1, "page": 1}, timeout=10)
                            return (idx, nom, mode, r.status_code, r.json() if r.status_code == 200 else None)
                        except Exception as exc:
                            return (idx, nom, mode, -1, str(exc))

                    _pg_bar = st.progress(0.0, text=f"Enrichissement Sirene — 0 / {len(_tasks_f)} (7 workers)")
                    _results_f = []
                    if _tasks_f:
                        with ThreadPoolExecutor(max_workers=7) as pool:
                            futures_f = {pool.submit(_search_four, t): t for t in _tasks_f}
                            for _n, fut in enumerate(as_completed(futures_f)):
                                _results_f.append(fut.result())
                                _pg_bar.progress((_n + 1) / len(_tasks_f),
                                                   text=f"Enrichissement Sirene — {_n+1} / {len(_tasks_f)}")

                    # 3) Application des resultats (sequentiel, ecriture dans consol_four)
                    _four_cells = set(st.session_state.get("_four_sirene_cells", set()))
                    for idx, nom, mode, status, data in _results_f:
                        if status != 200 or not data:
                            _f_ko += 1; continue
                        results = data.get("results", [])
                        if not results:
                            _f_ko += 1; continue
                        ent = results[0]; siege = ent.get("siege", {})
                        # Par SIREN : identification exacte, pas de controle de similarite
                        if mode != "siren":
                            nom_n = norm_piv(nom); best_n = norm_piv(ent.get("nom_complet", ""))
                            if not (nom_n == best_n or nom_n in best_n or best_n in nom_n or
                                    (len(nom_n) > 3 and len(best_n) > 3 and len(set(nom_n) & set(best_n)) / max(len(set(nom_n)), len(set(best_n))) > 0.6)):
                                _f_ko += 1; continue
                        cr = consol_four[idx]
                        for fld, val in [("Raison sociale", ent.get("nom_complet", ent.get("nom_raison_sociale", ""))),
                                         ("Siret", siege.get("siret", "")), ("Code postal", siege.get("code_postal", "")),
                                         ("Ville", siege.get("libelle_commune", "")), ("APE / NAF", ent.get("activite_principale", "")),
                                         ("Forme juridique", _normalize_forme_juridique(ent.get("nature_juridique", "")))]:
                            if val and (not cr.get(fld) or cr.get(fld) in ("", "NC")):
                                consol_four[idx][fld] = val
                                _four_cells.add((idx, fld))
                        siren = ent.get("siren", "")
                        if siren and (not cr.get("TVA intracommunautaire") or cr.get("TVA intracommunautaire") in ("", "NC")):
                            try:
                                consol_four[idx]["TVA intracommunautaire"] = f"FR{(12+3*(int(siren)%97))%97:02d}{siren}"
                                _four_cells.add((idx, "TVA intracommunautaire"))
                            except ValueError: pass
                        if not cr.get("Adresse"):
                            pts = [siege.get("numero_voie", ""), siege.get("type_voie", ""), siege.get("libelle_voie", "")]
                            a = " ".join(p for p in pts if p)
                            if a:
                                consol_four[idx]["Adresse"] = a
                                _four_cells.add((idx, "Adresse"))
                        _f_ok += 1

                    _pg_bar.progress(1.0, text=f"✅ Termine — {len(_tasks_f)} / {len(_tasks_f)}")
                    st.session_state["_four_consol"] = consol_four
                    st.session_state["_four_sirene_cells"] = _four_cells
                    st.session_state["_four_sirene_result"] = f"Enrichissement : {_f_ok} enrichi(s), {_f_complete} deja complet(s), {_f_ko} non trouve(s), {_f_skip} ignore(s)"
                    st.rerun()
                if st.session_state.get("_four_sirene_result"):
                    st.success(st.session_state["_four_sirene_result"])

                # --- 2ème lame : propositions Sirene pour les fournisseurs non identifies ---
                _four_cells_cur = st.session_state.get("_four_sirene_cells", set())
                _four_enriched_rows = {r for (r, _c) in _four_cells_cur}
                _non_enrichis_f = []
                for i, cr in enumerate(consol_four):
                    nom = cr.get("Raison sociale", "")
                    code = cr.get("Code", "")
                    # non enrichi ET sans Siret connu
                    if i not in _four_enriched_rows and nom and nom != "nan" and not cr.get("Siret"):
                        _non_enrichis_f.append((i, nom, code))

                if _non_enrichis_f:
                    st.divider()
                    st.subheader(f"🔍 2ème lame — {len(_non_enrichis_f)} fournisseur(s) non identifie(s)")
                    st.caption("Les 2 resultats Sirene les plus probables sont proposes. Selectionnez le bon ou ignorez.")
                    if st.button(f"🔍 Rechercher les {len(_non_enrichis_f)} propositions", type="primary", use_container_width=True, key="btn_2eme_lame_four"):
                        def _search_2eme_four(task):
                            idx, nom, code = task
                            try:
                                sq = " ".join(nom.replace("nan", "").split()).strip()
                                if not sq: return (idx, nom, code, [])
                                r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                                 params={"q": sq, "per_page": 2, "page": 1}, timeout=10)
                                if r.status_code == 429:
                                    time.sleep(2)
                                    r = requests.get("https://recherche-entreprises.api.gouv.fr/search",
                                                     params={"q": sq, "per_page": 2, "page": 1}, timeout=10)
                                if r.status_code != 200: return (idx, nom, code, [])
                                results = r.json().get("results", [])
                                props = []
                                for res in results[:2]:
                                    rsie = res.get("siege") or {}
                                    props.append({
                                        "nom": res.get("nom_complet", res.get("nom_raison_sociale", "")),
                                        "siren": res.get("siren", ""),
                                        "ville": rsie.get("libelle_commune", ""),
                                        "activite": _naf_label(rsie.get("activite_principale", "")),
                                        "_raw": res,
                                    })
                                return (idx, nom, code, props)
                            except Exception:
                                return (idx, nom, code, [])

                        _all_props_f = []
                        _pg_2 = st.progress(0.0, text=f"Recherche propositions Sirene — 0 / {len(_non_enrichis_f)}")
                        with ThreadPoolExecutor(max_workers=7) as pool:
                            futures_2 = {pool.submit(_search_2eme_four, t): t for t in _non_enrichis_f}
                            for _n, fut in enumerate(as_completed(futures_2)):
                                _all_props_f.append(fut.result())
                                _pg_2.progress((_n + 1) / len(_non_enrichis_f),
                                                 text=f"Recherche propositions Sirene — {_n+1} / {len(_non_enrichis_f)}")
                        _props_dict_f = {}
                        for idx, nom, code, props in _all_props_f:
                            if props:
                                _props_dict_f[idx] = {"fournisseur": nom, "code": code, "propositions": props}
                        st.session_state["_2eme_lame_four_props"] = _props_dict_f
                        st.session_state["_2eme_lame_four_result"] = f"2eme lame : {len(_props_dict_f)} fournisseur(s) avec propositions sur {len(_non_enrichis_f)} recherche(s)"
                        st.rerun()

                    if st.session_state.get("_2eme_lame_four_result"):
                        st.success(st.session_state["_2eme_lame_four_result"])

                    # Afficher les propositions pour validation
                    _props_f = st.session_state.get("_2eme_lame_four_props", {})
                    if _props_f:
                        st.divider()
                        st.subheader(f"📋 {len(_props_f)} proposition(s) a valider")
                        _accepted_f = []
                        for idx, info in sorted(_props_f.items()):
                            props = info["propositions"]
                            options = ["— Ignorer"] + [
                                f"{p['nom']} | SIREN {p['siren']} | {p['ville']} | {p['activite']}" for p in props
                            ]
                            sel = st.selectbox(f"**{info['fournisseur']}** ({info['code']})", options, key=f"prop2_four_{idx}")
                            if sel != "— Ignorer":
                                _sel_idx = options.index(sel) - 1
                                _accepted_f.append((idx, props[_sel_idx]))

                        if _accepted_f:
                            if st.button(f"✅ Appliquer {len(_accepted_f)} selection(s)", type="primary", use_container_width=True, key="btn_apply_2eme_four"):
                                _four_cells2 = set(st.session_state.get("_four_sirene_cells", set()))
                                for idx, prop in _accepted_f:
                                    ent = prop["_raw"]; siege = ent.get("siege", {})
                                    siren = ent.get("siren", ""); siret = siege.get("siret", "")
                                    cr = consol_four[idx]
                                    def _upd_f(fld, val, _idx=idx):
                                        if val and (not cr.get(fld) or cr.get(fld) in ("", "NC")):
                                            consol_four[_idx][fld] = val
                                            _four_cells2.add((_idx, fld))
                                    _upd_f("Siret", siret)
                                    _upd_f("Code postal", siege.get("code_postal", ""))
                                    _upd_f("Ville", siege.get("libelle_commune", ""))
                                    _upd_f("APE / NAF", ent.get("activite_principale", ""))
                                    _upd_f("Forme juridique", _normalize_forme_juridique(ent.get("nature_juridique", "")))
                                    if siren:
                                        try:
                                            tv = f"FR{(12 + 3 * (int(siren) % 97)) % 97:02d}{siren}"
                                            _upd_f("TVA intracommunautaire", tv)
                                        except ValueError: pass
                                    pts = [siege.get("numero_voie", ""), siege.get("type_voie", ""), siege.get("libelle_voie", "")]
                                    _upd_f("Adresse", " ".join(p for p in pts if p))
                                # Retirer les props appliquees
                                for idx, _ in _accepted_f:
                                    if idx in st.session_state["_2eme_lame_four_props"]:
                                        del st.session_state["_2eme_lame_four_props"][idx]
                                st.session_state["_four_consol"] = consol_four
                                st.session_state["_four_sirene_cells"] = _four_cells2
                                st.rerun()

                # --- Vérification champs obligatoires ---
                st.divider()
                st.subheader("⚠️ Vérification avant injection")
                _four_required = {"Raison sociale": "name", "Code": "code", "Code postal": "postcode", "Ville": "town", "Code pays (ISO 2)": "iso2"}
                _four_missing = []
                for i, cr in enumerate(consol_four):
                    _manques = []
                    for fld in _four_required:
                        val = cr.get(fld, "")
                        if not val or val in ("NC", "nan"):
                            _manques.append(fld)
                    if _manques:
                        _four_missing.append({"idx": i, "Code": cr.get("Code",""), "Nom": cr.get("Raison sociale",""), "Champs manquants": ", ".join(_manques)})

                if _four_missing:
                    st.warning(f"⚠️ {len(_four_missing)} fournisseur(s) avec champs obligatoires manquants.")
                    if st.button("🔄 Compléter les vides (CP=00000, Ville=Inconnue, Pays=FR)", use_container_width=True, key="btn_fill_four"):
                        for mr in _four_missing:
                            _defaults_f = {"Code postal": "00000", "Ville": "Inconnue", "Code pays (ISO 2)": "FR"}
                            for fld, dval in _defaults_f.items():
                                _cur = consol_four[mr["idx"]].get(fld, "")
                                if not _cur or _cur in ("NC", "nan"):
                                    consol_four[mr["idx"]][fld] = dval
                        st.session_state["_four_consol"] = consol_four
                        st.rerun()
                    st.dataframe(pd.DataFrame(_four_missing)[["Code", "Nom", "Champs manquants"]], use_container_width=True, hide_index=True)
                else:
                    st.success("✅ Tous les fournisseurs ont les champs obligatoires remplis.")

                # --- Récapitulatif final éditable ---
                st.divider()
                st.subheader(f"📋 Récapitulatif final — {len(consol_four)} fournisseurs")
                _four_show_cols = ["Source", "Code", "Raison sociale", "Code postal", "Ville", "Code pays (ISO 2)",
                                   "Siret", "TVA intracommunautaire", "Forme juridique", "APE / NAF", "Adresse", "Telephone"]
                _df_four_final = pd.DataFrame([{k: v for k, v in c.items() if not k.startswith("_")} for c in consol_four])
                if "Source" not in _df_four_final.columns:
                    _df_four_final.insert(0, "Source", [c["_source"] for c in consol_four])
                _four_show_cols = [c for c in _four_show_cols if c in _df_four_final.columns]
                _df_four_final = _df_four_final[_four_show_cols]
                for _cs in ["Siret"]:
                    if _cs in _df_four_final.columns:
                        _df_four_final[_cs] = _df_four_final[_cs].apply(lambda v: to_clean_str(v) if not pd.isna(v) else "")
                with st.expander(f"👁️ Voir / éditer les {len(consol_four)} fournisseurs", expanded=False):
                    _four_edited = st.data_editor(_df_four_final, use_container_width=True, hide_index=True,
                                                   disabled=["Source"], key=f"four_recap_{st.session_state.get('meg_editor_ver', 0)}")
                    # Persister les modifications
                    for _i in _four_edited.index:
                        for _col in [c for c in _four_show_cols if c != "Source"]:
                            if _col in _four_edited.columns:
                                _new_v = str(_four_edited.at[_i, _col]).strip() if not pd.isna(_four_edited.at[_i, _col]) else ""
                                _old_v = consol_four[_i].get(_col, "")
                                if _new_v != _old_v and _new_v:
                                    consol_four[_i][_col] = _new_v
                                    st.session_state["_four_consol"] = consol_four

                # Injection
                st.divider()
                st.subheader("🚀 Injection fournisseurs dans Evoliz")
                inject_four_btn = st.button("🚀 Injecter les fournisseurs", type="primary", use_container_width=True, key="btn_inject_four", disabled=not has_api_f or bool(_four_missing))
                if inject_four_btn and has_api_f:
                    headers_f = st.session_state.token_headers_105; cid_f = st.session_state.company_id_105
                    # URL avec fallback (mono/multi indifferent)
                    _url_fp_pri = f"https://www.evoliz.io/api/v1/companies/{cid_f}/suppliers"
                    _url_fp_fb = "https://www.evoliz.io/api/v1/suppliers"
                    url_four = _url_fp_pri
                    def _http_four(method, suffix, payload):
                        for _u in [f"{_url_fp_pri}{suffix}", f"{_url_fp_fb}{suffix}"]:
                            try:
                                if method == "POST":
                                    r = requests.post(_u, headers=headers_f, json=payload, timeout=15)
                                else:
                                    r = requests.patch(_u, headers=headers_f, json=payload, timeout=15)
                                if r.status_code in (200, 201, 204): return r
                                if r.status_code not in (403, 404): return r
                            except Exception:
                                pass
                        return r
                    created_f=updated_f=skipped_f=0; errors_f=[]; inject_log_f=[]
                    progress_f = st.progress(0, text="Injection fournisseurs...")
                    for i, cr in enumerate(consol_four):
                        nom = cr.get("Raison sociale", "")
                        if not nom: skipped_f += 1; continue
                        code = cr.get("Code", "")
                        cp = cr.get("Code postal", ""); ville = cr.get("Ville", "")
                        iso2 = cr.get("Code pays (ISO 2)", "FR").upper()[:2] if cr.get("Code pays (ISO 2)") else "FR"
                        payload_f = {
                            "name": nom,
                            "address": {"postcode": cp or "00000", "town": ville or "NC", "iso2": iso2},
                        }
                        if code: payload_f["code"] = code[:20]
                        tva_v = cr.get("TVA intracommunautaire", "")
                        payload_f["vat_number"] = tva_v if tva_v and tva_v != "NC" else "N/C"
                        siret_v = cr.get("Siret", "")
                        payload_f["business_number"] = siret_v if siret_v and siret_v != "NC" else "N/C"
                        for fld, ak in [("Forme juridique","legalform"),("APE / NAF","activity_number"),
                                        ("Telephone","phone"),("Portable","mobile"),("Fax","fax"),
                                        ("Site web","website"),("Commentaires","comment")]:
                            v = cr.get(fld, "")
                            if v and v != "NC": payload_f[ak] = v
                        adr = cr.get("Adresse", "")
                        if adr: payload_f["address"]["addr"] = adr

                        entity_id = cr.get("_entityid")
                        if entity_id:
                            r_f = _http_four("PATCH", f"/{entity_id}", payload_f)
                            if r_f.status_code in (200, 204): updated_f += 1; inject_log_f.append({"Code": code, "Nom": nom, "Action": "🔄 MAJ", "Statut": "✅ OK", "Detail": ""})
                            else: errors_f.append(f"MAJ '{nom}': HTTP {r_f.status_code}"); inject_log_f.append({"Code": code, "Nom": nom, "Action": "🔄 MAJ", "Statut": f"❌ {r_f.status_code}", "Detail": r_f.text[:80]})
                        else:
                            r_f = _http_four("POST", "", payload_f)
                            if r_f.status_code in (200, 201): created_f += 1; inject_log_f.append({"Code": code, "Nom": nom, "Action": "➕ Creation", "Statut": "✅ OK", "Detail": ""})
                            elif r_f.status_code == 400 and "already been taken" in r_f.text: updated_f += 1; inject_log_f.append({"Code": code, "Nom": nom, "Action": "🔄 Existant", "Statut": "✅ OK", "Detail": ""})
                            else: errors_f.append(f"Creation '{nom}': HTTP {r_f.status_code}"); inject_log_f.append({"Code": code, "Nom": nom, "Action": "➕ Creation", "Statut": f"❌ {r_f.status_code}", "Detail": r_f.text[:80]})
                        progress_f.progress((i + 1) / len(consol_four), text=f"{i + 1}/{len(consol_four)} - {nom[:30]}")
                        time.sleep(0.15)
                    progress_f.empty()
                    st.divider(); st.subheader("📊 Synthese")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("➕ Créés", created_f); c2.metric("🔄 MAJ", updated_f); c3.metric("⏭️ Ignorés", skipped_f); c4.metric("❌ Erreurs", len(errors_f))
                    if created_f + updated_f > 0: st.success(f"Injection : {created_f} créé(s), {updated_f} mis à jour")
                    elif errors_f: st.error(f"{len(errors_f)} erreur(s)")
                    if inject_log_f:
                        df_il_f = pd.DataFrame(inject_log_f)
                        st.dataframe(df_il_f, use_container_width=True, hide_index=True)


# --- Onglet Bascule Factures ---
if _connected and mod_factures:
 with m_fac:
    if not (st.session_state.get('company_id_105') and st.session_state.get('token_headers_105')):
        st.warning("⛔ Connectez-vous a l'API et selectionnez un dossier (onglet **🔑 Connexion API**) avant d'utiliser cet onglet.")
    st.subheader("🧾 Bascule Factures, Avoirs & Paiements MEG")
    f_meg_fac = st.session_state.get("imp_file_factures")
    f_meg_cli2 = None  # utilise le gabarit genere a l'etape clients
    if not f_meg_fac:
        st.info("Importez d'abord un fichier factures dans l'onglet 📁 Import fichiers.")
    c1, c2 = st.columns(2)
    meg_tva1 = c1.number_input("TVA 1 (%)", value=20.0, step=0.5, key="meg_tva1")
    meg_tva2 = c2.number_input("TVA 2 (%)", value=0.0, step=0.5, key="meg_tva2")
    c3, c4 = st.columns(2)
    meg_pf = c3.text_input("Prefixe Facture", value="FAC", key="meg_pf")
    meg_pa = c4.text_input("Prefixe Avoir", value="AVR", key="meg_pa")
    if f_meg_fac and st.button("Generer Gabarit Facture, Avoir & Paiement", key="btn_meg_fac"):
        df_cli = None
        if f_meg_cli2: df_cli = _read_meg(f_meg_cli2)
        elif os.path.exists(GABARIT_CLIENT_PATH): df_cli = pd.read_excel(GABARIT_CLIENT_PATH, header=0)
        if df_cli is None:
            st.error("Generez d'abord le Gabarit Client ou uploadez-le.")
        else:
            with st.spinner("Traitement en cours..."):
                df_f = _read_meg(f_meg_fac)
                st.dataframe(df_f.head(10), use_container_width=True)
                cl_lk = {}
                for _, r in df_cli.iterrows():
                    cd=to_clean_str(r.iloc[0]); nm=to_clean_str(r.iloc[2]) if len(r)>2 else ""; fm=to_clean_str(r.iloc[5]) if len(r)>5 else ""
                    lk=f"{fm} {nm}".strip()
                    if lk: cl_lk[lk]=cd
                    if nm: cl_lk[nm]=cd
                def _fcc(name):
                    cn=name.strip()
                    if cn in cl_lk: return cl_lk[cn]
                    for k,v in cl_lk.items():
                        if cn in k or k in cn: return v
                    return "Client non present"
                wb_f,ws_f=_make_wb(H_FACTURE); wb_a,ws_a=_make_wb(H_AVOIR); wb_p,ws_p=_make_wb(H_PAIEMENT)
                has_av=has_pa=tva_err=False
                fi={h.split(" *")[0]:i for i,h in enumerate(H_FACTURE)}
                ai={h.split(" *")[0]:i for i,h in enumerate(H_AVOIR)}
                pi={h.split(" *")[0]:i for i,h in enumerate(H_PAIEMENT)}
                ta=[meg_tva1,meg_tva2]
                def _gr(rate):
                    for t in ta:
                        if abs(rate-t)<=0.1: return t
                    return rate
                def _af(r,ht,tv=0):
                    o=[None]*len(H_FACTURE); o[fi["N facture externe"]]=to_clean_str(r.iloc[1]); o[fi["Date facture"]]=_parse_date(r.iloc[2])
                    cn=to_clean_str(r.iloc[3]); o[fi["Client"]]=cn; o[fi["Code client"]]=_fcc(cn)
                    o[fi["Commentaires"]]=to_clean_str(r.iloc[4]); o[fi["Conditions de reglement"]]="A reception"
                    o[fi["Designation"]]="NC"; o[fi["Qte"]]=1; o[fi["PU HT"]]=round(float(ht),2); o[fi["TVA"]]=tv
                    ws_f.append(o); return ws_f.max_row
                def _aa(r,ht,tv=0):
                    has_av=True
                    o=[None]*len(H_AVOIR); o[ai["N avoir externe"]]=to_clean_str(r.iloc[1]); o[ai["Date avoir"]]=_parse_date(r.iloc[2])
                    cn=to_clean_str(r.iloc[3]); o[ai["Client"]]=cn; o[ai["Code client"]]=_fcc(cn)
                    o[ai["Commentaires"]]=to_clean_str(r.iloc[4]); o[ai["Conditions de reglement"]]="A reception"
                    o[ai["Designation"]]="NC"; o[ai["Qte"]]=1; o[ai["PU HT"]]=round(float(ht),2); o[ai["TVA"]]=tv
                    ws_a.append(o); return ws_a.max_row
                def _ap(r,paid):
                    has_pa=True
                    o=[None]*len(H_PAIEMENT); o[pi["Facture n"]]=to_clean_str(r.iloc[1]); o[pi["Date paiement"]]=_parse_date(r.iloc[2])
                    o[pi["Libelle"]]="Reglement client"; o[pi["Mode de paiement"]]="Autres"; o[pi["Montant"]]=round(float(paid),2)
                    ws_p.append(o)
                for _,row in df_f.iterrows():
                    ht=_safe_float(row.iloc[5]); ttc=_safe_float(row.iloc[6]); paid=_safe_float(row.iloc[7])
                    doc=to_clean_str(row.iloc[1]); stat=to_clean_str(row.iloc[0])
                    if ht==0: continue
                    is_f=doc.startswith(meg_pf); is_a=doc.startswith(meg_pa)
                    if ttc-ht==0:
                        if is_f: _af(row,ht)
                        if is_a: _aa(row,ht)
                    else:
                        tva_am=ttc-ht; tva_r=round(tva_am/ht*100,2); tva_r=_gr(tva_r)
                        if tva_r==meg_tva1 or tva_r==meg_tva2:
                            if is_f: _af(row,ht,tva_r)
                            if is_a: _aa(row,ht,tva_r)
                        else:
                            if meg_tva2/100-meg_tva1/100!=0: ht2=round((ttc-ht*(1+meg_tva1/100))/(meg_tva2/100-meg_tva1/100),2)
                            else: ht2=0
                            ht1=ht-ht2; tc=round(ht1*meg_tva1/100+ht2*meg_tva2/100,2)
                            for hv,tv in [(ht1,meg_tva1),(ht2,meg_tva2)]:
                                if is_f:
                                    lr=_af(row,hv,tv)
                                    if tva_am!=tc: ws_f.cell(row=lr,column=fi["TVA"]+1).value="TVA a verifier"; tva_err=True
                                if is_a:
                                    lr=_aa(row,hv,tv)
                                    if tva_am!=tc: ws_a.cell(row=lr,column=ai["TVA"]+1).value="TVA a verifier"; tva_err=True
                    ttc9=_safe_float(row.iloc[8]) if len(row)>8 else ttc
                    if is_f and stat!="Annulee" and ttc9!=ttc and paid!=0: _ap(row,paid)
                if tva_err: st.warning("Des erreurs de TVA detectees. Verifiez la colonne TVA.")
                zb = io.BytesIO()
                with zipfile.ZipFile(zb,"w",zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr("3. Gabarit Facture.xlsx",_wb_bytes(wb_f))
                    if has_pa: zf.writestr("4. Gabarit Paiement.xlsx",_wb_bytes(wb_p))
                    if has_av: zf.writestr("5. Gabarit Avoir.xlsx",_wb_bytes(wb_a))
                zb.seek(0)
                st.success("Gabarits generes")
                st.download_button("Telecharger le ZIP",data=zb.getvalue(),file_name="Gabarit_Facture_Paiement_Avoir.zip",mime="application/zip",key="dl_meg_fac")

# --- Onglet Bascule Articles ---
if _connected and mod_articles:
 with m_art:
    if not (st.session_state.get('company_id_105') and st.session_state.get('token_headers_105')):
        st.warning("⛔ Connectez-vous a l'API et selectionnez un dossier (onglet **🔑 Connexion API**) avant d'utiliser cet onglet.")
    st.subheader("📦 Articles — Injection API Evoliz")
    f_meg_art = st.session_state.get("imp_file_articles")
    _pdfs_art_imp = st.session_state.get("imp_file_articles_pdfs", [])
    if not f_meg_art and not _pdfs_art_imp:
        st.info("Importez un fichier Excel et/ou des factures PDF dans l'onglet 📁 Import fichiers.")
    if True:
        has_h = bool(st.session_state.token_headers_105)
        if not has_h:
            st.warning("Connectez-vous d'abord a l'API Evoliz (onglet Balance & Cles API)")
        elif f_meg_art or _pdfs_art_imp:
            cid = st.session_state.company_id_105
            headers = st.session_state.token_headers_105
            if not cid:
                st.error("companyid non disponible. Reconnectez-vous.")
            else:
                # L'Excel est optionnel : on peut partir d'une base vide si seuls les PDFs sont importes
                if f_meg_art:
                    df_art = _read_meg(f_meg_art)
                    st.caption(f"Fichier Excel lu : **{len(df_art)} lignes**, **{len(df_art.columns)} colonnes**")
                    st.dataframe(df_art.head(5), use_container_width=True, hide_index=True)
                    _art_file_id = f_meg_art.name + str(len(df_art))
                else:
                    df_art = pd.DataFrame()
                    _art_file_id = f"pdf_only_{len(_pdfs_art_imp)}"
                    st.caption(f"📸 **{len(_pdfs_art_imp)} PDF** importe(s) — pas d'Excel. Les lignes seront extraites via l'outil PDF.")

                # Placeholder pour l'extraction PDF qui sera rendue AVANT la consolidation
                # (le code de l'extraction se trouve plus bas mais affiche dans ce conteneur)
                _pdf_container = st.container()

                _art_already = st.session_state.get("_art_consol_id") == _art_file_id
                _art_auto = not _art_already
                _art_manual = st.button("🔄 Etape 2 — Consolider avec Evoliz", use_container_width=True, key="btn_consol_art")

                if _art_auto or _art_manual:
                    st.session_state["_art_consol_id"] = _art_file_id
                    with st.spinner("Lecture articles et classifications Evoliz..."):
                        url_art = f"https://www.evoliz.io/api/v1/companies/{cid}/articles"
                        ev_articles = {}; page = 1
                        while True:
                            r = requests.get(url_art, headers=headers, params={"per_page": 100, "page": page}, timeout=15)
                            if r.status_code != 200: break
                            d = r.json()
                            for it in d.get("data", []):
                                ref = (it.get("reference_clean") or it.get("reference") or "").strip().upper()
                                if ref:
                                    # On stocke la reponse API complete pour afficher tous les champs disponibles
                                    _raw = dict(it)
                                    _raw["_id"] = it.get("articleid")
                                    ev_articles[ref] = _raw
                            if page >= d.get("meta", {}).get("last_page", 1): break
                            page += 1

                        sale_cl = {}; sale_cl_by_id = {}; url_sc = f"https://www.evoliz.io/api/v1/companies/{cid}/sale-classifications"; page = 1
                        while True:
                            r = requests.get(url_sc, headers=headers, params={"per_page": 100, "page": page}, timeout=15)
                            if r.status_code != 200: break
                            d = r.json()
                            for it in d.get("data", []):
                                c = str(it.get("code", "")).strip().upper()
                                sid = it.get("classificationid") or it.get("id")
                                lbl = it.get("label", "")
                                if c and sid:
                                    sale_cl[c] = sid
                                    sale_cl_by_id[sid] = c
                            if page >= d.get("meta", {}).get("last_page", 1): break
                            page += 1

                    # Consolidation : on fusionne les champs du fichier avec ceux de l'API
                    # Mapping des champs API -> colonnes affichees (snake_case -> libelles lisibles)
                    _field_map_api_to_display = {
                        "reference": "Reference",
                        "designation": "Designation",
                        "unit_price": "PU HT",
                        "vat_rate": "TVA %",
                        # Les autres champs API seront ajoutes dynamiquement (voir plus bas)
                    }
                    # Champs API usuels a exposer (titres lisibles). Les autres champs API seront ajoutes dynamiquement.
                    _api_extra_fields = {
                        "brand": "Marque",
                        "description": "Description",
                        "note": "Note",
                        "ean": "EAN / Code-barres",
                        "unit": "Unite",
                        "purchase_price": "Prix d'achat",
                        "min_sale_price": "Prix mini",
                        "supplier_reference": "Ref. fournisseur",
                        "weight": "Poids",
                        "height": "Hauteur", "width": "Largeur", "depth": "Profondeur",
                        "stock_management": "Gestion stock",
                    }

                    # Preserver les entrees provenant des factures PDF (extraites avant consolidation)
                    _prev_consol = st.session_state.get("_art_consol", [])
                    _pdf_entries = [dict(a) for a in _prev_consol if str(a.get("_source", "")).startswith("📸")]

                    consol_art = []; seen_art_ids = set()
                    # 1) D'abord les lignes PDF (au debut de la liste)
                    for _pdf_e in _pdf_entries:
                        _rk_pdf = str(_pdf_e.get("Reference", "")).upper()
                        if _rk_pdf in ev_articles and _rk_pdf != "NOREF":
                            _ev = ev_articles[_rk_pdf]
                            seen_art_ids.add(_ev["_id"])
                            _pdf_e["_source"] = "📸+☁️ Doublon"
                            _pdf_e["_entityid"] = _ev["_id"]
                            for _api_f, _disp_f in _api_extra_fields.items():
                                _v = _ev.get(_api_f)
                                if _v not in (None, "") and not _pdf_e.get(_disp_f):
                                    _pdf_e[_disp_f] = _v
                        consol_art.append(_pdf_e)

                    # 2) Puis les lignes Excel (en evitant les doublons avec les PDF deja ajoutes)
                    _pdf_refs = {str(p.get("Reference", "")).upper() for p in _pdf_entries}
                    for _, row in df_art.iterrows():
                        ref = to_clean_str(row.iloc[0]); des = to_clean_str(row.iloc[1]) if len(row) > 1 else ""
                        if not ref and not des: continue
                        if ref.upper() in _pdf_refs: continue  # deja present via PDF
                        pu = _safe_float(row.iloc[6]) if len(row) > 6 else None
                        tv = _safe_float(row.iloc[7]) if len(row) > 7 else None
                        cc = to_clean_str(row.iloc[4]).upper() if len(row) > 4 else ""
                        entry = {
                            "Reference": ref or "NOREF",
                            "Designation": des or ref,
                            "PU HT": round(pu, 2) if pu else "",
                            "TVA %": round(tv, 2) if tv else "",
                            "Classification vente": cc,
                            "_source": "📄 Nouveau",
                            "_entityid": None,
                        }
                        # Initialiser les champs extra vides
                        for _api_f, _disp_f in _api_extra_fields.items():
                            entry[_disp_f] = ""
                        ref_key = entry["Reference"].upper()
                        if ref_key in ev_articles:
                            ev = ev_articles[ref_key]
                            seen_art_ids.add(ev["_id"])
                            entry["_source"] = "📄+☁️ Doublon"
                            entry["_entityid"] = ev["_id"]
                            # Completer les champs extra depuis l'API
                            for _api_f, _disp_f in _api_extra_fields.items():
                                _v = ev.get(_api_f)
                                if _v not in (None, ""):
                                    entry[_disp_f] = _v
                        consol_art.append(entry)

                    for ref_key, ev in ev_articles.items():
                        if ev["_id"] not in seen_art_ids:
                            _e = {
                                "Reference": ev.get("reference", ""),
                                "Designation": ev.get("designation", ""),
                                "PU HT": ev.get("unit_price", ""),
                                "TVA %": ev.get("vat_rate", ""),
                                "Classification vente": sale_cl_by_id.get(ev.get("sale_classificationid"), ""),
                                "_source": "☁️ Evoliz seul",
                                "_entityid": ev["_id"],
                            }
                            for _api_f, _disp_f in _api_extra_fields.items():
                                _v = ev.get(_api_f)
                                _e[_disp_f] = _v if _v not in (None,) else ""
                            consol_art.append(_e)

                    st.session_state["_art_consol"] = consol_art
                    st.session_state["_art_sale_cl"] = sale_cl
                    st.session_state["_art_consol_stats"] = {
                        "fichier": sum(1 for c in consol_art if "Nouveau" in c["_source"]),
                        "doublons": sum(1 for c in consol_art if "Doublon" in c["_source"]),
                        "evoliz_seul": sum(1 for c in consol_art if "Evoliz seul" in c["_source"]),
                        "total_evoliz": len(ev_articles),
                    }
                    st.rerun()

                # Affichage consolidation
                consol_art = st.session_state.get("_art_consol", [])
                stats_a = st.session_state.get("_art_consol_stats")
                sale_cl = st.session_state.get("_art_sale_cl", {})
                if stats_a:
                    st.divider()
                    st.subheader("📊 Consolidation")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("📄 Nouveaux", stats_a["fichier"])
                    c2.metric("📄+☁️ Doublons", stats_a["doublons"])
                    c3.metric("☁️ Evoliz seul", stats_a["evoliz_seul"])
                    c4.metric("☁️ Total Evoliz", stats_a["total_evoliz"])

                # --- Extraction PDF : rendue dans le conteneur place AVANT la consolidation ---
                with _pdf_container:
                    _pdfs = st.session_state.get("imp_file_articles_pdfs", [])
                    _n_pdfs = len(_pdfs) if _pdfs else 0
                    _pdf_already_added = sum(1 for a in consol_art if a.get("_source") == "📸 Facture")
                    with st.expander(f"📸 Etape 1 — Extraire les lignes depuis {_n_pdfs} facture(s) PDF ({_pdf_already_added} deja ajoutee(s))", expanded=(_n_pdfs > 0 and _pdf_already_added == 0)):
                        if not _HAS_PDFPLUMBER:
                            st.warning("Installer `pdfplumber` pour activer cette fonctionnalite : `pip install pdfplumber`")
                        elif _n_pdfs == 0:
                            st.info("Aucun PDF importe. Ajoutez des factures PDF dans l'onglet **📁 Import fichiers** (ligne Factures PDF).")
                        else:
                            st.caption(f"📦 **{_n_pdfs} facture(s)** prete(s) pour extraction → mapping manuel des colonnes.")

                            c_ex1, c_ex2 = st.columns(2)
                            _extract_mode = c_ex1.radio("Mode d'extraction", ["Lignes de texte brut", "Tables structurees"], horizontal=False, key="pdf_extract_mode")
                            _show_raw = c_ex2.checkbox("Afficher le texte brut extrait (debug)", value=False, key="pdf_show_raw")

                            if st.button("🔍 Extraire", key="btn_extract_pdf_art", type="primary"):
                                # --- Extraction parallele (jusqu'a 100 PDFs) ---
                                import io as _io_pdf
                                def _extract_single_pdf(pdf_file, mode, show_raw):
                                    """Extraction d'un seul PDF. Retourne (rows, raw_pages, err)."""
                                    _rows = []; _raw_pages = []; _err = None
                                    try:
                                        pdf_file.seek(0)
                                        _bytes = pdf_file.read()
                                        with pdfplumber.open(_io_pdf.BytesIO(_bytes)) as pdf:
                                            for _pg_num, page in enumerate(pdf.pages):
                                                if mode == "Tables structurees":
                                                    _tables = page.extract_tables() or []
                                                    for _t_idx, _tbl in enumerate(_tables):
                                                        if not _tbl: continue
                                                        _max_w = max(len(r) for r in _tbl) if _tbl else 0
                                                        for _row in _tbl:
                                                            _row_p = list(_row) + [None] * (_max_w - len(_row))
                                                            _clean = [str(c).strip() if c else "" for c in _row_p]
                                                            if not any(_clean): continue
                                                            _entry = {f"Col {_k+1}": _clean[_k] for _k in range(_max_w)}
                                                            _entry["_pdf"] = pdf_file.name
                                                            _entry["_page"] = _pg_num + 1
                                                            _entry["_tbl"] = _t_idx + 1
                                                            _rows.append(_entry)
                                                else:
                                                    # Mode "Lignes de texte brut" : on utilise les coordonnees
                                                    # des mots pour detecter les colonnes reelles (pas de split par espaces).
                                                    # Les mots proches horizontalement sont groupes dans la meme colonne.
                                                    _words = page.extract_words(extra_attrs=["fontname", "size"]) or []
                                                    if not _words:
                                                        continue
                                                    # Groupement par ligne (coordonnee y similaire, tolerance 3pt)
                                                    _lines_by_y = []
                                                    _words_sorted = sorted(_words, key=lambda w: (w["top"], w["x0"]))
                                                    for _w in _words_sorted:
                                                        _placed = False
                                                        for _grp in _lines_by_y:
                                                            if abs(_grp[0]["top"] - _w["top"]) <= 3:
                                                                _grp.append(_w); _placed = True; break
                                                        if not _placed:
                                                            _lines_by_y.append([_w])
                                                    # Pour chaque ligne, clusteriser les mots en colonnes via x-gap
                                                    for _grp in _lines_by_y:
                                                        _grp.sort(key=lambda w: w["x0"])
                                                        # Seuil de gap : derive de la mediane des gaps entre mots
                                                        _gaps = [_grp[_i]["x0"] - _grp[_i-1]["x1"] for _i in range(1, len(_grp))]
                                                        if _gaps:
                                                            _sorted_gaps = sorted(_gaps)
                                                            _median_gap = _sorted_gaps[len(_sorted_gaps)//2]
                                                            _gap_threshold = max(5, _median_gap * 2.5)
                                                        else:
                                                            _gap_threshold = 5
                                                        _cols_words = [[_grp[0]]]
                                                        for _w in _grp[1:]:
                                                            _gap = _w["x0"] - _cols_words[-1][-1]["x1"]
                                                            if _gap > _gap_threshold:
                                                                _cols_words.append([_w])
                                                            else:
                                                                _cols_words[-1].append(_w)
                                                        _cols = [" ".join(_ww["text"] for _ww in _col_w).strip() for _col_w in _cols_words]
                                                        if not any(_cols): continue
                                                        _entry = {f"Col {_k+1}": _cols[_k] for _k in range(len(_cols))}
                                                        _entry["_pdf"] = pdf_file.name
                                                        _entry["_page"] = _pg_num + 1
                                                        _entry["_ligne"] = " | ".join(_cols)[:200]
                                                        _rows.append(_entry)
                                                if show_raw:
                                                    _raw_pages.append(page.extract_text() or "")
                                    except Exception as _exc:
                                        _err = str(_exc)
                                    return (pdf_file.name, _rows, _raw_pages, _err)

                                _all_rows = []
                                _raw_texts = {}
                                _errs_pdf = []
                                _total_pdf = len(_pdfs)
                                _pg_ocr = st.progress(0.0, text=f"Extraction — 0 / {_total_pdf} (8 workers)")
                                # Jusqu'a 8 workers en parallele (threads : pdfplumber est IO-bound)
                                with ThreadPoolExecutor(max_workers=min(8, _total_pdf)) as _pool:
                                    _futures = {_pool.submit(_extract_single_pdf, p, _extract_mode, _show_raw): p for p in _pdfs}
                                    for _n, _fut in enumerate(as_completed(_futures)):
                                        _name, _rows, _raw_pages, _err = _fut.result()
                                        if _err:
                                            _errs_pdf.append(f"{_name}: {_err}")
                                        _all_rows.extend(_rows)
                                        if _show_raw:
                                            _raw_texts[_name] = _raw_pages
                                        _pg_ocr.progress((_n + 1) / _total_pdf, text=f"Extraction — {_n+1} / {_total_pdf}")
                                _pg_ocr.empty()
                                if _errs_pdf:
                                    with st.expander(f"⚠️ {len(_errs_pdf)} PDF en erreur", expanded=False):
                                        for e in _errs_pdf: st.text(e)
                                st.session_state["_art_pdf_rows"] = _all_rows
                                st.session_state["_art_pdf_raw"] = _raw_texts if _show_raw else {}
                                st.rerun()

                            # Affichage texte brut si demande
                            _raw = st.session_state.get("_art_pdf_raw", {})
                            if _raw:
                                for _fn, _pgs in _raw.items():
                                    with st.expander(f"📄 Texte brut : {_fn}", expanded=False):
                                        for _i, _t in enumerate(_pgs):
                                            st.text_area(f"Page {_i+1}", value=_t, height=200, key=f"_raw_txt_{_fn}_{_i}")

                            _all_rows = st.session_state.get("_art_pdf_rows", [])
                            if _all_rows:
                                # Normaliser les cles pour creer un DataFrame homogene
                                _all_keys = []
                                for _r in _all_rows:
                                    for _k in _r.keys():
                                        if _k not in _all_keys:
                                            _all_keys.append(_k)
                                _df_rows = pd.DataFrame(_all_rows, columns=_all_keys)
                                _data_cols = [c for c in _all_keys if not c.startswith("_")]

                                st.success(f"✅ {len(_df_rows)} ligne(s) brute(s) extraite(s) sur {len({r['_pdf'] for r in _all_rows})} PDF")

                                # Filtre : ignorer les lignes vides de la plupart des colonnes (en-tetes, separations)
                                _only_data = st.checkbox("Filtrer les lignes probablement hors-table (ne garder que celles avec 2+ colonnes non vides)", value=True, key="pdf_only_data")
                                if _only_data:
                                    _mask = _df_rows[_data_cols].apply(lambda r: (r.astype(str).str.strip() != "").sum() >= 2, axis=1)
                                    _df_rows_show = _df_rows[_mask].reset_index(drop=True)
                                else:
                                    _df_rows_show = _df_rows.reset_index(drop=True)

                                # Afficher le dump brut dans un expander (debug)
                                with st.expander(f"🔎 Voir les {len(_df_rows_show)} ligne(s) brutes extraites", expanded=False):
                                    # Masquer les colonnes techniques volumineuses (_pdf, _page, _tbl, _ligne)
                                    _cols_hide = [c for c in _df_rows_show.columns if c.startswith("_")]
                                    st.dataframe(_df_rows_show.drop(columns=_cols_hide).head(100), use_container_width=True, hide_index=True)

                                # Helper de conversion numerique (defini AVANT l'auto-detection)
                                def _to_num(v):
                                    if v is None or str(v).strip() == "": return None
                                    s = str(v).replace(",", ".").replace("€", "").replace("\u00a0", "").replace(" ", "").strip()
                                    try: return round(float(s), 4)
                                    except ValueError: return None

                                # --- 🤖 Detection MULTI-BLOCS par headers Qte ---
                                # Un document peut contenir plusieurs blocs d'articles (plusieurs pages,
                                # plusieurs tableaux imbriques). Chaque bloc a son propre header Qte.
                                # Pour chaque bloc, les cellules du header row servent de noms de champs.
                                _qte_pat = re.compile(r"(?i)(^|\s|\(|\[|\/)(qte|qté|quant(?:ité|ite)?s?|qty|q\.)\b")
                                _total_re = re.compile(r"(?i)\b(total|sous[-\s]?total|net\s*a\s*payer|montant\s*(ht|ttc)|a\s*payer|tva\s*\d)\b")

                                # 1) Trouver TOUS les headers Qte (col, row) avec >= 1 valeur numerique en dessous
                                _headers_found = []  # liste de (row_idx, qte_col, header_cells_dict)
                                for _ri, _row in _df_rows_show.iterrows():
                                    for _c in _data_cols:
                                        _cell = str(_row.get(_c, "")).strip()
                                        if not _cell: continue
                                        if _qte_pat.search(_cell) and len(_cell) < 40:
                                            _below = _df_rows_show.iloc[_ri+1:_ri+20][_c].astype(str).str.strip().tolist()
                                            _num_count = sum(1 for v in _below if _to_num(v) is not None and _to_num(v) > 0)
                                            if _num_count >= 1:
                                                # Capturer les headers de la ligne (pour cette vue-bloc)
                                                _header_cells = {_hc: str(_row.get(_hc, "")).strip() for _hc in _data_cols}
                                                _headers_found.append((_ri, _c, _header_cells))
                                                break  # un seul Qte-header par ligne

                                # 2) Pour chaque header trouve, extraire les lignes d'articles jusqu'a rupture
                                _all_articles = []  # liste de dicts {header_label -> value, "_block": idx, "_pdf": name}
                                _blocks_info = []   # pour restitution
                                for _blk_idx, (_hdr_ri, _qte_c, _hdr_cells) in enumerate(_headers_found):
                                    _start = _hdr_ri + 1
                                    # Ne pas repasser sur une ligne deja dans un autre bloc
                                    _next_hdr_ri = _headers_found[_blk_idx + 1][0] if _blk_idx + 1 < len(_headers_found) else len(_df_rows_show)
                                    _end = _next_hdr_ri
                                    _block_articles = []
                                    for _ri in range(_start, _end):
                                        _row = _df_rows_show.iloc[_ri]
                                        _qte_val = _to_num(_row.get(_qte_c, ""))
                                        if _qte_val is None or _qte_val <= 0:
                                            _row_txt = " ".join(str(_row.get(_c, "")).strip() for _c in _data_cols)
                                            # Ligne vide = tolerer
                                            if not _row_txt.strip(): continue
                                            # Total = arreter ce bloc
                                            if _total_re.search(_row_txt): break
                                            # Header Qte repete (page break) = continuer
                                            if _qte_pat.search(str(_row.get(_qte_c, "")).strip()): continue
                                            # Autre ligne = arreter ce bloc
                                            break
                                        # Ligne d'article : extraire les valeurs par colonne avec leur header
                                        _art = {"_block": _blk_idx + 1, "_pdf": _row.get("_pdf", ""), "_page": _row.get("_page", "")}
                                        for _c in _data_cols:
                                            _label = _hdr_cells.get(_c, "").strip() or _c  # fallback : nom de colonne brut
                                            _val = str(_row.get(_c, "")).strip()
                                            # Exclure la colonne Qte du tableau final
                                            if _c == _qte_c:
                                                _art["_qte"] = _qte_val
                                                continue
                                            if _label and _val:
                                                _art[_label] = _val
                                        _block_articles.append(_art)
                                    _all_articles.extend(_block_articles)
                                    _blocks_info.append({
                                        "Bloc": _blk_idx + 1,
                                        "Header ligne": _hdr_ri + 1,
                                        "Col Qte": _qte_c,
                                        "Libelle Qte": _hdr_cells.get(_qte_c, ""),
                                        "Articles extraits": len(_block_articles),
                                    })

                                # Etat compatibilite avec le code de mapping/insertion existant
                                _detected_qte_col = None
                                _detected_header_row = None
                                _detect_source = ""
                                if _headers_found:
                                    # Pour retro-compat avec le mapping manuel qui preselectionne
                                    _detected_qte_col = _headers_found[0][1]
                                    _detected_header_row = _headers_found[0][0]

                                if _blocks_info:
                                    st.info(f"🤖 **{len(_blocks_info)} bloc(s) d'articles** detecte(s) — {sum(b['Articles extraits'] for b in _blocks_info)} ligne(s) au total.")
                                    with st.expander("Details des blocs detectes", expanded=False):
                                        st.dataframe(pd.DataFrame(_blocks_info), use_container_width=True, hide_index=True)

                                # Tableau de restitution : reprendre tous les headers (sauf Qte) trouves dans les blocs
                                if _all_articles:
                                    # Union des cles (hors techniques)
                                    _all_headers = []
                                    for _a in _all_articles:
                                        for _k in _a.keys():
                                            if _k.startswith("_"): continue
                                            if _k not in _all_headers:
                                                _all_headers.append(_k)
                                    _df_articles = pd.DataFrame([
                                        {**{_h: _a.get(_h, "") for _h in _all_headers},
                                         "Bloc": _a.get("_block"),
                                         "Qte": _a.get("_qte", ""),
                                         "PDF": _a.get("_pdf", ""),
                                        }
                                        for _a in _all_articles
                                    ])
                                    st.markdown(f"**📦 {len(_df_articles)} ligne(s) d'article extraite(s)** (colonnes issues des headers de chaque bloc)")
                                    st.dataframe(_df_articles, use_container_width=True, hide_index=True)
                                    # Sauvegarder pour usage ulterieur (mapping + injection)
                                    st.session_state["_art_pdf_extracted_rows"] = _all_articles
                                    # Pour compat : remplacer _df_rows_show par les lignes d'article detectees
                                    _df_rows_show = _df_rows_show.iloc[[_h[0] for _h in _headers_found] + [_ri for _a in _all_articles for _ri in []]].reset_index(drop=True) if False else _df_rows_show

                                # --- Mapping : choisir quels headers extraits deviennent quoi ---
                                # Utilise les noms de champs reels (headers PDF) et non plus Col 1, Col 2...
                                _extracted = st.session_state.get("_art_pdf_extracted_rows", []) or _all_articles
                                if _extracted:
                                    _field_options = ["— Non mappee —"]
                                    _union_hdrs = []
                                    for _a in _extracted:
                                        for _k in _a.keys():
                                            if _k.startswith("_"): continue
                                            if _k not in _union_hdrs:
                                                _union_hdrs.append(_k)
                                    _field_options += _union_hdrs

                                    # Heuristique d'auto-selection
                                    def _auto_pick(keywords):
                                        for _h in _union_hdrs:
                                            _hn = _h.lower()
                                            for kw in keywords:
                                                if kw in _hn:
                                                    return _h
                                        return "— Non mappee —"
                                    _auto_ref = _auto_pick(["ref", "code", "article"])
                                    _auto_des = _auto_pick(["design", "libell", "produit", "descr"])
                                    _auto_pu = _auto_pick(["pu", "prix", "unitaire", "unit"])
                                    _auto_tva = _auto_pick(["tva", "vat", "taxe"])

                                    st.markdown("**🗺️ Mapping des headers PDF → champs article**")
                                    m1, m2, m3, m4, m5 = st.columns(5)
                                    _map_ref = m1.selectbox("Reference", _field_options,
                                                             index=_field_options.index(_auto_ref), key="pdf_map_ref")
                                    _map_des = m2.selectbox("Designation", _field_options,
                                                             index=_field_options.index(_auto_des), key="pdf_map_des")
                                    _map_pu = m3.selectbox("PU HT", _field_options,
                                                            index=_field_options.index(_auto_pu), key="pdf_map_pu")
                                    _map_tva = m4.selectbox("TVA %", _field_options,
                                                             index=_field_options.index(_auto_tva), key="pdf_map_tva")
                                    _map_cl = m5.selectbox("Classif. vente", _field_options, key="pdf_map_cl")

                                    c_add1, c_add2 = st.columns(2)
                                    if c_add1.button(f"➕ Ajouter les {len(_extracted)} ligne(s) a la consolidation", type="primary", key="btn_add_pdf_to_consol"):
                                        _existing_refs = {str(a.get("Reference", "")).upper(): i for i, a in enumerate(consol_art)}
                                        _existing_des = {str(a.get("Designation", "")).strip().upper(): i for i, a in enumerate(consol_art) if a.get("Designation")}
                                        # Dedup local sur les lignes extraites (par Ref puis Designation)
                                        _seen_refs_local = set()
                                        _seen_des_local = set()
                                        _added = _merged = _skip_line = _skip_dup = 0
                                        for _a in _extracted:
                                            _ref = str(_a.get(_map_ref, "")).strip() if _map_ref != "— Non mappee —" else ""
                                            _des = str(_a.get(_map_des, "")).strip() if _map_des != "— Non mappee —" else ""
                                            _pu_n = _to_num(_a.get(_map_pu, "")) if _map_pu != "— Non mappee —" else None
                                            _tva_n = _to_num(_a.get(_map_tva, "")) if _map_tva != "— Non mappee —" else None
                                            _cl_v = str(_a.get(_map_cl, "")).strip().upper() if _map_cl != "— Non mappee —" else ""

                                            if not _ref and not _des:
                                                _skip_line += 1; continue
                                            if re.search(r"(?i)\b(total|sous-total|net\s*a\s*payer|tva|montant\s*ht|ttc)\b", _des):
                                                _skip_line += 1; continue

                                            _rk = (_ref or "NOREF").upper()
                                            _dk = _des.strip().upper()

                                            # Dedup INTRA-extraction (deux PDFs / lignes avec meme Ref ou meme Designation)
                                            if _rk != "NOREF" and _rk in _seen_refs_local:
                                                _skip_dup += 1; continue
                                            if _dk and _dk in _seen_des_local:
                                                _skip_dup += 1; continue

                                            _entry = {
                                                "Reference": _ref or "NOREF",
                                                "Designation": _des or _ref,
                                                "PU HT": _pu_n if _pu_n is not None else "",
                                                "TVA %": _tva_n if _tva_n is not None else "",
                                                "Classification vente": _cl_v,
                                                "_source": "📸 Facture",
                                                "_entityid": None,
                                            }

                                            # Fusion avec une entree existante dans consol_art (match sur Ref OU Designation)
                                            _match_i = None
                                            if _rk != "NOREF" and _rk in _existing_refs:
                                                _match_i = _existing_refs[_rk]
                                            elif _dk and _dk in _existing_des:
                                                _match_i = _existing_des[_dk]

                                            if _match_i is not None:
                                                for _f in ["Designation", "PU HT", "TVA %", "Classification vente"]:
                                                    if not consol_art[_match_i].get(_f) and _entry.get(_f):
                                                        consol_art[_match_i][_f] = _entry[_f]
                                                _merged += 1
                                            else:
                                                consol_art.append(_entry)
                                                _added += 1

                                            if _rk != "NOREF": _seen_refs_local.add(_rk)
                                            if _dk: _seen_des_local.add(_dk)
                                        st.session_state["_art_consol"] = consol_art
                                        st.session_state["_art_pdf_rows"] = []
                                        st.session_state["_art_pdf_extracted_rows"] = []
                                        _msg = f"✅ {_added} ajoute(s), {_merged} fusionne(s), {_skip_line} ignore(s), {_skip_dup} doublon(s) deduplique(s)"
                                        st.success(_msg)
                                        st.rerun()

                                    if c_add2.button("🗑️ Vider l'extraction", key="btn_clear_pdf_extraction"):
                                        st.session_state["_art_pdf_rows"] = []
                                        st.session_state["_art_pdf_raw"] = {}
                                        st.session_state["_art_pdf_extracted_rows"] = []
                                        st.rerun()

                if consol_art:
                    # --- Creation de classifications de vente (queue + batch) ---
                    if "_art_pending_classifs" not in st.session_state:
                        st.session_state._art_pending_classifs = []  # liste de dicts {label, accountid, account_label}
                    _pending_cl = st.session_state._art_pending_classifs

                    with st.expander(f"➕ Creer de nouvelles classifications de vente ({len(_pending_cl)} en attente)", expanded=False):
                        st.caption("Les classifs sont d'abord **mises en attente**. Elles seront creees via API **au moment de l'injection des articles**, et les IDs obtenus seront utilises dans les payloads article.")

                        # Chargement des comptes si pas encore disponibles
                        _ev_accounts = st.session_state.get("ev_acc_105", {})
                        if not _ev_accounts:
                            if st.button("📖 Charger les comptes comptables depuis Evoliz", key="btn_load_acc_art"):
                                with st.spinner("Lecture des comptes Evoliz..."):
                                    st.session_state.ev_acc_105 = fetch_evoliz_data("accounts", headers, company_id=cid)
                                st.rerun()
                            st.info("Aucun compte comptable en session. Cliquez sur le bouton ci-dessus pour les charger.")
                        else:
                            _only_7xx = st.checkbox("Afficher uniquement les comptes 7xx (ventes)", value=True, key="art_only_7xx")
                            _accs = list(_ev_accounts.values())
                            _seen_acc_ids = set()
                            _accs_unique = []
                            for _a in _accs:
                                if _a['id'] in _seen_acc_ids: continue
                                _seen_acc_ids.add(_a['id'])
                                if _only_7xx and not str(_a.get('code', '')).startswith('7'): continue
                                _accs_unique.append(_a)
                            _accs_unique.sort(key=lambda x: str(x.get('code', '')))
                            _acc_options = [f"{a['code']} - {a['label']}" for a in _accs_unique]
                            _acc_map = {f"{a['code']} - {a['label']}": a['id'] for a in _accs_unique}

                            if not _acc_options:
                                st.warning("Aucun compte disponible avec le filtre courant.")
                            else:
                                st.caption("**Mode 1 — Ajout rapide (plusieurs classifs pour un meme compte)**")
                                _acc_sel = st.selectbox("Compte comptable a lier", _acc_options, key="new_cl_acc")
                                _labels_multi = st.text_area(
                                    "Libelles des classifications (un par ligne)",
                                    placeholder="BOULANGERIE\nPATISSERIE\nBOISSONS",
                                    key="new_cl_labels_multi",
                                    height=100,
                                )
                                if st.button("📥 Ajouter a la liste a creer", key="btn_add_cl_multi"):
                                    _lines = [ln.strip() for ln in _labels_multi.split("\n") if ln.strip()]
                                    if not _lines:
                                        st.warning("Saisissez au moins un libelle.")
                                    else:
                                        _acc_id = _acc_map[_acc_sel]
                                        _added = 0
                                        _existing_labels = {p["label"].upper() for p in _pending_cl}
                                        for _lbl in _lines:
                                            if _lbl.upper() in _existing_labels: continue
                                            if _lbl.upper() in sale_cl: continue  # deja cree dans Evoliz
                                            _pending_cl.append({"label": _lbl, "accountid": _acc_id, "account_label": _acc_sel})
                                            _added += 1
                                        st.session_state._art_pending_classifs = _pending_cl
                                        st.success(f"✅ {_added} classif(s) ajoute(s) a la liste.")
                                        st.rerun()

                                st.divider()
                                st.caption("**Mode 2 — Ajout detaille (compte distinct par classif)**")
                                _nrows = st.number_input("Nombre de classifs a ajouter", min_value=1, max_value=50, value=3, step=1, key="new_cl_nrows")
                                _to_add_list = []
                                for _i in range(int(_nrows)):
                                    c1, c2 = st.columns([2, 3])
                                    _lbl_i = c1.text_input(f"Libelle {_i+1}", key=f"new_cl_lbl_{_i}")
                                    _acc_i = c2.selectbox(f"Compte {_i+1}", [""] + _acc_options, key=f"new_cl_acc_{_i}")
                                    if _lbl_i and _acc_i:
                                        _to_add_list.append((_lbl_i, _acc_map[_acc_i], _acc_i))
                                if _to_add_list:
                                    if st.button(f"📥 Ajouter les {len(_to_add_list)} classif(s) a la liste", key="btn_add_cl_detail"):
                                        _added = 0
                                        _existing_labels = {p["label"].upper() for p in _pending_cl}
                                        for _lbl, _acc_id, _acc_lbl in _to_add_list:
                                            if _lbl.upper() in _existing_labels: continue
                                            if _lbl.upper() in sale_cl: continue
                                            _pending_cl.append({"label": _lbl, "accountid": _acc_id, "account_label": _acc_lbl})
                                            _added += 1
                                        st.session_state._art_pending_classifs = _pending_cl
                                        st.success(f"✅ {_added} classif(s) ajoute(s) a la liste.")
                                        st.rerun()

                        # --- Liste des classifs en attente avec suppression ---
                        if _pending_cl:
                            st.divider()
                            st.markdown(f"**📋 {len(_pending_cl)} classification(s) en attente de creation**")
                            _df_pending = pd.DataFrame([
                                {"Libelle": p["label"], "Compte comptable": p["account_label"]}
                                for p in _pending_cl
                            ])
                            st.dataframe(_df_pending, use_container_width=True, hide_index=True)
                            c_p1, c_p2 = st.columns(2)
                            if c_p1.button("🗑️ Vider la liste", key="btn_clear_pending_cl"):
                                st.session_state._art_pending_classifs = []
                                st.rerun()
                            if c_p2.button("🚀 Creer maintenant (sans attendre l'injection)", key="btn_create_now_pending_cl"):
                                _ok = _ko = 0; _errs = []
                                _pg = st.progress(0.0, text=f"Creation — 0 / {len(_pending_cl)}")
                                _remaining = []
                                for _i, _p in enumerate(_pending_cl):
                                    ok, resp = inject_flux("VENTE", _p["label"][:50], _p["label"], headers,
                                                            acc_id=_p["accountid"], company_id=cid)
                                    if ok:
                                        _ok += 1
                                        _new_code = _p["label"][:50].upper()
                                        _new_id = resp.get("classificationid") or resp.get("id") if isinstance(resp, dict) else None
                                        if _new_id:
                                            sale_cl[_new_code] = _new_id
                                    else:
                                        _ko += 1; _errs.append(f"{_p['label']}: {resp}")
                                        _remaining.append(_p)
                                    _pg.progress((_i + 1) / len(_pending_cl), text=f"Creation — {_i+1} / {len(_pending_cl)}")
                                st.session_state["_art_sale_cl"] = sale_cl
                                st.session_state._art_pending_classifs = _remaining
                                if _ok: st.success(f"✅ {_ok} classification(s) creee(s)")
                                if _ko:
                                    st.error(f"❌ {_ko} erreur(s)")
                                    with st.expander("Detail erreurs"):
                                        for e in _errs: st.text(e)
                                st.rerun()

                    # --- Mapping rapide des classifications ---
                    # Options = classifs existantes + classifs en attente (prefixees ⏳)
                    _pending_labels = [f"⏳ {p['label'].upper()}" for p in st.session_state.get("_art_pending_classifs", [])]
                    _cl_options = [""] + sorted(sale_cl.keys()) + _pending_labels

                    # Mapping en masse : classifs inconnues -> reaffecter
                    _unknown_cl = sorted({str(a.get("Classification vente", "")).strip().upper()
                                           for a in consol_art
                                           if a.get("Classification vente") and str(a.get("Classification vente")).strip().upper() not in sale_cl})
                    if _unknown_cl:
                        with st.expander(f"🔀 Remapper les {len(_unknown_cl)} classification(s) inconnue(s) en masse", expanded=True):
                            st.caption("Choisissez la classification Evoliz correspondante. Tous les articles concernes seront mis a jour.")
                            _bulk_map = {}
                            for _uc in _unknown_cl:
                                _nb = sum(1 for a in consol_art if str(a.get("Classification vente", "")).strip().upper() == _uc)
                                _bulk_map[_uc] = st.selectbox(
                                    f"**{_uc}** ({_nb} article(s))",
                                    _cl_options,
                                    key=f"bulk_cl_{_uc}",
                                    format_func=lambda x: "— Laisser vide —" if not x else x,
                                )
                            if st.button("✅ Appliquer le remapping", key="btn_apply_bulk_cl"):
                                for i, a in enumerate(consol_art):
                                    _cur = str(a.get("Classification vente", "")).strip().upper()
                                    if _cur in _bulk_map and _bulk_map[_cur]:
                                        consol_art[i]["Classification vente"] = _bulk_map[_cur]
                                st.session_state["_art_consol"] = consol_art
                                st.rerun()

                    # --- Tableau consolide : tous champs editables + masquage colonnes ---
                    df_art_preview = pd.DataFrame([{k: v for k, v in c.items() if not k.startswith("_")} for c in consol_art])
                    _sources_a = [c["_source"] for c in consol_art]
                    df_art_preview.insert(0, "Source", _sources_a)

                    _all_art_cols = [c for c in df_art_preview.columns if c != "Source"]
                    _hidden_default = st.session_state.get("_art_hidden_cols", [])
                    _hidden_cols = st.multiselect(
                        "🙈 Colonnes a masquer (non-envoyees a Evoliz lors de l'injection)",
                        options=_all_art_cols,
                        default=[c for c in _hidden_default if c in _all_art_cols],
                        key="_art_hidden_cols",
                        help="Cocher les attributs a ignorer. Les colonnes masquees ne seront pas envoyees a l'API.",
                    )
                    _visible_cols = ["Source"] + [c for c in _all_art_cols if c not in _hidden_cols]
                    df_art_display = df_art_preview[_visible_cols]

                    _edited_art = st.data_editor(
                        df_art_display,
                        use_container_width=True, hide_index=True,
                        disabled=["Source"],
                        column_config={
                            "Classification vente": st.column_config.SelectboxColumn(
                                "Classification vente",
                                help="Selectionnez une classification de vente existante dans Evoliz. Editable et saisissable.",
                                options=_cl_options,
                                required=False,
                            ),
                            "Prix d'achat": st.column_config.NumberColumn("Prix d'achat", step=0.01, format="%.2f"),
                            "Prix mini": st.column_config.NumberColumn("Prix mini", step=0.01, format="%.2f"),
                            "Poids": st.column_config.NumberColumn("Poids (kg)", step=0.01, format="%.2f"),
                            "Hauteur": st.column_config.NumberColumn("Hauteur", step=0.1, format="%.1f"),
                            "Largeur": st.column_config.NumberColumn("Largeur", step=0.1, format="%.1f"),
                            "Profondeur": st.column_config.NumberColumn("Profondeur", step=0.1, format="%.1f"),
                            "Gestion stock": st.column_config.CheckboxColumn("Gestion stock"),
                            "Marque": st.column_config.TextColumn("Marque"),
                            "Description": st.column_config.TextColumn("Description"),
                            "Note": st.column_config.TextColumn("Note"),
                            "EAN / Code-barres": st.column_config.TextColumn("EAN"),
                            "Unite": st.column_config.TextColumn("Unite", help="kg, l, piece, etc."),
                            "Ref. fournisseur": st.column_config.TextColumn("Ref. fournisseur"),
                            "Reference": st.column_config.TextColumn("Reference", help="Cle unique de l'article. Editable."),
                            "Designation": st.column_config.TextColumn("Designation", help="Editable. Copie-colle OK."),
                            "PU HT": st.column_config.NumberColumn("PU HT", help="Prix unitaire HT. Editable.", step=0.01, format="%.2f"),
                            "TVA %": st.column_config.NumberColumn("TVA %", help="Taux TVA. Editable.", step=0.1, format="%.2f"),
                        },
                        key="art_editor",
                    )
                    # Persister tous les changements de cellules editees (sur les colonnes visibles)
                    _art_changed = False
                    for _i, _row in _edited_art.iterrows():
                        for _col in _visible_cols:
                            if _col == "Source": continue
                            _new_v = _row.get(_col)
                            _old_v = consol_art[_i].get(_col)
                            # Normalisation selon type
                            _numeric_cols = ("PU HT", "TVA %", "Prix d'achat", "Prix mini", "Poids", "Hauteur", "Largeur", "Profondeur")
                            if _col in _numeric_cols:
                                try: _new_n = float(_new_v) if _new_v not in (None, "") and not (isinstance(_new_v, float) and pd.isna(_new_v)) else ""
                                except (TypeError, ValueError): _new_n = ""
                                try: _old_n = float(_old_v) if _old_v not in (None, "") else ""
                                except (TypeError, ValueError): _old_n = ""
                                if _new_n != _old_n:
                                    consol_art[_i][_col] = _new_n
                                    _art_changed = True
                            elif _col == "Gestion stock":
                                _new_b = bool(_new_v) if _new_v is not None else False
                                _old_b = bool(_old_v) if _old_v else False
                                if _new_b != _old_b:
                                    consol_art[_i][_col] = _new_b
                                    _art_changed = True
                            else:
                                _new_s = str(_new_v).strip() if _new_v is not None and not (isinstance(_new_v, float) and pd.isna(_new_v)) else ""
                                _old_s = str(_old_v).strip() if _old_v is not None and not (isinstance(_old_v, float) and pd.isna(_old_v)) else ""
                                if _col == "Classification vente":
                                    _new_s = _new_s.upper()
                                    _old_s = _old_s.upper()
                                if _new_s != _old_s:
                                    consol_art[_i][_col] = _new_s
                                    _art_changed = True
                    if _art_changed:
                        st.session_state["_art_consol"] = consol_art

                    # Verification classif manquantes (apres mapping)
                    _art_missing_cl = []
                    for i, a in enumerate(consol_art):
                        cc = str(a.get("Classification vente", "")).strip().upper()
                        if cc and cc not in sale_cl:
                            _art_missing_cl.append({"Reference": a.get("Reference"), "Classification": cc})
                    if _art_missing_cl:
                        st.warning(f"⚠️ {len(_art_missing_cl)} article(s) avec classification inconnue (ignoree a l'injection). Utilisez le remappage ci-dessus ou le selecteur dans la colonne.")

                    # Injection
                    st.divider()
                    st.subheader("🚀 Injection Articles dans Evoliz")
                    _pending_now = st.session_state.get("_art_pending_classifs", [])
                    _lbl_inject = "🚀 Injecter les articles"
                    if _pending_now:
                        _lbl_inject = f"🚀 Creer {len(_pending_now)} classif(s) + Injecter les articles"
                    inject_art_btn = st.button(_lbl_inject, type="primary", use_container_width=True, key="btn_inject_art")
                    if inject_art_btn:
                        # --- Etape 1 : Creer les classifications en attente ---
                        _pending_now = st.session_state.get("_art_pending_classifs", [])
                        if _pending_now:
                            st.markdown("### 🏷️ Etape 1 : Creation des classifications en attente")
                            _ok_cl = _ko_cl = 0; _errs_cl = []
                            _remaining_cl = []
                            _pg_cl = st.progress(0.0, text=f"Creation classifs — 0 / {len(_pending_now)}")
                            for _i, _p in enumerate(_pending_now):
                                ok, resp = inject_flux("VENTE", _p["label"][:50], _p["label"], headers,
                                                        acc_id=_p["accountid"], company_id=cid)
                                if ok:
                                    _ok_cl += 1
                                    _new_code = _p["label"][:50].upper()
                                    _new_id = resp.get("classificationid") or resp.get("id") if isinstance(resp, dict) else None
                                    if _new_id:
                                        sale_cl[_new_code] = _new_id
                                else:
                                    _ko_cl += 1; _errs_cl.append(f"{_p['label']}: {resp}")
                                    _remaining_cl.append(_p)
                                _pg_cl.progress((_i + 1) / len(_pending_now), text=f"Creation classifs — {_i+1} / {len(_pending_now)}")
                            st.session_state["_art_sale_cl"] = sale_cl
                            st.session_state._art_pending_classifs = _remaining_cl
                            if _ok_cl: st.success(f"✅ {_ok_cl} classification(s) creee(s)")
                            if _ko_cl:
                                st.error(f"❌ {_ko_cl} erreur(s) de creation classif (les articles concernes seront injectes sans classif)")
                                with st.expander("Detail erreurs"):
                                    for e in _errs_cl: st.text(e)
                            st.markdown("### 📦 Etape 2 : Injection des articles")
                        # URLs avec fallback (mono/multi indifferent)
                        _url_ap_pri = f"https://www.evoliz.io/api/v1/companies/{cid}/articles"
                        _url_ap_fb = "https://www.evoliz.io/api/v1/articles"
                        url_art = _url_ap_pri
                        def _http_art(method, suffix, payload):
                            for _u in [f"{_url_ap_pri}{suffix}", f"{_url_ap_fb}{suffix}"]:
                                try:
                                    if method == "POST":
                                        r = requests.post(_u, headers=headers, json=payload, timeout=15)
                                    else:
                                        r = requests.patch(_u, headers=headers, json=payload, timeout=15)
                                    if r.status_code in (200, 201, 204): return r
                                    if r.status_code not in (403, 404): return r
                                except Exception:
                                    pass
                            return r
                        created_a = updated_a = skipped_a = 0; errors_a = []; inject_log_a = []
                        # On n'envoie que les articles du fichier (Nouveaux + Doublons)
                        _to_inject = [a for a in consol_art if "Evoliz seul" not in a["_source"]]
                        # Colonnes masquees -> non envoyees
                        _hidden_now = set(st.session_state.get("_art_hidden_cols", []))
                        prog = st.progress(0.0, text=f"Injection articles — 0 / {len(_to_inject)}")
                        for i, a in enumerate(_to_inject):
                            payload = {"reference": a["Reference"], "designation": a["Designation"]}
                            if "PU HT" not in _hidden_now and a.get("PU HT") not in (None, ""):
                                try: payload["unit_price"] = round(float(a["PU HT"]), 2)
                                except (TypeError, ValueError): pass
                            if "TVA %" not in _hidden_now and a.get("TVA %") not in (None, ""):
                                try: payload["vat_rate"] = round(float(a["TVA %"]), 2)
                                except (TypeError, ValueError): pass
                            cc = str(a.get("Classification vente", "")).strip().upper()
                            # Nettoyer le prefixe ⏳ ajoute pour les classifs en attente
                            if cc.startswith("⏳"):
                                cc = cc.replace("⏳", "").strip()
                            if "Classification vente" not in _hidden_now and cc and cc in sale_cl:
                                payload["sale_classificationid"] = sale_cl[cc]
                            # Champs extra API (envoyes si non masques et non vides)
                            _extra_payload_map = {
                                "Marque": "brand", "Description": "description", "Note": "note",
                                "EAN / Code-barres": "ean", "Unite": "unit",
                                "Prix d'achat": "purchase_price", "Prix mini": "min_sale_price",
                                "Ref. fournisseur": "supplier_reference", "Poids": "weight",
                                "Hauteur": "height", "Largeur": "width", "Profondeur": "depth",
                                "Gestion stock": "stock_management",
                            }
                            for _disp_f, _api_f in _extra_payload_map.items():
                                if _disp_f in _hidden_now: continue
                                _v = a.get(_disp_f, "")
                                if _v in (None, ""): continue
                                # Champs numeriques
                                if _api_f in ("purchase_price", "min_sale_price", "weight", "height", "width", "depth"):
                                    try: payload[_api_f] = round(float(_v), 2)
                                    except (TypeError, ValueError): pass
                                elif _api_f == "stock_management":
                                    # Accepte booleens ou strings
                                    _vs = str(_v).strip().lower()
                                    payload[_api_f] = _vs in ("true", "1", "oui", "yes", "vrai")
                                else:
                                    payload[_api_f] = str(_v).strip()
                            entity_id = a.get("_entityid")
                            try:
                                if entity_id:
                                    r = _http_art("PATCH", f"/{entity_id}", payload)
                                    if r.status_code in (200, 204):
                                        updated_a += 1; inject_log_a.append({"Reference": a["Reference"], "Action": "🔄 MAJ", "Statut": "✅ OK", "Detail": ""})
                                    else:
                                        errors_a.append(f"MAJ '{a['Reference']}': HTTP {r.status_code}")
                                        inject_log_a.append({"Reference": a["Reference"], "Action": "🔄 MAJ", "Statut": f"❌ {r.status_code}", "Detail": r.text[:80]})
                                else:
                                    r = _http_art("POST", "", payload)
                                    if r.status_code in (200, 201):
                                        created_a += 1; inject_log_a.append({"Reference": a["Reference"], "Action": "➕ Creation", "Statut": "✅ OK", "Detail": ""})
                                    elif r.status_code == 400 and "already been taken" in r.text:
                                        updated_a += 1; inject_log_a.append({"Reference": a["Reference"], "Action": "🔄 Existant", "Statut": "✅ OK", "Detail": ""})
                                    else:
                                        errors_a.append(f"Creation '{a['Reference']}': HTTP {r.status_code}")
                                        inject_log_a.append({"Reference": a["Reference"], "Action": "➕ Creation", "Statut": f"❌ {r.status_code}", "Detail": r.text[:80]})
                            except Exception as e:
                                errors_a.append(f"'{a['Reference']}': {e}")
                                inject_log_a.append({"Reference": a["Reference"], "Action": "?", "Statut": "❌ Exception", "Detail": str(e)[:80]})
                            prog.progress((i + 1) / max(len(_to_inject), 1), text=f"Injection articles — {i+1} / {len(_to_inject)} — {a['Reference'][:30]}")
                        prog.empty()
                        st.divider(); st.subheader("📊 Synthese")
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("➕ Crees", created_a); c2.metric("🔄 MAJ", updated_a)
                        c3.metric("⏭️ Ignores", skipped_a); c4.metric("❌ Erreurs", len(errors_a))
                        if created_a + updated_a > 0:
                            st.success(f"Injection : {created_a} cree(s), {updated_a} mis a jour")
                        elif errors_a:
                            st.error(f"{len(errors_a)} erreur(s)")
                        if inject_log_a:
                            df_il_a = pd.DataFrame(inject_log_a)
                            def _ci_a(row):
                                s = str(row.get("Statut", "")); a = str(row.get("Action", ""))
                                if "OK" in s and "Creation" in a: return ["background-color:#d4edda"] * len(row)
                                if "OK" in s: return ["background-color:#e8f4fd"] * len(row)
                                if "❌" in s: return ["background-color:#f8d7da"] * len(row)
                                return [""] * len(row)
                            st.dataframe(df_il_a.style.apply(_ci_a, axis=1), use_container_width=True, hide_index=True)